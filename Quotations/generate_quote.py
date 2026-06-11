import os
import re
import sys
import unicodedata
from datetime import datetime
import pandas as pd
from typing import Iterable, Optional, Tuple
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
try:
    from rapidfuzz import process as rf_process, fuzz as rf_fuzz  # type: ignore
except Exception:
    rf_process = None  # type: ignore
    rf_fuzz = None  # type: ignore

# Support running as a script
try:
    from .qtool_loader import load_input_template  # type: ignore
    from .qtool_loader import STD_COLS  # type: ignore
    from .data_sources import (
        load_main_ports, load_transit_time, load_horse_puerto, load_cost_per_km,
        map_factory_to_port, find_port_by_country,
    )
    from .rules import flow_by_incoterm
    # Prefer local module name 'Distances' (Windows FS retains this casing)
    try:
        from .Distances import GeoIndex, resolve_point, road_km_between  # type: ignore
    except Exception:
        from .distances import GeoIndex, resolve_point, road_km_between  # type: ignore
    try:
        from .geo_online import geocode_city_online_if_allowed  # type: ignore
    except Exception:
        geocode_city_online_if_allowed = None  # type: ignore
except Exception:
    sys.path.append(os.path.dirname(os.path.dirname(__file__)))
    from Quotations.qtool_loader import load_input_template  # type: ignore
    from Quotations.qtool_loader import STD_COLS  # type: ignore
    from Quotations.data_sources import (
        load_main_ports, load_transit_time, load_horse_puerto, load_cost_per_km,
        map_factory_to_port, find_port_by_country,
    )
    from Quotations.rules import flow_by_incoterm  # type: ignore
    try:
        from Quotations.Distances import GeoIndex, resolve_point, road_km_between  # type: ignore
    except Exception:
        from Quotations.distances import GeoIndex, resolve_point, road_km_between  # type: ignore
    try:
        from Quotations.geo_online import geocode_city_online_if_allowed  # type: ignore
    except Exception:
        geocode_city_online_if_allowed = None  # type: ignore


QTOOL_DIR = r"C:\Users\OLMEDOJorge\OneDrive - Horse\Exchange VRAC\02_Engineering Department\08. New tools & technologies\QTool"
INPUT_FILE = os.path.join(QTOOL_DIR, "upload_Quotation Template.xlsx")
OUTPUT_PREFIX = "Horse_TPTQuotation"
OUTPUT_EXT = ".xlsx"
DEFAULT_INCOTERM = os.environ.get("QINCOTERM", "FCA")


def find_qtool_data_file() -> str | None:
    """Find the QUOTATION TOOL DATA Excel in QTOOL_DIR."""
    # 0) Prefer repository-bundled file (works in Streamlit Cloud / GitHub deploy)
    try:
        repo_local = os.path.join(os.path.dirname(__file__), "QUOTATION TOOL DATA.xlsx")
        if os.path.exists(repo_local):
            return repo_local
    except Exception:
        pass
    # 1) Prefer exact canonical filename if present
    try:
        exact = os.path.join(QTOOL_DIR, "QUOTATION TOOL DATA.xlsx")
        if os.path.exists(exact):
            return exact
    except Exception:
        pass
    # 2) Otherwise, pick the latest file that starts with that name
    candidates = []
    try:
        for name in os.listdir(QTOOL_DIR):
            low = name.lower()
            if low.startswith("quotation tool data") and low.endswith((".xlsx", ".xlsm", ".xls")):
                full = os.path.join(QTOOL_DIR, name)
                try:
                    mtime = os.path.getmtime(full)
                except Exception:
                    mtime = 0
                # Prefer the canonical file by giving it a boost in score
                is_canonical = 1 if low == "quotation tool data.xlsx" else 0
                candidates.append((is_canonical, mtime, full))
    except FileNotFoundError:
        return None
    if not candidates:
        return None
    # Sort by (is_canonical desc, mtime desc)
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return candidates[0][2]


def next_output_path(directory: str) -> str:
    date_tag = datetime.now().strftime("%Y%m%d")
    pattern_new = re.compile(rf"^{re.escape(OUTPUT_PREFIX)}_{date_tag}_(\d+){re.escape(OUTPUT_EXT)}$", re.IGNORECASE)
    # Backward compatibility: keep correlativo based on old naming if files already exist.
    pattern_old = re.compile(rf"^Horse_Quotation_{date_tag}_descarga_(\d+){re.escape(OUTPUT_EXT)}$", re.IGNORECASE)
    max_n = 0
    try:
        for name in os.listdir(directory):
            m_new = pattern_new.match(name)
            m_old = pattern_old.match(name)
            m = m_new or m_old
            if m is not None:
                try:
                    n = int(m.group(1))
                    if n > max_n:
                        max_n = n
                except Exception:
                    pass
    except FileNotFoundError:
        os.makedirs(directory, exist_ok=True)
    next_n = max_n + 1
    return os.path.join(directory, f"{OUTPUT_PREFIX}_{date_tag}_{next_n}{OUTPUT_EXT}")


def _find_reference_output_file() -> str | None:
    """Return a path to a reference 'download_quotation-output*.xlsx' in QTOOL_DIR, preferring non-numbered name."""
    exact = os.path.join(QTOOL_DIR, "download_quotation-output.xlsx")
    if os.path.exists(exact):
        return exact
    # Fallback to latest numbered
    try:
        candidates = []
        for name in os.listdir(QTOOL_DIR):
            if re.match(r"download_quotation-output_\d+\.xlsx", name, flags=re.I):
                full = os.path.join(QTOOL_DIR, name)
                try:
                    mtime = os.path.getmtime(full)
                except Exception:
                    mtime = 0
                candidates.append((mtime, full))
        if candidates:
            candidates.sort(reverse=True)
            return candidates[0][1]
    except FileNotFoundError:
        return None
    return None


def _load_reference_quote_columns() -> list[str] | None:
    ref = _find_reference_output_file()
    if not ref:
        return None
    try:
        # Read only header (no rows) from 'Quote' sheet
        df0 = pd.read_excel(ref, sheet_name="Quote", nrows=0)
        return list(df0.columns)
    except Exception:
        return None


def build_output(input_df: pd.DataFrame, out_path: str):
    # Load data sources
    data_file = find_qtool_data_file()
    if not data_file:
        raise FileNotFoundError("QUOTATION TOOL DATA file not found in QTool directory")
    try:
        df_mp = load_main_ports(data_file)
        try:
            df_tt = load_transit_time(data_file)
        except Exception:
            df_tt = pd.DataFrame()
        df_hp = load_horse_puerto(data_file)
        df_cpkm = load_cost_per_km(data_file)
        # Optional ports locations sheet (coordinates for POL/POD)
        try:
            df_ports = pd.read_excel(data_file, sheet_name="Ports Locations")
        except Exception:
            df_ports = pd.DataFrame()
        # Optional ZIP coordinates sheet (ZIP -> Lat/Lon)
        try:
            try:
                df_zip_coords = pd.read_excel(data_file, sheet_name="ZIP_COORDS")
            except Exception:
                df_zip_coords = pd.read_excel(data_file, sheet_name="ZIP_COORDINATES")
        except Exception:
            df_zip_coords = pd.DataFrame()
        # Optional city→ZIP mapping sheet (database-managed)
        try:
            df_city_zips = pd.read_excel(data_file, sheet_name="CITY_ZIPS")
        except Exception:
            df_city_zips = pd.DataFrame()
        # Optional city coordinates sheet (database-managed)
        try:
            # Accept either GEO_LOCATIONS or CITY_COORDS sheet name
            try:
                df_geo_cities = pd.read_excel(data_file, sheet_name="GEO_LOCATIONS")
            except Exception:
                df_geo_cities = pd.read_excel(data_file, sheet_name="CITY_COORDS")
        except Exception:
            df_geo_cities = pd.DataFrame()
        # Optional city alias mapping sheet (database-managed)
        try:
            df_city_aliases = pd.read_excel(data_file, sheet_name="CITY_ALIASES")
        except Exception:
            df_city_aliases = pd.DataFrame()
        # Packaging codes reference sheet
        try:
            df_packaging = pd.read_excel(data_file, sheet_name="PACKAGING")
        except Exception:
            df_packaging = pd.DataFrame()
        # Optional VTT table used by VTT2.py for POL/POD transit time lookups
        try:
            vtt_data_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "VTT Tool", "VTT DATA.xlsx")
            df_vtt_routes = pd.read_excel(vtt_data_path)
        except Exception:
            df_vtt_routes = pd.DataFrame()
    except PermissionError as e:
        raise PermissionError(f"No se pudo leer QUOTATION TOOL DATA (bloqueado/abierto): {data_file}") from e

    # Default packaging code when input is missing or not found in PACKAGING sheet
    DEFAULT_PACKAGING_CODE = "CAR-S*2466"

    def resolve_packaging_code(raw_code: str | None) -> str:
        """Return the validated packaging code from PACKAGING sheet, or DEFAULT_PACKAGING_CODE."""
        code = str(raw_code or "").strip()
        if not code or code.upper() in ("", "NAN", "NONE", "NULL", "-"):
            return DEFAULT_PACKAGING_CODE
        if df_packaging is None or df_packaging.empty:
            return code
        valid_codes = df_packaging["Packaging Code"].astype(str).str.strip()
        if code in valid_codes.values:
            return code
        # Case-insensitive match
        match = valid_codes[valid_codes.str.upper() == code.upper()]
        if not match.empty:
            return str(match.iloc[0])
        # Not found in database -> use default
        return DEFAULT_PACKAGING_CODE

    def lookup_packaging_data(pn: str, raw_pack_code: str | None, resolved_pack_code: str) -> dict:
        """Lookup PACKAGING data with priority:
        1) Reference (PN) for Part Weight
        2) Packaging Code from input
        3) Default packaging code when input code has no match
        """
        result = {
            "pkg_volume_m3": None,
            "pkg_snp": None,
            "pkg_length_mm": None,
            "pkg_width_mm": None,
            "pkg_height_mm": None,
            "pkg_weight_part": None,
            "pkg_weight_empty": None,
            "pkg_weight_full": None,
            "pkg_debug": "",
        }
        if df_packaging is None or df_packaging.empty:
            result["pkg_debug"] = "PACKAGING no disponible"
            return result
        try:
            pn_u = str(pn or "").strip().upper()
            raw_pc_u = str(raw_pack_code or "").strip().upper()
            res_pc_u = str(resolved_pack_code or "").strip().upper()
            ref_col = df_packaging["Reference"].astype(str).str.strip().str.upper()
            pcode_col = df_packaging["Packaging Code"].astype(str).str.strip().str.upper()

            row = None
            # 1) Priority for Weight/part and base packaging data: match by Reference (PN)
            m_ref = df_packaging[ref_col == pn_u]
            if not m_ref.empty:
                row = m_ref.iloc[0]
                result["pkg_debug"] = "Packaging: Weight/part por Reference (PN)"
            else:
                # 2) Try input packaging code
                if raw_pc_u and raw_pc_u not in ("NAN", "NONE", "NULL", "-"):
                    m_raw = df_packaging[pcode_col == raw_pc_u]
                    if not m_raw.empty:
                        row = m_raw.iloc[0]
                        result["pkg_debug"] = "Packaging: Weight/part por Packaging Code input"
                # 3) If input code does not match, use resolved/default packaging code
                if row is None and res_pc_u:
                    m_res = df_packaging[pcode_col == res_pc_u]
                    if not m_res.empty:
                        row = m_res.iloc[0]
                        if raw_pc_u and raw_pc_u != res_pc_u:
                            result["pkg_debug"] = "Packaging: sin match input; usando Packaging Code default"
                        else:
                            result["pkg_debug"] = "Packaging: Weight/part por Packaging Code"

            if row is None:
                result["pkg_debug"] = "Packaging: sin match en Reference ni Packaging Code"
                return result

            h = pd.to_numeric(row.get("Height (mm)"), errors="coerce")
            w = pd.to_numeric(row.get("Width (mm)"), errors="coerce")
            l = pd.to_numeric(row.get("Lenght (mm)"), errors="coerce")
            part_w = pd.to_numeric(row.get("Part Weight (kg)"), errors="coerce")
            empty_w = pd.to_numeric(row.get("Weight EMPTY (kg)"), errors="coerce")
            snp = pd.to_numeric(row.get("SNP / Pack (PN / Packaging)"), errors="coerce")
            if pd.notna(h) and pd.notna(w) and pd.notna(l):
                result["pkg_volume_m3"] = round(h * w * l / 1_000_000_000, 6)
                result["pkg_length_mm"] = float(l)
                result["pkg_width_mm"] = float(w)
                result["pkg_height_mm"] = float(h)
            result["pkg_snp"] = float(snp) if pd.notna(snp) else None
            result["pkg_weight_part"] = float(part_w) if pd.notna(part_w) else None
            result["pkg_weight_empty"] = float(empty_w) if pd.notna(empty_w) else None
            if pd.notna(empty_w) and pd.notna(part_w) and pd.notna(snp):
                result["pkg_weight_full"] = round(float(empty_w) + float(part_w) * float(snp), 3)
        except Exception:
            pass
        return result

    # Empaquetado por tipo de transporte (lógica basada en Empower3D, sin stackability manual)
    TRANSPORT_OPERATIVE_DIMS = {
        "OVERSEAS": (12032, 2352, 2550),      # Container 40 HC
        "INLAND": (13620, 2480, 2900),        # Mega Trailer 90m3
    }
    TRANSPORT_MAX_WEIGHT = {
        "OVERSEAS": 24750.0,                  # kg
        "INLAND": 25000.0,                    # kg
    }

    def _max_packs_by_volume(container_dim: tuple[int, int, int], box_dim: tuple[float, float, float]) -> int:
        """Best-fit quantity by volume using two planar rotations, unlimited stack up to container height."""
        Lc, Wc, Hc = container_dim
        l1, w1, h = box_dim
        if l1 <= 0 or w1 <= 0 or h <= 0:
            return 0
        l2, w2 = w1, l1
        nh = int(Hc // h)
        if nh <= 0:
            return 0

        # Option 1: main orientation (l1, w1)
        nl1 = int(Lc // l1)
        nw1 = int(Wc // w1)
        sw = Wc - (nw1 * w1)
        sl = Lc - (nl1 * l1)
        nl2 = nl1
        nw2 = int(sw // w2) if sw >= w2 else 0
        nl3 = int(sl // l2) if sl >= l2 else 0
        nw3 = nw1
        nl4 = int(sl // l2) if sl >= l2 else 0
        nw4 = int(sw // w2) if sw >= w2 else 0
        total1 = (nl1 * nw1 + nl2 * nw2 + nl3 * nw3 + nl4 * nw4) * nh

        # Option 2: swapped main orientation (w1, l1)
        nl1b = int(Lc // w1)
        nw1b = int(Wc // l1)
        swb = Wc - (nw1b * l1)
        slb = Lc - (nl1b * w1)
        nl2b = nl1b
        nw2b = int(swb // w2) if swb >= w2 else 0
        nl3b = int(slb // l2) if slb >= l2 else 0
        nw3b = nw1b
        nl4b = int(slb // l2) if slb >= l2 else 0
        nw4b = int(swb // w2) if swb >= w2 else 0
        total2 = (nl1b * nw1b + nl2b * nw2b + nl3b * nw3b + nl4b * nw4b) * nh

        return int(max(total1, total2))

    def calc_pack_per_container(flow_type: str, pkg: dict) -> int | None:
        flow_u = str(flow_type or "").strip().upper()
        if flow_u not in TRANSPORT_OPERATIVE_DIMS:
            return None
        l = pkg.get("pkg_length_mm")
        w = pkg.get("pkg_width_mm")
        h = pkg.get("pkg_height_mm")
        if l is None or w is None or h is None:
            return None
        by_volume = _max_packs_by_volume(TRANSPORT_OPERATIVE_DIMS[flow_u], (float(l), float(w), float(h)))
        if by_volume <= 0:
            return 0
        wf = pkg.get("pkg_weight_full")
        if wf is None or float(wf) <= 0:
            return int(by_volume)
        by_weight = int(TRANSPORT_MAX_WEIGHT[flow_u] // float(wf))
        return int(min(by_volume, by_weight))

    # Helpers
    def norm(s):
        return str(s).strip().upper() if pd.notna(s) else ""

    # Accent/diacritic-insensitive canonicalization helpers
    def _strip_accents(text: str | None) -> str:
        if not text:
            return ""
        try:
            return ''.join(ch for ch in unicodedata.normalize('NFKD', str(text)) if not unicodedata.combining(ch))
        except Exception:
            return str(text)

    def _canon(text: str | None) -> str:
        if not text:
            return ""
        t = _strip_accents(str(text)).upper()
        t = re.sub(r"[\./,;:_-]+", " ", t)
        t = re.sub(r"\s+", " ", t).strip()
        return t

    def _fuzzy_pick(query: str, choices: Iterable[str], score_cutoff: int = 90) -> Tuple[Optional[str], Optional[int]]:
        if not query:
            return None, None
        try:
            if rf_process is not None and rf_fuzz is not None:
                res = rf_process.extractOne(query, list(choices), scorer=rf_fuzz.token_sort_ratio, score_cutoff=score_cutoff)
                if res is None:
                    return None, None
                cand, score, _ = res
                return str(cand), int(score)
        except Exception:
            pass
        # fallback exact canon match
        qn = _canon(query)
        for c in choices:
            if _canon(c) == qn:
                return str(c), 100
        return None, None

    # Static inland distance estimates removed: distances are now fully dynamic via coordinates

    def get_pol(origin_cc: str):
        # Try specific country mapping in HORSE-PUERTO
        rec = find_port_by_country(df_hp, origin_cc)
        if rec and rec.get("Port"):
            return str(rec.get("Port"))
        return ""

    def get_pod(dest_plant: str, dest_cc: str):
        # Prefer exact factory match
        rec = map_factory_to_port(df_hp, dest_plant)
        if rec and rec.get("Port"):
            return str(rec.get("Port"))
        # Fallback by destination country
        rec2 = find_port_by_country(df_hp, dest_cc)
        if rec2 and rec2.get("Port"):
            return str(rec2.get("Port"))
        return ""

    def ports_for_country_from_hp(cc: str) -> list[str]:
        try:
            if not cc:
                return []
            mask = df_hp["Country Code"].astype(str).str.upper() == str(cc).strip().upper()
            m = df_hp.loc[mask]
            if m.empty:
                return []
            vals = m.get("Port")
            ports = vals.dropna().astype(str).str.strip().tolist() if vals is not None else []
            seen = set(); res = []
            for p in ports:
                pu = p.upper()
                if pu not in seen:
                    seen.add(pu)
                    res.append(p)
            return res
        except Exception:
            return []

    def _unique_ports(col_name: str) -> list[str]:
        try:
            vals = df_mp[col_name].dropna().astype(str).str.strip()
            # preserve original casing for output but dedupe by upper
            seen = set(); res = []
            for v in vals.tolist():
                vu = v.upper()
                if vu not in seen:
                    seen.add(vu)
                    res.append(v)
            return res
        except Exception:
            return []

    # Country-specific preferred port codes (by column name POL/POD)
    PORT_PREFERENCES: dict[str, dict[str, list[str]]] = {
        # China: prefer Shanghai as POL when available
        "CN": {"POL": ["CNSHA"]},
    }

    def _find_country_cols():
        """Detect country columns in MAIN PORTS for POL/POD sides.
        1) Name-based heuristic.
        2) Later refined using UN/LOC country prefix match vs POL/POD columns (if resolvable).
        Returns (pol_country_col, pod_country_col) or (None, None) if not found.
        """
        cols = [str(c) for c in df_mp.columns]
        lc = [c.lower().strip() for c in cols]
        pol_cc = None
        pod_cc = None
        patterns_pol = [
            "pol country code", "pol country", "origin country code", "origin country",
            "origin country (code)", "origin cc", "pol cc"
        ]
        patterns_pod = [
            "pod country code", "pod country", "destination country code", "destination country",
            "destination country (code)", "destination cc", "pod cc"
        ]
        for i, name in enumerate(lc):
            if name in patterns_pol and pol_cc is None:
                pol_cc = cols[i]
            if name in patterns_pod and pod_cc is None:
                pod_cc = cols[i]
        if pol_cc is None:
            for i, name in enumerate(lc):
                if ("pol" in name or "origin" in name) and ("country" in name or "code" in name or name.endswith(" cc") or " cc" in name):
                    pol_cc = cols[i]
                    break
        if pod_cc is None:
            for i, name in enumerate(lc):
                if ("pod" in name or "destination" in name) and ("country" in name or "code" in name or name.endswith(" cc") or " cc" in name):
                    pod_cc = cols[i]
                    break
        return pol_cc, pod_cc

    pol_cc_col, pod_cc_col = _find_country_cols()

    # Country cell comparator using only MAIN PORTS values (no HORSE-PUERTO)
    def _normalize(s):
        return str(s).strip().upper() if s is not None and str(s).strip() != "" else ""

    def _country_matches_any(cell_val: str, targets: set[str]) -> bool:
        if not cell_val or not targets:
            return False
        v = _normalize(cell_val)
        if not v:
            return False
        return v in targets

    # Resolve actual POL/POD column names in MAIN PORTS and TRANSITTIME
    def _resolve_port_cols(df: pd.DataFrame) -> tuple[str | None, str | None]:
        candidates_pol = ["POL", "Port of Loading", "Origin Port", "POL CODE", "POL Code"]
        candidates_pod = ["POD", "Port of Discharge", "Destination Port", "POD CODE", "POD Code"]
        lcmap = {str(c).lower().strip(): str(c) for c in df.columns}
        pol_col = None
        pod_col = None
        for cand in candidates_pol:
            k = cand.lower().strip()
            if k in lcmap:
                pol_col = lcmap[k]
                break
        for cand in candidates_pod:
            k = cand.lower().strip()
            if k in lcmap:
                pod_col = lcmap[k]
                break
        # Contains-based fallback
        if pol_col is None:
            for c in df.columns:
                cl = str(c).lower()
                if ("pol" in cl or "loading" in cl or "origin port" in cl) and ("port" in cl or "code" in cl):
                    pol_col = str(c)
                    break
        if pod_col is None:
            for c in df.columns:
                cl = str(c).lower()
                if ("pod" in cl or "discharge" in cl or "destination port" in cl) and ("port" in cl or "code" in cl):
                    pod_col = str(c)
                    break
        return pol_col, pod_col

    pol_col_mp, pod_col_mp = _resolve_port_cols(df_mp)
    pol_col_tt, pod_col_tt = _resolve_port_cols(df_tt)

    def _refine_country_cols(pol_cc_guess: str | None, pod_cc_guess: str | None) -> tuple[str | None, str | None]:
        """Use UN/LOC country prefix from POL/POD codes to assign the most likely country columns.
        We choose the column where POL's first-2-letter country code matches column value most often (for POL side),
        and likewise for POD.
        If scores are zero or ties remain, keep original guesses.
        """
        try:
            if not pol_col_mp or pol_col_mp not in df_mp.columns or not pod_col_mp or pod_col_mp not in df_mp.columns:
                return pol_cc_guess, pod_cc_guess
            # Candidate country-like columns
            cand_cols = []
            for c in df_mp.columns:
                cl = str(c).lower()
                if ("country" in cl or "code" in cl or cl.endswith(" cc") or " cc" in cl):
                    cand_cols.append(str(c))
            if not cand_cols:
                return pol_cc_guess, pod_cc_guess
            # Sample rows
            sample = df_mp.dropna(subset=[pol_col_mp, pod_col_mp]).head(500)
            def cc_from_port(pval: str) -> str:
                s = str(pval).strip().upper()
                return s[:2] if len(s) >= 2 else ""
            pol_scores = {}
            pod_scores = {}
            for cc_col in cand_cols:
                try:
                    pol_match = 0; pol_total = 0
                    for _, rr in sample.iterrows():
                        cc_val = str(rr.get(cc_col, "")).strip().upper()
                        pol_port = rr.get(pol_col_mp)
                        if pd.notna(cc_val) and pd.notna(pol_port):
                            pol_total += 1
                            if cc_val == cc_from_port(pol_port):
                                pol_match += 1
                    pol_scores[cc_col] = (pol_match, pol_total)
                    pod_match = 0; pod_total = 0
                    for _, rr in sample.iterrows():
                        cc_val = str(rr.get(cc_col, "")).strip().upper()
                        pod_port = rr.get(pod_col_mp)
                        if pd.notna(cc_val) and pd.notna(pod_port):
                            pod_total += 1
                            if cc_val == cc_from_port(pod_port):
                                pod_match += 1
                    pod_scores[cc_col] = (pod_match, pod_total)
                except Exception:
                    continue
            def best_col(scores: dict) -> str | None:
                items = []
                for k, (m, t) in scores.items():
                    ratio = (m / t) if t else 0.0
                    items.append((ratio, m, k))
                if not items:
                    return None
                items.sort(reverse=True)
                return items[0][2]
            pol_best = best_col(pol_scores)
            pod_best = best_col(pod_scores)
            # If bests are None or zero-ratio, keep guesses
            def ratio_of(scores, col):
                if col is None or col not in scores:
                    return 0.0
                m, t = scores[col]
                return (m / t) if t else 0.0
            if ratio_of(pol_scores, pol_best) == 0.0 and ratio_of(pod_scores, pod_best) == 0.0:
                return pol_cc_guess, pod_cc_guess
            # Avoid assigning the same column to both sides if avoidable
            if pol_best and pod_best and pol_best == pod_best:
                # pick second best for POD if available
                sorted_pod = sorted([( (m/t if t else 0.0), m, k) for k,(m,t) in pod_scores.items() ], reverse=True)
                if len(sorted_pod) > 1:
                    pod_best = sorted_pod[1][2]
            return pol_best or pol_cc_guess, pod_best or pod_cc_guess
        except Exception:
            return pol_cc_guess, pod_cc_guess

    pol_cc_col, pod_cc_col = _refine_country_cols(pol_cc_col, pod_cc_col)

    def _port_country_from_hp(port: str) -> str | None:
        try:
            m = df_hp[df_hp["Port"].astype(str).str.upper() == str(port).strip().upper()]
            if not m.empty:
                cc = m.iloc[0].get("Country Code")
                return str(cc).strip().upper() if pd.notna(cc) else None
        except Exception:
            return None
        return None

    def pick_port_from_main_ports(cc: str, kind: str = "POL", other_port: str | None = None, near_point: tuple[float, float] | None = None) -> str:
        """Pick a port in MAIN PORTS for given country.
        - Enforce country match via MAIN PORTS country columns when present; else via HORSE-PUERTO mapping.
        - If other_port is provided, restrict to rows pairing with it.
        - If near_point provided and we can resolve coordinates, choose nearest; otherwise most frequent.
        """
        # Use resolved column names
        col = (pol_col_mp if kind.upper() == "POL" else pod_col_mp) or ("POL" if kind.upper() == "POL" else "POD")
        other_col = (pod_col_mp if kind.upper() == "POL" else pol_col_mp) or ("POD" if kind.upper() == "POL" else "POL")

        # Start from full scope and apply pair filter
        df_scope = df_mp
        if other_port:
            df_scope = df_scope[df_scope[other_col].astype(str).str.upper() == str(other_port).strip().upper()]

        # Enforce country on scope using MAIN PORTS columns if available
        if kind.upper() == "POL" and pol_cc_col is not None and pol_cc_col in df_scope.columns:
            df_scope = df_scope[df_scope[pol_cc_col].apply(lambda v: _country_matches_any(v, {_normalize(cc)}))]
        if kind.upper() == "POD" and pod_cc_col is not None and pod_cc_col in df_scope.columns:
            df_scope = df_scope[df_scope[pod_cc_col].apply(lambda v: _country_matches_any(v, {_normalize(cc)}))]

        # Collect candidate ports from scope
        scope_ports = df_scope[col].dropna().astype(str).str.strip().tolist() if col in df_scope.columns else []

        # If MAIN PORTS lacks country columns, intersect with HORSE-PUERTO mapping (or validate port->country)
        # No HORSE-PUERTO based filtering here by request; rely solely on MAIN PORTS columns

        # Dedupe preserve order
        seen = set(); valid = []
        for p in scope_ports:
            pu = p.upper()
            if pu not in seen:
                seen.add(pu)
                valid.append(p)
        if not valid:
            return ""

        # Country-specific preferences
        prefs = PORT_PREFERENCES.get(cc.upper(), {}).get(col, [])
        for pref in prefs:
            for p in valid:
                if str(p).strip().upper() == pref.upper():
                    return p

        # Proximity if possible
        if near_point is not None and cc:
            try:
                scored_dist = []
                for p in valid:
                    plat, plon, _ = resolve_point(geo, cc, port=p)
                    if plat is not None:
                        km = road_km_between((plat, plon), near_point)
                        scored_dist.append((km, p))
                if scored_dist:
                    scored_dist.sort()
                    return scored_dist[0][1]
            except Exception:
                pass

        # Fallback to most frequent in scope
        scored = []
        for p in valid:
            try:
                count = len(df_scope[df_scope[col].astype(str).str.upper() == str(p).strip().upper()])
            except Exception:
                count = 0
            scored.append((count, p))
        scored.sort(reverse=True)
        return scored[0][1]

    def _port_coords_for_distance(cc: str, port: str, origin_point: tuple[float, float] | None) -> tuple[float | None, float | None]:
        # Prefer normalized Ports Locations with swap-guard; fallback to geo index resolve
        f_lat, f_lon = _port_coords_from_ports_locations(port)
        if f_lat is not None and f_lon is not None:
            b_lat, b_lon, _ = _best_port_coords_for_origin(origin_point, f_lat, f_lon)
            return b_lat, b_lon
        plat, plon, _ = resolve_point(geo, cc, port=port)
        return (plat, plon)

    def select_pol_pod_pair(oc: str, dc: str, origin_point: tuple[float, float] | None, oc_name: str | None = None, dc_name: str | None = None) -> tuple[str, str, str]:
        """Select a (POL, POD) pair strictly from MAIN PORTS matching origin and destination countries.
        Returns (pol, pod, reason) where reason is 'preferencia', 'cercania', 'frecuencia' or 'sin-candidatos'.
        """
        if not oc or not dc or df_mp.empty:
            return "", "", "sin-candidatos"
        df_scope = df_mp
        # If MAIN PORTS lacks country columns for either side, do not guess; leave blank by design
        if not (pol_cc_col and pol_cc_col in df_mp.columns and pod_cc_col and pod_cc_col in df_mp.columns):
            return "", "", "sin-columnas-pais"
        # Helper: if UN/LOC prefixes indicate reversed assignment, swap them
        def _maybe_swap(pol: str, pod: str, why: str) -> tuple[str, str, str]:
            oc2 = (oc or "").strip().upper()
            dc2 = (dc or "").strip().upper()
            if len(oc2) == 2 and len(dc2) == 2 and pol and pod:
                polcc = str(pol).strip().upper()[:2]
                podcc = str(pod).strip().upper()[:2]
                if polcc == dc2 and podcc == oc2 and not (polcc == oc2 and podcc == dc2):
                    return pod, pol, why + "+swap-unloc"
            return pol, pod, why
        # Apply country filters when columns exist
        if pol_cc_col is not None and pol_cc_col in df_scope.columns:
            oc_targets = {_normalize(oc), _normalize(oc_name)} if oc_name else {_normalize(oc)}
            df_scope = df_scope[df_scope[pol_cc_col].apply(lambda v: _country_matches_any(v, oc_targets))]
        if pod_cc_col is not None and pod_cc_col in df_scope.columns:
            dc_targets = {_normalize(dc), _normalize(dc_name)} if dc_name else {_normalize(dc)}
            df_scope = df_scope[df_scope[pod_cc_col].apply(lambda v: _country_matches_any(v, dc_targets))]
        # Drop rows without POL/POD
        if (pol_col_mp is None or pol_col_mp not in df_scope.columns) or (pod_col_mp is None or pod_col_mp not in df_scope.columns):
            return "", "", "sin-candidatos"
        df_scope = df_scope.dropna(subset=[pol_col_mp, pod_col_mp])
        # If scope is empty, try symmetric filter (in case country columns are swapped in data)
        if df_scope.empty:
            df_alt = df_mp
            if pol_cc_col is not None and pol_cc_col in df_alt.columns:
                dc_targets = {_normalize(dc), _normalize(dc_name)} if dc_name else {_normalize(dc)}
                df_alt = df_alt[df_alt[pol_cc_col].apply(lambda v: _country_matches_any(v, dc_targets))]
            if pod_cc_col is not None and pod_cc_col in df_alt.columns:
                oc_targets = {_normalize(oc), _normalize(oc_name)} if oc_name else {_normalize(oc)}
                df_alt = df_alt[df_alt[pod_cc_col].apply(lambda v: _country_matches_any(v, oc_targets))]
            if (pol_col_mp is not None and pol_col_mp in df_alt.columns) and (pod_col_mp is not None and pod_col_mp in df_alt.columns):
                df_alt = df_alt.dropna(subset=[pol_col_mp, pod_col_mp])
                if not df_alt.empty:
                    # Build pairs and select using same strategy, then UN/LOC swap will correct direction
                    pairs = []
                    seen = set()
                    for _, rr in df_alt.iterrows():
                        pol = str(rr[pol_col_mp]).strip(); pod = str(rr[pod_col_mp]).strip()
                        key = (pol.upper(), pod.upper())
                        if key not in seen:
                            seen.add(key); pairs.append((pol, pod))
                    # Preferencia
                    prefs = PORT_PREFERENCES.get(oc.upper(), {}).get("POL", [])
                    for pref in prefs:
                        for pol, pod in pairs:
                            if pol.strip().upper() == pref.upper():
                                pol, pod, why = _maybe_swap(pol, pod, "preferencia-swap-cols")
                                return pol, pod, why
                    # Cercania
                    if origin_point is not None:
                        try:
                            scored = []
                            for pol, pod in pairs:
                                plat, plon, _ = resolve_point(geo, oc, port=pol)
                                if plat is not None:
                                    km = road_km_between((plat, plon), origin_point)
                                    scored.append((km, pol, pod))
                            if scored:
                                scored.sort()
                                _, pol_b, pod_b = scored[0]
                                pol_b, pod_b, why = _maybe_swap(pol_b, pod_b, "cercania-swap-cols")
                                return pol_b, pod_b, why
                        except Exception:
                            pass
                    # Frecuencia
                    counts = {}
                    for pol, pod in pairs:
                        mask = (df_alt[pol_col_mp].astype(str).str.upper() == pol.upper()) & (df_alt[pod_col_mp].astype(str).str.upper() == pod.upper())
                        counts[(pol, pod)] = int(mask.sum())
                    best = sorted(pairs, key=lambda pp: counts.get(pp, 0), reverse=True)[0]
                    pol_b, pod_b, why = _maybe_swap(best[0], best[1], "frecuencia-swap-cols")
                    return pol_b, pod_b, why

        if df_scope.empty:
            # Fallback A: elegir POL por país de origen y luego POD por país destino desde las filas con ese POL
            pol_only, why_pol = select_pol_from_origin(oc, origin_point, oc_name)
            if pol_only:
                pol_c = pol_col_mp or "POL"
                pod_c = pod_col_mp or "POD"
                rows_pol = df_mp[df_mp[pol_c].astype(str).str.upper() == pol_only.upper()]
                if pod_cc_col is not None and pod_cc_col in rows_pol.columns:
                    dc_targets = {_normalize(dc), _normalize(dc_name)} if dc_name else {_normalize(dc)}
                    rows_pol = rows_pol[rows_pol[pod_cc_col].apply(lambda v: _country_matches_any(v, dc_targets))]
                if not rows_pol.empty:
                    pod_counts = rows_pol[pod_c].astype(str).str.upper().value_counts()
                    pod_best_u = pod_counts.index[0]
                    orig = rows_pol[rows_pol[pod_c].astype(str).str.upper() == pod_best_u].iloc[0][pod_c]
                    polonly, podorig, why = pol_only, str(orig), f"fallback-pol({why_pol})"
                    polonly, podorig, why = _maybe_swap(polonly, podorig, why)
                    return polonly, podorig, why
            # Fallback B: elegir POD por país destino y emparejar con un POL válido del país origen
            rows_pod = df_mp
            pol_c = pol_col_mp or "POL"
            pod_c = pod_col_mp or "POD"
            if pod_cc_col is not None and pod_cc_col in rows_pod.columns:
                dc_targets = {_normalize(dc), _normalize(dc_name)} if dc_name else {_normalize(dc)}
                rows_pod = rows_pod[rows_pod[pod_cc_col].apply(lambda v: _country_matches_any(v, dc_targets))]
            if not rows_pod.empty:
                if pol_cc_col is not None and pol_cc_col in rows_pod.columns:
                    oc_targets = {_normalize(oc), _normalize(oc_name)} if oc_name else {_normalize(oc)}
                    rows_pod = rows_pod[rows_pod[pol_cc_col].apply(lambda v: _country_matches_any(v, oc_targets))]
                if not rows_pod.empty:
                    pair_counts = rows_pod.groupby([rows_pod[pol_c].astype(str).str.upper(), rows_pod[pod_c].astype(str).str.upper()]).size().sort_values(ascending=False)
                    pol_best_u, pod_best_u = pair_counts.index[0]
                    row = rows_pod[(rows_pod[pol_c].astype(str).str.upper() == pol_best_u) & (rows_pod[pod_c].astype(str).str.upper() == pod_best_u)].iloc[0]
                    polv, podv, why = str(row[pol_c]), str(row[pod_c]), "fallback-pod(frecuencia)"
                    polv, podv, why = _maybe_swap(polv, podv, why)
                    return polv, podv, why
            return "", "", "sin-candidatos"

        # Deduplicate to a list of pairs preserving order of appearance
        pairs = []
        seen = set()
        for _, rr in df_scope.iterrows():
            pol = str(rr[pol_col_mp]).strip()
            pod = str(rr[pod_col_mp]).strip()
            key = (pol.upper(), pod.upper())
            if key not in seen:
                seen.add(key)
                pairs.append((pol, pod))

        # Proximity FIRST: choose pair whose POL is closest to origin_point
        if origin_point is not None:
            try:
                scored = []
                for pol, pod in pairs:
                    plat, plon = _port_coords_for_distance(oc, pol, origin_point)
                    if plat is not None and plon is not None:
                        km = road_km_between((plat, plon), origin_point)
                        scored.append((km, pol, pod))
                if scored:
                    scored.sort()
                    _, pol_b, pod_b = scored[0]
                    pol_b, pod_b, why = _maybe_swap(pol_b, pod_b, "cercania")
                    return pol_b, pod_b, why
            except Exception:
                pass

        # Preference: OC-specific preferred POL (used when no origin point)
        prefs = PORT_PREFERENCES.get(oc.upper(), {}).get("POL", [])
        for pref in prefs:
            for pol, pod in pairs:
                if pol.strip().upper() == pref.upper():
                    pol, pod, why = _maybe_swap(pol, pod, "preferencia")
                    return pol, pod, why

        # Frequency: choose most frequent pair occurrence in df_scope
        counts = {}
        for pol, pod in pairs:
            mask = (df_scope[pol_col_mp].astype(str).str.upper() == pol.upper()) & (df_scope[pod_col_mp].astype(str).str.upper() == pod.upper())
            counts[(pol, pod)] = int(mask.sum())
        best = sorted(pairs, key=lambda pp: counts.get(pp, 0), reverse=True)[0]
        pol_b, pod_b, why = _maybe_swap(best[0], best[1], "frecuencia")
        return pol_b, pod_b, why

    def select_pol_from_origin(oc: str, origin_point: tuple[float, float] | None, oc_name: str | None = None) -> tuple[str, str]:
        """Select POL strictly by origin country from MAIN PORTS.
        Returns (pol, why) where why is one of {preferencia, cercania, frecuencia, sin-candidatos}.
        """
        if not oc:
            return "", "sin-candidatos"
        # Scope filtered by origin country if MP has such column
        df_scope = df_mp
        if pol_cc_col is not None:
            oc_targets = {_normalize(oc), _normalize(oc_name)} if oc_name else {_normalize(oc)}
            df_scope = df_scope[df_scope[pol_cc_col].apply(lambda v: _country_matches_any(v, oc_targets))]
        # Candidate POLs from scope
        pols = df_scope[pol_col_mp].dropna().astype(str).str.strip().tolist() if (pol_col_mp and pol_col_mp in df_scope.columns) else []
        # If MP lacks country columns, intersect with HORSE-PUERTO ports for oc
        if pol_cc_col is None:
            hp_ports = set([p.upper() for p in ports_for_country_from_hp(oc)])
            if hp_ports:
                pols = [p for p in pols if p.upper() in hp_ports]
            else:
                pols = [p for p in pols if _port_country_from_hp(p) == oc.upper()]
        # Dedupe
        seen = set(); valid = []
        for p in pols:
            pu = p.upper()
            if pu not in seen:
                seen.add(pu)
                valid.append(p)
        if not valid:
            return "", "sin-candidatos"
        # Proximity FIRST
        if origin_point is not None:
            try:
                scored = []
                for p in valid:
                    plat, plon = _port_coords_for_distance(oc, p, origin_point)
                    if plat is not None and plon is not None:
                        km = road_km_between((plat, plon), origin_point)
                        scored.append((km, p))
                if scored:
                    scored.sort()
                    return scored[0][1], "cercania"
            except Exception:
                pass
        # Preference then frequency when no origin point
        prefs = PORT_PREFERENCES.get(oc.upper(), {}).get("POL", [])
        for pref in prefs:
            for p in valid:
                if p.strip().upper() == pref.upper():
                    return p, "preferencia"
        # Frequency
        scored = []
        for p in valid:
            try:
                count = len(df_scope[df_scope[pol_col_mp].astype(str).str.upper() == p.strip().upper()])
            except Exception:
                count = 0
            scored.append((count, p))
        scored.sort(reverse=True)
        return scored[0][1], "frecuencia"

    def _resolve_col(df: pd.DataFrame, candidates: list[str], contains_any: list[str] | None = None) -> str | None:
        """Find a column in df by case-insensitive exact name in candidates, else by contains keywords in order."""
        cols = [str(c) for c in df.columns]
        lcmap = {str(c).lower().strip(): c for c in cols}
        for cand in candidates:
            k = cand.lower().strip()
            if k in lcmap:
                return lcmap[k]
        if contains_any:
            lower_cols = [(c, c.lower()) for c in cols]
            for key in contains_any:
                for orig, low in lower_cols:
                    if key.lower() in low:
                        return orig
        return None

    def get_ocean_rate_and_tt(pol: str, pod: str):
        rate = None
        tt_days = None
        if pol and pod and not df_vtt_routes.empty:
            try:
                pol_vtt = _resolve_col(df_vtt_routes, ["POL"], ["pol"])
                pod_vtt = _resolve_col(df_vtt_routes, ["POD"], ["pod"])
                tt_vtt = _resolve_col(df_vtt_routes, ["Transit time", "Transit Time"], ["transit time", "transit"])
                sec_vtt = _resolve_col(df_vtt_routes, ["Time for security"], ["time for security", "security"])
                if pol_vtt and pod_vtt and tt_vtt:
                    mv = df_vtt_routes[
                        (df_vtt_routes[pol_vtt].astype(str).str.upper().str.strip() == str(pol).upper().strip()) &
                        (df_vtt_routes[pod_vtt].astype(str).str.upper().str.strip() == str(pod).upper().strip())
                    ]
                    if not mv.empty:
                        try:
                            tvals = pd.to_numeric(mv[tt_vtt], errors="coerce")
                            if sec_vtt and sec_vtt in mv.columns:
                                svals = pd.to_numeric(mv[sec_vtt], errors="coerce")
                            else:
                                svals = pd.Series([float("nan")] * len(mv), index=mv.index)
                            totals = tvals.fillna(0) + svals.fillna(0)
                            valid = tvals.notna() | svals.notna()
                            vv = totals[valid]
                            tt_days = float(vv.min()) if not vv.empty else None
                        except Exception:
                            tt_days = None
            except Exception:
                pass
        if pol and pod and not df_mp.empty:
            try:
                pol_c = pol_col_mp or "POL"
                pod_c = pod_col_mp or "POD"
                m = df_mp[(df_mp[pol_c].astype(str).str.upper() == pol.upper()) & (df_mp[pod_c].astype(str).str.upper() == pod.upper())]
                if not m.empty:
                    # Rate
                    rate_col = _resolve_col(
                        df_mp,
                        ["Rate 40ft all-in", "Rate 40FT ALL-IN", "Rate 40ft", "Ocean Rate"],
                        ["rate", "40", "all"]
                    )
                    if rate_col:
                        try:
                            rate = float(m.iloc[0].get(rate_col)) if pd.notna(m.iloc[0].get(rate_col)) else None
                        except Exception:
                            rate = None
                    # TT fallback in MAIN PORTS (VTT POL/POD table has priority)
                    if tt_days is None:
                        tt_col = _resolve_col(
                            df_mp,
                            ["TT_OVS", "TT", "TT (days)", "TT(days)", "Transit Time"],
                            ["tt", "transit time"]
                        )
                        if tt_col:
                            try:
                                tt_days = float(m.iloc[0].get(tt_col)) if pd.notna(m.iloc[0].get(tt_col)) else None
                            except Exception:
                                tt_days = None
            except Exception:
                pass
        if tt_days is None and pol and pod and not df_tt.empty:
            try:
                # Find appropriate TT column in TRANSITTIME sheet
                tt_col2 = _resolve_col(df_tt, ["Transit Time", "TT", "TT (days)", "TT(days)"], ["transit", "tt"])
                if tt_col2 is None:
                    tt_col2 = "Transit Time" if "Transit Time" in df_tt.columns else None
                if tt_col2 is not None:
                    pol_t = pol_col_tt or "POL"
                    pod_t = pod_col_tt or "POD"
                    m2 = df_tt[(df_tt[pol_t].astype(str).str.upper() == pol.upper()) & (df_tt[pod_t].astype(str).str.upper() == pod.upper())]
                    if not m2.empty:
                        try:
                            tt_days = float(m2.iloc[0].get(tt_col2)) if pd.notna(m2.iloc[0].get(tt_col2)) else None
                        except Exception:
                            tt_days = None
            except Exception:
                pass
        return rate, tt_days

    # Helper: parse city to extract a ZIP token if embedded (e.g., "Mundhwa 34190")
    def _parse_city_zip(city_val: str, zip_val: str) -> tuple[str, str, str]:
        note = ""
        city_raw = (city_val or "").strip()
        zip_raw = (zip_val or "").strip()
        if city_raw and not zip_raw:
            # find trailing 4-6 digit group anywhere; prefer the last group
            m = re.findall(r"(\d{4,6})", city_raw)
            if m:
                zip_raw = m[-1]
                # remove that group and surrounding separators from city
                city_clean = re.sub(r"[\s,-]*" + re.escape(zip_raw) + r"\b", "", city_raw).strip()
                note = f"ciudad parseada: '{city_raw}'→ ciudad='{city_clean}', zip='{zip_raw}'"
                return city_clean, zip_raw, note
        return city_raw, zip_raw, note

    # Helper: city alias from database sheet CITY_ALIASES (columns: Country Code, From City, To City)
    def _city_alias(cc: str, city: str) -> str | None:
        try:
            if df_city_aliases is None or df_city_aliases.empty:
                return None
            cc_u = (cc or "").strip().upper()
            city_u = (city or "").strip().upper()
            # Try common column name variants
            cols = {c.lower().strip(): c for c in df_city_aliases.columns}
            cc_col = cols.get("country code") or cols.get("cc") or "Country Code"
            from_col = cols.get("from city") or cols.get("from") or "From City"
            to_col = cols.get("to city") or cols.get("to") or "To City"
            m = df_city_aliases[
                df_city_aliases.get(cc_col, pd.Series()).astype(str).str.upper().eq(cc_u) &
                df_city_aliases.get(from_col, pd.Series()).astype(str).str.strip().str.upper().eq(city_u)
            ]
            if not m.empty:
                val = m.iloc[0].get(to_col)
                return str(val).strip() if pd.notna(val) else None
        except Exception:
            return None
        return None

    def _normalize_zip_token(s: str | None) -> str:
        if not s:
            return ""
        z = str(s).strip().upper().replace(" ", "").replace("-", "")
        # Keep only alphanumerics
        z = "".join([ch for ch in z if ch.isalnum()])
        return z

    def _normalize_city_for_country(cc: str, city: str | None) -> str:
        """Normalize city names by country to improve matching.
        - CN: strip common suffixes like ' SHI' or ' CITY' (e.g., 'NANCHANG SHI' -> 'NANCHANG').
        """
        c = (city or "").strip()
        if not c:
            return c
        # Remove accents universally to align with DB entries that might include diacritics
        c = ''.join(ch for ch in unicodedata.normalize('NFKD', c) if not unicodedata.combining(ch))
        if (cc or "").strip().upper() == "CN":
            cu = c.upper()
            for suf in (" SHI", " CITY", " SHI CITY"):
                if cu.endswith(suf):
                    return c[: -len(suf)].strip()
        return c

    def _country_zip_is_valid(cc: str, z: str) -> bool:
        if not z:
            return False
        cc = (cc or "").strip().upper()
        if cc == "ES":
            return z.isdigit() and len(z) == 5
        if cc == "IN":
            return z.isdigit() and len(z) == 6
        if cc == "BR":
            return z.isdigit() and len(z) == 8
        if cc == "TR":
            return z.isdigit() and len(z) == 5
        # Default: accept 3-10 alphanumeric
        return 3 <= len(z) <= 10

    # NOTE: No ZIP enrichment from HORSE-PUERTO. ZIPs must come from CITY_ZIPS (or alias) per data-first policy.

    def _zip_from_city_sheet(cc: str, city: str) -> str | None:
        try:
            if df_city_zips is None or df_city_zips.empty:
                return None
            cc_u = (cc or "").strip().upper()
            city_u = (city or "").strip().upper()
            cols = {c.lower().strip(): c for c in df_city_zips.columns}
            cc_col = cols.get("country code") or cols.get("cc") or "Country Code"
            city_col = cols.get("city") or "City"
            zip_col = cols.get("zip") or cols.get("zip code") or cols.get("postal code") or "ZIP"
            m = df_city_zips[
                df_city_zips.get(cc_col, pd.Series()).astype(str).str.upper().eq(cc_u) &
                df_city_zips.get(city_col, pd.Series()).astype(str).str.strip().str.upper().eq(city_u)
            ]
            if not m.empty:
                val = m.iloc[0].get(zip_col)
                return str(val).strip() if pd.notna(val) else None
            # Accent-insensitive and fuzzy city matching within the same country
            df_cc = df_city_zips[df_city_zips.get(cc_col, pd.Series()).astype(str).str.upper().eq(cc_u)]
            if not df_cc.empty:
                try:
                    df_cc = df_cc.copy()
                    df_cc['__canon_city'] = df_cc.get(city_col, pd.Series()).astype(str).map(_canon)
                    key = _canon(city)
                    m2 = df_cc[df_cc['__canon_city'] == key]
                    if not m2.empty:
                        val2 = m2.iloc[0].get(zip_col)
                        return str(val2).strip() if pd.notna(val2) else None
                    choices = df_cc.get(city_col, pd.Series()).dropna().astype(str).tolist()
                    best, score = _fuzzy_pick(city, choices, score_cutoff=92)
                    if best:
                        m3 = df_cc[df_cc.get(city_col, pd.Series()).astype(str).str.upper() == str(best).upper()]
                        if not m3.empty:
                            val3 = m3.iloc[0].get(zip_col)
                            return str(val3).strip() if pd.notna(val3) else None
                except Exception:
                    pass
        except Exception:
            return None
        return None

    def validate_and_enrich_zip(cc: str, city: str, zip_in: str) -> tuple[str, str | None]:
        """Return (zip_out, reason) if we corrected/assigned a ZIP using city+country context.
        Reasons: 'city_zips', 'alias_city_zips', None if unchanged or no candidate.
        """
        z = _normalize_zip_token(zip_in)
        if _country_zip_is_valid(cc, z):
            return z, None
        city_u = (city or "").strip().upper()
        # Try CITY_ZIPS sheet
        z2 = _zip_from_city_sheet(cc, city_u)
        z2n = _normalize_zip_token(z2)
        if z2n and _country_zip_is_valid(cc, z2n):
            return z2n, "city_zips"
        # Try alias then CITY_ZIPS again
        alias = _city_alias(cc, city_u)
        if alias:
            z3 = _zip_from_city_sheet(cc, alias)
            z3n = _normalize_zip_token(z3)
            if z3n and _country_zip_is_valid(cc, z3n):
                return z3n, "alias_city_zips"
        # No change
        return z, None

    # Direct coordinate fallback from 'Ports Locations' when GeoIndex cannot resolve a port by country
    def _port_coords_from_ports_locations(port_code: str) -> tuple[float | None, float | None]:
        try:
            if not port_code or df_ports is None or df_ports.empty:
                return None, None
            m = df_ports[df_ports["POL/POD"].astype(str).str.upper() == str(port_code).strip().upper()]
            if not m.empty:
                def _coerce_coord(val, kind: str) -> float | None:
                    try:
                        if pd.isna(val):
                            return None
                        if isinstance(val, str):
                            s = val.strip().replace(" ", "")
                            # Replace comma decimal with dot
                            s = s.replace(",", ".")
                            f = float(s)
                        else:
                            f = float(val)
                        # Fix common thousand-without-decimal issue (e.g., 39450 for 39.450)
                        if kind == "lat" and abs(f) > 90 and abs(f) <= 180000:
                            f = f / 1000.0
                        if kind == "lon" and abs(f) > 180 and abs(f) <= 360000:
                            f = f / 1000.0
                        # Final sanity
                        if kind == "lat" and abs(f) <= 90:
                            return f
                        if kind == "lon" and abs(f) <= 180:
                            return f
                        return None
                    except Exception:
                        return None
                lat = _coerce_coord(m.iloc[0].get("LAT"), "lat")
                lon = _coerce_coord(m.iloc[0].get("LONG"), "lon")
                if lat is not None and lon is not None:
                    return lat, lon
        except Exception:
            return None, None
        return None, None

    def _best_port_coords_for_origin(origin_point: tuple[float, float] | None,
                                     lat: float | None,
                                     lon: float | None) -> tuple[float | None, float | None, str]:
        """Return the most plausible (lat, lon) for a port, trying given (lat,lon) and swapped,
        choosing the one closest to origin_point if provided. Returns (lat, lon, tag) where tag indicates 'ports_locations' or 'ports_locations(swapped)'."""
        if lat is None or lon is None:
            return None, None, ""
        if origin_point is None:
            return lat, lon, "ports_locations"
        try:
            km_direct = road_km_between((lat, lon), origin_point)
            km_swapped = road_km_between((lon, lat), origin_point)
            if km_swapped < km_direct:
                return lon, lat, "ports_locations(swapped)"
            return lat, lon, "ports_locations"
        except Exception:
            return lat, lon, "ports_locations"

    def validate_ports_country(pol: str | None, pod: str | None, oc: str, dc: str, oc_name: str | None = None, dc_name: str | None = None) -> tuple[bool, bool]:
        """Validate that POL belongs to Origin Country and POD belongs to Destination Country using only MAIN PORTS.
        If MAIN PORTS lacks the corresponding country column, validation for that side is treated as True (cannot verify).
        """
        pol_ok = True
        pod_ok = True
        # Validate POL
        if pol and pol_col_mp and pol_cc_col and pol_cc_col in df_mp.columns:
            oc_targets = {_normalize(oc)} | ({_normalize(oc_name)} if oc_name else set())
            try:
                rows = df_mp[df_mp[pol_col_mp].astype(str).str.upper() == str(pol).strip().upper()]
                if not rows.empty and pol_cc_col in rows.columns:
                    pol_ok = bool(rows[rows[pol_cc_col].apply(lambda v: _country_matches_any(v, oc_targets))].shape[0] > 0)
            except Exception:
                pol_ok = True
        # Validate POD
        if pod and pod_col_mp and pod_cc_col and pod_cc_col in df_mp.columns:
            dc_targets = {_normalize(dc)} | ({_normalize(dc_name)} if dc_name else set())
            try:
                rows = df_mp[df_mp[pod_col_mp].astype(str).str.upper() == str(pod).strip().upper()]
                if not rows.empty and pod_cc_col in rows.columns:
                    pod_ok = bool(rows[rows[pod_cc_col].apply(lambda v: _country_matches_any(v, dc_targets))].shape[0] > 0)
            except Exception:
                pod_ok = True
        return pol_ok, pod_ok

    # New selection: choose POL/POD strictly from Ports Locations by country and proximity
    def _row_country_code_ports(rr) -> str:
        try:
            cval = rr.get("Country")
            c = str(cval).strip().upper() if pd.notna(cval) else ""
            # If it already looks like a code, keep; else map name→code
            return coerce_country_code(c, c)
        except Exception:
            return ""

    def _all_ports_for_country_from_ports_locations(cc: str) -> list[tuple[str, float | None, float | None]]:
        """Return list of (port_code, lat, lon) from Ports Locations matching given country by either:
        - Country column (name/code normalized), OR
        - UN/LOCODE prefix of POL/POD (first 2 letters).
        This avoids misses when the Country column uses 3-letter codes or names not in the HP map.
        """
        out = []
        try:
            if df_ports is None or df_ports.empty or not cc:
                return out
            cc_u = (cc or "").strip().upper()
            cc2 = cc_u[:2]
            for _, rr in df_ports.iterrows():
                try:
                    code = str(rr.get("POL/POD", "")).strip().upper()
                    if not code:
                        continue
                    # Accept row if Country matches target OR UN/LOC prefix matches
                    row_cc = _row_country_code_ports(rr)
                    pol_cc2 = code[:2]
                    if not (row_cc == cc_u or pol_cc2 == cc2):
                        continue
                    la, lo = _port_coords_from_ports_locations(code)
                    out.append((code, la, lo))
                except Exception:
                    continue
        except Exception:
            return out
        return out

    def select_port_nearest_from_ports_locations(cc: str, near_point: tuple[float, float] | None, side: str) -> tuple[str, str]:
        """Pick nearest port (POL or POD) in Ports Locations by country and proximity to near_point.
        Returns (port_code, reason). If near_point is None, returns most frequent/first available.
        """
        cands = _all_ports_for_country_from_ports_locations(cc)
        if not cands:
            return "", f"sin-candidatos-ports-locations-{side}"
        # Proximity first when near_point available
        if near_point is not None:
            scored = []
            for code, la, lo in cands:
                if la is None or lo is None:
                    continue
                bla, blo, _ = _best_port_coords_for_origin(near_point, la, lo)
                if bla is None or blo is None:
                    continue
                try:
                    km = road_km_between((bla, blo), near_point)
                except Exception:
                    continue
                scored.append((km, code))
            if scored:
                scored.sort()
                return scored[0][1], f"cercania-ports-locations-{side}"
        # Fallback: pick first available
        return cands[0][0], f"primero-ports-locations-{side}"

    def _ports_candidates_debug(cc: str, near_point: tuple[float, float] | None) -> list[dict]:
        """Return detailed candidate list for debug: [{code, has_coords, km}] sorted by km asc (None last)."""
        info = []
        cands = _all_ports_for_country_from_ports_locations(cc)
        for code, la, lo in cands:
            entry = {"code": code, "has_coords": False, "km": None}
            if la is not None and lo is not None and near_point is not None:
                bla, blo, _ = _best_port_coords_for_origin(near_point, la, lo)
                if bla is not None and blo is not None:
                    try:
                        entry["km"] = round(road_km_between((bla, blo), near_point), 1)
                        entry["has_coords"] = True
                    except Exception:
                        entry["has_coords"] = True
                else:
                    entry["has_coords"] = True
            else:
                entry["has_coords"] = (la is not None and lo is not None)
            info.append(entry)
        info.sort(key=lambda x: (float('inf') if x["km"] is None else x["km"]))
        return info

    def _city_coords_from_db(cc: str, city: str) -> tuple[float | None, float | None]:
        """Lookup city coordinates from GEO_LOCATIONS/CITY_COORDS sheet, with alias fallback."""
        try:
            cc_u = (cc or "").strip().upper()
            key_u = (city or "").strip().upper()
            key_c = _canon(city)
            # Manual fallback requested by user
            if cc_u == "CN" and key_c == "YIWU":
                return 29.3151, 120.0768
            if cc_u == "CN" and key_u == "WUHAN":
                return 30.5928, 114.3055
            if cc_u == "MA" and key_u == "TANGER":
                return 35.7595, -5.8340
            if cc_u == "CZ" and key_c in {"FRENSTAT POD RADHOSTEM", "FRENSTAT"}:
                return 49.5489, 18.2108

            if df_geo_cities is None or df_geo_cities.empty or not city:
                return None, None
            cols = {c.lower().strip(): c for c in df_geo_cities.columns}
            cc_col = cols.get("country code") or cols.get("cc") or "Country Code"
            city_col = cols.get("city") or "City"
            lat_col = cols.get("lat") or cols.get("latitude") or "Lat"
            lon_col = cols.get("lon") or cols.get("long") or cols.get("longitude") or "Long"
            m = df_geo_cities[
                df_geo_cities.get(cc_col, pd.Series()).astype(str).str.upper().eq(cc_u) &
                df_geo_cities.get(city_col, pd.Series()).astype(str).str.strip().str.upper().eq(key_u)
            ]
            if not m.empty:
                la = m.iloc[0].get(lat_col); lo = m.iloc[0].get(lon_col)
                if pd.notna(la) and pd.notna(lo):
                    return float(la), float(lo)
            # Alias fallback
            alias = _city_alias(cc, city)
            if alias:
                au = alias.strip().upper()
                m2 = df_geo_cities[
                    df_geo_cities.get(cc_col, pd.Series()).astype(str).str.upper().eq(cc_u) &
                    df_geo_cities.get(city_col, pd.Series()).astype(str).str.strip().str.upper().eq(au)
                ]
                if not m2.empty:
                    la = m2.iloc[0].get(lat_col); lo = m2.iloc[0].get(lon_col)
                    if pd.notna(la) and pd.notna(lo):
                        return float(la), float(lo)
            # Accent-insensitive exact and fuzzy fallback within country
            try:
                df_cc = df_geo_cities[df_geo_cities.get(cc_col, pd.Series()).astype(str).str.upper().eq(cc_u)]
                if not df_cc.empty:
                    df_cc = df_cc.copy()
                    df_cc['__canon_city'] = df_cc.get(city_col, pd.Series()).astype(str).map(_canon)
                    key = _canon(city)
                    m3 = df_cc[df_cc['__canon_city'] == key]
                    if not m3.empty:
                        la = m3.iloc[0].get(lat_col); lo = m3.iloc[0].get(lon_col)
                        if pd.notna(la) and pd.notna(lo):
                            return float(la), float(lo)
                    choices = df_cc.get(city_col, pd.Series()).dropna().astype(str).tolist()
                    best, score = _fuzzy_pick(city, choices, score_cutoff=90)
                    if best:
                        m4 = df_cc[df_cc.get(city_col, pd.Series()).astype(str).str.upper() == str(best).upper()]
                        if not m4.empty:
                            la = m4.iloc[0].get(lat_col); lo = m4.iloc[0].get(lon_col)
                            if pd.notna(la) and pd.notna(lo):
                                return float(la), float(lo)
            except Exception:
                pass
        except Exception:
            return None, None
        return None, None

    def _city_coords_online(cc: str, city: str) -> tuple[float | None, float | None, bool]:
        """Online geocoding fallback using Nominatim (geopy) when allowed.
        Returns (lat, lon, used) where used indicates if online lookup succeeded.
        """
        try:
            if geocode_city_online_if_allowed is None:
                return None, None, False
            if not city:
                return None, None, False
            la, lo, src = geocode_city_online_if_allowed(cc, city)
            if la is not None and lo is not None:
                return la, lo, True
        except Exception:
            pass
        return None, None, False

    def _zip_coords_from_db(cc: str, z: str) -> tuple[float | None, float | None]:
        """Lookup ZIP coordinates from ZIP_COORDS/ZIP_COORDINATES when available."""
        try:
            cc_u = (cc or "").strip().upper()
            z_u = _normalize_zip_token(z)
            if not z_u:
                return None, None
            # Manual fallback requested by user
            if cc_u == "CN" and z_u == "322000":
                return 29.3151, 120.0768
            if cc_u == "CN" and z_u == "430000":
                return 30.5928, 114.3055
            if cc_u == "MA" and z_u == "90010":
                return 35.7595, -5.8340
            if cc_u == "CZ" and z_u == "74401":
                return 49.5489, 18.2108

            if 'df_zip_coords' not in locals() or df_zip_coords is None or df_zip_coords.empty or not z:
                return None, None
            cols = {c.lower().strip(): c for c in df_zip_coords.columns}
            cc_col = cols.get("country code") or cols.get("cc") or "Country Code"
            zip_col = cols.get("zip") or cols.get("zip code") or cols.get("postal code") or "ZIP"
            lat_col = cols.get("lat") or cols.get("latitude") or "Lat"
            lon_col = cols.get("lon") or cols.get("long") or cols.get("longitude") or "Long"
            m = df_zip_coords[
                df_zip_coords.get(cc_col, pd.Series()).astype(str).str.upper().eq(cc_u) &
                df_zip_coords.get(zip_col, pd.Series()).astype(str).str.upper().apply(_normalize_zip_token).eq(z_u)
            ]
            if not m.empty:
                la = m.iloc[0].get(lat_col); lo = m.iloc[0].get(lon_col)
                if pd.notna(la) and pd.notna(lo):
                    return float(la), float(lo)
        except Exception:
            return None, None
        return None, None

    def get_domestic_eur_per_km(cc: str):
        # Normalize common aliases used across files (e.g. GB in input vs UK in COSTPERKM)
        alias = {
            "GB": "UK",
            "UK": "UK",
            "EL": "GR",
            "TK": "TR",
        }
        cc_u = str(cc or "").strip().upper()
        cc_norm = alias.get(cc_u, cc_u)
        try:
            co = df_cpkm["Country of origin"].astype(str).str.upper().str.strip()
            cd = df_cpkm["Destination Country"].astype(str).str.upper().str.strip()
            m = df_cpkm[(co == cc_norm) & (cd == cc_norm)]
            if not m.empty:
                val = m.iloc[0].get("Eur/km")
                return float(val) if pd.notna(val) else None
        except Exception:
            return None
        return None

    def get_pair_eur_per_km(oc: str, dc: str):
        """Lookup €/km for a country pair. Try directional oc->dc, then symmetric dc->oc, then domestic if same country."""
        alias = {
            "GB": "UK",
            "UK": "UK",
            "EL": "GR",
            "TK": "TR",
        }
        oc_u = str(oc or "").strip().upper()
        dc_u = str(dc or "").strip().upper()
        oc_n = alias.get(oc_u, oc_u)
        dc_n = alias.get(dc_u, dc_u)
        try:
            co = df_cpkm["Country of origin"].astype(str).str.upper().str.strip()
            cd = df_cpkm["Destination Country"].astype(str).str.upper().str.strip()
            m = df_cpkm[(co == oc_n) & (cd == dc_n)]
            if not m.empty:
                val = m.iloc[0].get("Eur/km")
                if pd.notna(val):
                    return float(val)
            # symmetric fallback
            m2 = df_cpkm[(co == dc_n) & (cd == oc_n)]
            if not m2.empty:
                val2 = m2.iloc[0].get("Eur/km")
                if pd.notna(val2):
                    return float(val2)
            if oc_n == dc_n:
                return get_domestic_eur_per_km(oc_n)
        except Exception:
            return None
        return None

    def get_pair_tt_road(oc: str, dc: str):
        """Lookup inland transit time (TT_ROAD) by country pair, with symmetric fallback."""
        alias = {
            "GB": "UK",
            "UK": "UK",
            "EL": "GR",
            "TK": "TR",
        }
        oc_u = str(oc or "").strip().upper()
        dc_u = str(dc or "").strip().upper()
        oc_n = alias.get(oc_u, oc_u)
        dc_n = alias.get(dc_u, dc_u)
        try:
            co = df_cpkm["Country of origin"].astype(str).str.upper().str.strip()
            cd = df_cpkm["Destination Country"].astype(str).str.upper().str.strip()
            m = df_cpkm[(co == oc_n) & (cd == dc_n)]
            if not m.empty:
                val = m.iloc[0].get("TT_ROAD")
                if pd.notna(val):
                    return float(val)
            m2 = df_cpkm[(co == dc_n) & (cd == oc_n)]
            if not m2.empty:
                val2 = m2.iloc[0].get("TT_ROAD")
                if pd.notna(val2):
                    return float(val2)
        except Exception:
            return None
        return None

    PLANT_ALIASES = {
        "HORSE BRASIL": "HORSE BRAZIL",
        "HORSE BRAZIL": "HORSE BRAZIL",
        "HORSE VALLADOLID": "HORSE MOTORES",
        "HORSE MOTORES": "HORSE MOTORES",
        "WEST HORSE POWERTRAIN PORTUGAL": "HORSE CACIA",
        "HORSE CACIA": "HORSE CACIA",
    }

    def canonical_plant_name(name: str | None) -> str:
        raw = str(name or "").strip().upper()
        if not raw:
            return ""
        return PLANT_ALIASES.get(raw, raw)

    def get_hp_eur_per_km(plant_name: str, port_code: str):
        """Lookup €/km from HORSE-PUERTO using Plant + POL/POD (column K / Eur/km)."""
        try:
            if df_hp is None or df_hp.empty or not plant_name or not port_code:
                return None
            plant_name = canonical_plant_name(plant_name)
            cols = {str(c).lower().strip(): str(c) for c in df_hp.columns}
            plant_col = cols.get("plant") or "Plant"
            port_col = cols.get("pol/pod") or "POL/POD"
            eur_col = cols.get("eur/km")
            # Fallback to column K (index 10) if Eur/km header is unavailable
            if eur_col is None and len(df_hp.columns) > 10:
                eur_col = str(df_hp.columns[10])
            if eur_col is None:
                return None

            m = df_hp[
                df_hp.get(plant_col, pd.Series()).astype(str).str.strip().str.upper().eq(str(plant_name).strip().upper()) &
                df_hp.get(port_col, pd.Series()).astype(str).str.strip().str.upper().eq(str(port_code).strip().upper())
            ]
            if not m.empty:
                val = m.iloc[0].get(eur_col)
                if pd.notna(val):
                    return float(val)
        except Exception:
            return None
        return None

    # Optional local GEO index to compute dynamic distances when possible
    # Build from Ports Locations (PORT) and HORSE-PUERTO (PLANT) when available
    geo = None
    try:
        geo_rows = []
        # Plant coordinates from HORSE-PUERTO
        if df_hp is not None and not df_hp.empty:
            for _, rr in df_hp.iterrows():
                try:
                    cc = str(rr.get("Country Code", "")).strip().upper()
                    plant = str(rr.get("Plant", "")).strip()
                    lat = rr.get("Plant Lat"); lon = rr.get("Plant Long")
                    if cc and plant and pd.notna(lat) and pd.notna(lon):
                        geo_rows.append({
                            "type": "PLANT",
                            "country_code": cc,
                            "key": plant.upper(),
                            "lat": float(lat),
                            "lon": float(lon),
                        })
                except Exception:
                    pass
        # Ports coordinates from Ports Locations
        if df_ports is not None and not df_ports.empty:
            for _, rr in df_ports.iterrows():
                try:
                    # Country might be name or code; try to coerce using the helper below (defined later)
                    ctry = (rr.get("Country") if pd.notna(rr.get("Country")) else "")
                    cc = str(ctry).strip().upper()
                    if len(cc) not in (2, 3):
                        # Resolve via HORSE-PUERTO mapping if available
                        # We'll normalize later once coerce_country_code is defined
                        pass
                    code = str(rr.get("POL/POD", "")).strip().upper()
                    lat = rr.get("LAT"); lon = rr.get("LONG")
                    if code and pd.notna(lat) and pd.notna(lon):
                        geo_rows.append({
                            "type": "PORT",
                            "country_code": cc,
                            "key": code,
                            "lat": float(lat),
                            "lon": float(lon),
                        })
                except Exception:
                    pass
        if geo_rows:
            geo_df = pd.DataFrame(geo_rows)
            # Try to normalize country codes using HORSE-PUERTO mapping where country_code seems like a name
            if not geo_df.empty:
                # We will normalize below after we build country_name_to_code
                pass
    except Exception:
        geo = None

    # Map Country (name) -> Country Code using HORSE-PUERTO to recover missing codes
    country_name_to_code: dict[str, str] = {}
    try:
        if not df_hp.empty and "Country" in df_hp.columns and "Country Code" in df_hp.columns:
            df_tmp = df_hp[["Country", "Country Code"]].dropna()
            for _, rr in df_tmp.iterrows():
                name = str(rr.get("Country", "")).strip().upper()
                code = str(rr.get("Country Code", "")).strip().upper()
                if name and code:
                    country_name_to_code[name] = code
    except Exception:
        pass

    # Minimal built-in fallback for common country names -> ISO2
    FALLBACK_COUNTRY_NAME_TO_ISO2 = {
        "CHINA": "CN",
        "INDIA": "IN",
        "SPAIN": "ES",
        "BRAZIL": "BR",
        "TURKEY": "TR",
        "FRANCE": "FR",
        "GERMANY": "DE",
        "ALLEMAGNE": "DE",
        "ALEMANIA": "DE",
        "DEUTSCHLAND": "DE",
        "UNITED STATES": "US",
        "USA": "US",
        "MEXICO": "MX",
        "UNITED KINGDOM": "GB",
        "UK": "GB",
        "ITALY": "IT",
        "PORTUGAL": "PT",
        "NETHERLANDS": "NL",
        "POLAND": "PL",
        "CZECH REPUBLIC": "CZ",
        "CZECHIA": "CZ",
        "CZECH": "CZ",
        "SLOVAKIA": "SK",
        "SOUTH KOREA": "KR",
        "KOREA, REPUBLIC OF": "KR",
        "JAPAN": "JP",
        "MOROCCO": "MA",
        "MAROC": "MA",
        "MARRUECOS": "MA",
        "MAROKKO": "MA",
    }

    def coerce_country_code(raw_code: str | None, raw_name: str | None) -> str:
        code = (raw_code or "").strip().upper()
        if len(code) in (2, 3):
            return code
        name = (raw_name or "").strip().upper()
        # First try HORSE-PUERTO mapping
        mapped = country_name_to_code.get(name)
        if mapped:
            return mapped
        # Then try built-in fallback map
        fb = FALLBACK_COUNTRY_NAME_TO_ISO2.get(name)
        if fb:
            return fb
        # As last resort, return given code/name (may be non-ISO)
        return code or name

    EUROPEAN_COUNTRY_CODES = {
        "AL", "AD", "AT", "BA", "BE", "BG", "BY", "CH", "CY", "CZ", "DE", "DK", "EE", "ES", "FI",
        "FO", "FR", "GB", "GI", "GR", "HR", "HU", "IE", "IS", "IT", "LI", "LT", "LU", "LV", "MC",
        "MD", "ME", "MK", "MT", "NL", "NO", "PL", "PT", "RO", "RS", "SE", "SI", "SK", "SM", "UA",
        "VA", "XK",
    }

    def is_morocco_europe_roro_route(origin_cc: str, dest_cc: str) -> bool:
        return (origin_cc or "").strip().upper() == "MA" and (dest_cc or "").strip().upper() in EUROPEAN_COUNTRY_CODES

    # Normalize missing/ambiguous country codes in geo_df (if created)
    try:
        if 'geo_df' in locals() and geo_df is not None and not geo_df.empty:
            def _norm_code(cc_or_name: str) -> str:
                c = (cc_or_name or '').strip().upper()
                if len(c) in (2, 3):
                    return c
                return country_name_to_code.get(c, c)
            geo_df['country_code'] = geo_df['country_code'].apply(_norm_code)
            geo = GeoIndex(geo_df)
        else:
            geo = GeoIndex.load_from_dir(QTOOL_DIR)
    except Exception:
        geo = GeoIndex.load_from_dir(QTOOL_DIR)

    # Build quote rows
    rows = []
    for _, r in input_df.iterrows():
        # Per-row incoterm (fallback to default)
        incoterm_row = (r.get("incoterm") or "").strip().upper() if pd.notna(r.get("incoterm")) else DEFAULT_INCOTERM
        debug_msgs = []
        try:
            included_legs, type_of_flow = flow_by_incoterm(incoterm_row)
        except Exception as e:
            debug_msgs.append(f"Incoterm no soportado '{r.get('incoterm')}', usando {DEFAULT_INCOTERM}")
            incoterm_row = DEFAULT_INCOTERM
            included_legs, type_of_flow = flow_by_incoterm(incoterm_row)
        pn = r.get("pn"); designation = r.get("designation"); supplier = r.get("supplier_plant")
        packaging_code_raw = r.get("packaging_code")
        packaging_code = resolve_packaging_code(packaging_code_raw)
        pkg_data = lookup_packaging_data(str(pn or ""), packaging_code_raw, packaging_code)
        if pkg_data.get("pkg_debug"):
            debug_msgs.append(str(pkg_data.get("pkg_debug")))
        oc = coerce_country_code(r.get("origin_country_code"), r.get("origin_country"))
        dc = coerce_country_code(r.get("dest_country_code"), r.get("dest_country"))
        # Business rule: Morocco -> Europe RoRo moves are handled as Inland transport.
        if is_morocco_europe_roro_route(oc, dc):
            type_of_flow = "Inland"
            included_legs = [1]
            debug_msgs.append("Ruta especial MA->Europa por RoRo: tratada como Inland")
        # If flow is Inland: per requirement, compute ONLY Leg1 (road from origin country to destination country)
        elif str(type_of_flow).strip().upper() == "INLAND":
            included_legs = [1]
        pack_per_container = calc_pack_per_container(type_of_flow, pkg_data)
        dest_plant = r.get("dest_plant")
        supplier_canon = canonical_plant_name(str(supplier or ""))
        dest_plant_canon = canonical_plant_name(str(dest_plant or ""))
        origin_city = str(r.get("origin_city")) if pd.notna(r.get("origin_city")) else ""
        origin_zip = str(r.get("origin_zip")) if pd.notna(r.get("origin_zip")) else ""
        dest_city = str(r.get("dest_city")) if pd.notna(r.get("dest_city")) else ""
        dest_zip = str(r.get("dest_zip")) if pd.notna(r.get("dest_zip")) else ""
        # Determine ports whenever flow is Overseas (even if buyer doesn't pay leg 2)
        pol_distance_km = None
        pod_distance_km = None
        if str(type_of_flow).strip().upper() == "OVERSEAS":
            # Compute origin point (parse and enrich ZIP first)
            oc_city_clean_pre, oc_zip_enriched_pre, parse_note_pre = _parse_city_zip(origin_city, origin_zip)
            if parse_note_pre:
                debug_msgs.append(parse_note_pre)
            ozip_final_pre, zip_reason_pre = validate_and_enrich_zip(oc, oc_city_clean_pre, oc_zip_enriched_pre)
            if zip_reason_pre:
                debug_msgs.append(f"ZIP origen corregido por {zip_reason_pre}: {oc_zip_enriched_pre or '-'}→{ozip_final_pre}")
            oc_city_norm_pre = _normalize_city_for_country(oc, oc_city_clean_pre)
            o_lat, o_lon, _ = resolve_point(geo, oc, zip_code=ozip_final_pre, city=oc_city_norm_pre, plant=supplier_canon)
            # ZIP_COORDS fallback
            if o_lat is None and ozip_final_pre:
                la, lo = _zip_coords_from_db(oc, ozip_final_pre)
                if la is not None:
                    o_lat, o_lon = la, lo
                    debug_msgs.append("Origen resuelto por zip_coords_db")
            if o_lat is None:
                # Fallback B: DB city coordinates (GEO_LOCATIONS/CITY_COORDS)
                o_lat2, o_lon2 = _city_coords_from_db(oc, oc_city_clean_pre)
                if o_lat2 is not None:
                    o_lat, o_lon = o_lat2, o_lon2
                    debug_msgs.append("Origen resuelto por geo_city_fallback")
            if o_lat is None:
                # Fallback C: Online (Nominatim) if allowed
                la, lo, used = _city_coords_online(oc, oc_city_clean_pre)
                if used and la is not None:
                    o_lat, o_lon = la, lo
                    debug_msgs.append("Origen resuelto por nominatim (online)")
            origin_point = (o_lat, o_lon) if o_lat is not None else None

            # Compute destination point early (for POD proximity)
            dc_city_clean_pre, dc_zip_enriched_pre, parse_note_pre2 = _parse_city_zip(dest_city, dest_zip)
            if parse_note_pre2:
                debug_msgs.append(parse_note_pre2)
            dzip_final_pre, dzip_reason_pre = validate_and_enrich_zip(dc, dc_city_clean_pre, dc_zip_enriched_pre)
            if dzip_reason_pre:
                debug_msgs.append(f"ZIP destino corregido por {dzip_reason_pre}: {dc_zip_enriched_pre or '-'}→{dzip_final_pre}")
            dc_city_norm_pre = _normalize_city_for_country(dc, dc_city_clean_pre)
            t_lat_pre, t_lon_pre, _ = resolve_point(geo, dc, zip_code=dzip_final_pre, city=dc_city_norm_pre, plant=dest_plant_canon)
            # ZIP_COORDS fallback
            if t_lat_pre is None and dzip_final_pre:
                la, lo = _zip_coords_from_db(dc, dzip_final_pre)
                if la is not None:
                    t_lat_pre, t_lon_pre = la, lo
                    debug_msgs.append("Destino resuelto por zip_coords_db")
            if t_lat_pre is None:
                # Fallback B: DB city coordinates
                t_lat2, t_lon2 = _city_coords_from_db(dc, dc_city_clean_pre)
                if t_lat2 is not None:
                    t_lat_pre, t_lon_pre = t_lat2, t_lon2
                    debug_msgs.append("Destino resuelto por geo_city_fallback")
            if t_lat_pre is None:
                la, lo, used = _city_coords_online(dc, dc_city_clean_pre)
                if used and la is not None:
                    t_lat_pre, t_lon_pre = la, lo
                    debug_msgs.append("Destino resuelto por nominatim (online)")
            dest_point = (t_lat_pre, t_lon_pre) if t_lat_pre is not None else None

            # Select POL/POD from Ports Locations by proximity to origin/destination
            pol_cands_debug = _ports_candidates_debug(oc, origin_point)
            pod_cands_debug = _ports_candidates_debug(dc, dest_point)
            pol, why_pol = select_port_nearest_from_ports_locations(oc, origin_point, side="POL")
            pod, why_pod = select_port_nearest_from_ports_locations(dc, dest_point, side="POD")
            if pol:
                debug_msgs.append(f"POL elegido por {why_pol}: {pol}")
            else:
                debug_msgs.append("Falta POL (Ports Locations no tiene candidatos para el país de origen)")
            if pod:
                debug_msgs.append(f"POD elegido por {why_pod}: {pod}")
            else:
                debug_msgs.append("Falta POD (Ports Locations no tiene candidatos para el país de destino)")
            # Add top-3 proximity candidates to Debug for traceability
            if pol_cands_debug:
                top = ", ".join([f"{d['code']}:{'?' if d['km'] is None else d['km']}km" for d in pol_cands_debug[:3]])
                debug_msgs.append(f"POL candidatos={len(pol_cands_debug)} top3[{top}]")
                if origin_point is None:
                    debug_msgs.append("Aviso: origen sin coordenadas → no se pueden calcular km de cercanía (agrega ZIP en CITY_ZIPS o coords de ciudad en GEO_LOCATIONS/CITY_COORDS)")
            if pod_cands_debug:
                top = ", ".join([f"{d['code']}:{'?' if d['km'] is None else d['km']}km" for d in pod_cands_debug[:3]])
                debug_msgs.append(f"POD candidatos={len(pod_cands_debug)} top3[{top}]")
                if dest_point is None:
                    debug_msgs.append("Aviso: destino sin coordenadas → no se pueden calcular km de cercanía (agrega ZIP en CITY_ZIPS o coords de ciudad en GEO_LOCATIONS/CITY_COORDS)")

            # Now, lookup Rate and TT for the chosen POL/POD pair
            ocean_rate, tt_days = get_ocean_rate_and_tt(pol, pod) if pol and pod else (None, None)
            # Compute POL/POD distances for audit if we have endpoints
            try:
                if pol and origin_point is not None:
                    plat_a, plon_a = _port_coords_for_distance(oc, pol, origin_point)
                    if plat_a is not None and plon_a is not None:
                        pol_distance_km = road_km_between((plat_a, plon_a), origin_point)
                if pod and dest_point is not None:
                    dlat_a, dlon_a = _port_coords_for_distance(dc, pod, dest_point)
                    if dlat_a is not None and dlon_a is not None:
                        pod_distance_km = road_km_between((dlat_a, dlon_a), dest_point)
            except Exception:
                pass

            if pol and pod and ocean_rate is None:
                debug_msgs.append("Falta tarifa leg2 (ocean) para el par en MAIN PORTS")
            if pol and pod and tt_days is None:
                debug_msgs.append("Falta TT leg2 para el par en VTT DATA / MAIN PORTS / TRANSITTIME")
        else:
            pol = ""
            pod = ""
            ocean_rate, tt_days = None, None
        # €/km lookups + Transit Time by flow
        if str(type_of_flow).strip().upper() == "INLAND":
            eurpkm_leg1 = get_pair_eur_per_km(oc, dc)
            eurpkm_leg3 = None
            transit_time_days = get_pair_tt_road(oc, dc)
        else:
            eurpkm_leg1 = get_domestic_eur_per_km(oc)
            eurpkm_leg3 = get_domestic_eur_per_km(dc)
            transit_time_days = tt_days

            # Business rule: for CIF/FOB/FCA maritime routes with POL/POD,
            # prefer HORSE-PUERTO Eur/km (Plant + POL/POD), but keep COSTPERKM
            # as fallback when the plant/port combination is missing.
            if incoterm_row in {"CIF", "FOB", "FCA"} and pol and pod:
                hp_leg1 = get_hp_eur_per_km(supplier_canon, str(pol))
                hp_leg3 = get_hp_eur_per_km(dest_plant_canon, str(pod))
                if hp_leg1 is None:
                    if eurpkm_leg1 is not None:
                        debug_msgs.append("Falta Eur/km HORSE-PUERTO para leg1 (Plant+POL); se usa COSTPERKM")
                    else:
                        debug_msgs.append("Falta Eur/km HORSE-PUERTO para leg1 (Plant+POL)")
                else:
                    eurpkm_leg1 = hp_leg1
                if hp_leg3 is None:
                    if eurpkm_leg3 is not None:
                        debug_msgs.append("Falta Eur/km HORSE-PUERTO para leg3 (Plant+POD); se usa COSTPERKM")
                    else:
                        debug_msgs.append("Falta Eur/km HORSE-PUERTO para leg3 (Plant+POD)")
                else:
                    eurpkm_leg3 = hp_leg3

        # KM computation (fully dynamic via coordinates)
        leg1_km = None
        leg3_km = None
        if str(type_of_flow).strip().upper() == "INLAND":
            # Road from origin location to destination location
            # Validate/enrich ZIPs from City+Country if needed
            ozip, oreason = validate_and_enrich_zip(oc, origin_city, origin_zip)
            if oreason:
                debug_msgs.append(f"ZIP origen corregido por {oreason}: {origin_zip}→{ozip}")
            dzip, dreason = validate_and_enrich_zip(dc, dest_city, dest_zip)
            if dreason:
                debug_msgs.append(f"ZIP destino corregido por {dreason}: {dest_zip}→{dzip}")
            olat, olon, osrc = resolve_point(geo, oc, zip_code=ozip, city=origin_city, plant=supplier_canon)
            dlat, dlon, dsrc = resolve_point(geo, dc, zip_code=dzip, city=dest_city, plant=dest_plant_canon)
            if (olat is None or olon is None) and ozip:
                la, lo = _zip_coords_from_db(oc, ozip)
                if la is not None:
                    olat, olon, osrc = la, lo, "zip_coords_db"
            if olat is None and origin_city:
                la, lo = _city_coords_from_db(oc, origin_city)
                if la is not None:
                    olat, olon, osrc = la, lo, "geo_city_fallback"
            if (dlat is None or dlon is None) and dzip:
                la, lo = _zip_coords_from_db(dc, dzip)
                if la is not None:
                    dlat, dlon, dsrc = la, lo, "zip_coords_db"
            if dlat is None and dest_city:
                la, lo = _city_coords_from_db(dc, dest_city)
                if la is not None:
                    dlat, dlon, dsrc = la, lo, "geo_city_fallback"
            if olat is not None and dlat is not None:
                leg1_km = road_km_between((olat, olon), (dlat, dlon))
                debug_msgs.append(f"Leg1 km dinámico ({osrc}→{dsrc})")
            else:
                debug_msgs.append("No se pudo calcular Leg1: faltan coordenadas (origen/destino)")
        else:
            # Overseas
            # Leg1: origin -> POL
            oc_city_clean, oc_zip_enriched, parse_note = _parse_city_zip(origin_city, origin_zip)
            if parse_note:
                debug_msgs.append(parse_note)
            oc_city_norm = _normalize_city_for_country(oc, oc_city_clean)
            # Validate/enrich ZIP for origin using city+country if invalid/missing
            oc_zip_final, zip_reason = validate_and_enrich_zip(oc, oc_city_clean, oc_zip_enriched)
            if zip_reason:
                debug_msgs.append(f"ZIP origen corregido por {zip_reason}: {oc_zip_enriched or '-'}→{oc_zip_final}")

            # Robust resolver for origin: try ZIP → cleaned city → supplier plant
            olat = olon = None
            osrc = ""
            tried = []
            if oc_zip_final:
                lat, lon, src = resolve_point(geo, oc, zip_code=oc_zip_final)
                tried.append(f"zip:{oc_zip_final}")
                # Accept only if truly resolved by ZIP, not country fallback
                if lat is not None and src.startswith("zip:"):
                    olat, olon, osrc = lat, lon, src
                elif olat is None:
                    la, lo = _zip_coords_from_db(oc, oc_zip_final)
                    if la is not None:
                        olat, olon, osrc = la, lo, "zip_coords_db"
            if olat is None and oc_city_norm:
                lat, lon, src = resolve_point(geo, oc, city=oc_city_norm)
                tried.append(f"city:{oc_city_norm}")
                if lat is not None and src.startswith("city:"):
                    olat, olon, osrc = lat, lon, src
            # City alias fallback from database (e.g., Mundhwa -> Pune)
            if olat is None and oc_city_norm:
                alias = _city_alias(oc, oc_city_norm)
                if alias:
                    lat, lon, src = resolve_point(geo, oc, city=alias)
                    tried.append(f"city-alias:{oc_city_norm}->{alias}")
                    if lat is not None and src.startswith("city:"):
                        olat, olon, osrc = lat, lon, src + "(alias)"
            if olat is None and supplier_canon:
                lat, lon, src = resolve_point(geo, oc, plant=supplier_canon)
                tried.append(f"plant:{supplier_canon}")
                if lat is not None and src.startswith("plant:"):
                    olat, olon, osrc = lat, lon, src
            # No HORSE-PUERTO fallback for city-level geocoding: use database city coords only
            # Last resort: city coordinates from database sheet (GEO_LOCATIONS/CITY_COORDS)
            if olat is None and oc_city_norm:
                try:
                    if df_geo_cities is not None and not df_geo_cities.empty:
                        cc_u = (oc or "").strip().upper()
                        key_u = oc_city_norm.strip().upper()
                        cols = {c.lower().strip(): c for c in df_geo_cities.columns}
                        cc_col = cols.get("country code") or cols.get("cc") or "Country Code"
                        city_col = cols.get("city") or "City"
                        lat_col = cols.get("lat") or cols.get("latitude") or "Lat"
                        lon_col = cols.get("lon") or cols.get("long") or cols.get("longitude") or "Long"
                        m = df_geo_cities[
                            df_geo_cities.get(cc_col, pd.Series()).astype(str).str.upper().eq(cc_u) &
                            df_geo_cities.get(city_col, pd.Series()).astype(str).str.strip().str.upper().eq(key_u)
                        ]
                        if not m.empty:
                            la = m.iloc[0].get(lat_col); lo = m.iloc[0].get(lon_col)
                            if pd.notna(la) and pd.notna(lo):
                                olat, olon = float(la), float(lo)
                                osrc = "geo_city"
                        if olat is None:
                            # Try alias via CITY_ALIASES sheet
                            alias = _city_alias(oc, oc_city_clean)
                            if alias:
                                m2 = df_geo_cities[
                                    df_geo_cities.get(cc_col, pd.Series()).astype(str).str.upper().eq(cc_u) &
                                    df_geo_cities.get(city_col, pd.Series()).astype(str).str.strip().str.upper().eq(alias.strip().upper())
                                ]
                                if not m2.empty:
                                    la = m2.iloc[0].get(lat_col); lo = m2.iloc[0].get(lon_col)
                                    if pd.notna(la) and pd.notna(lo):
                                        olat, olon = float(la), float(lo)
                                        osrc = "geo_city(alias)"
                except Exception:
                    pass
            if olat is None and oc_city_norm:
                # Last attempt: Online (Nominatim) if allowed
                la, lo, used = _city_coords_online(oc, oc_city_norm)
                if used and la is not None:
                    olat, olon, osrc = la, lo, "nominatim"
            if olat is None:
                debug_msgs.append("Origen no resuelto: intentos " + ", ".join(tried) if tried else "Origen no resuelto (sin datos)")

            # POL point
            plat, plon, psrc = resolve_point(geo, oc, port=pol) if pol else (None, None, "")
            if plat is None and pol:
                # Fallback: use coordinates directly from Ports Locations regardless of country filter
                f_lat, f_lon = _port_coords_from_ports_locations(pol)
                if f_lat is not None:
                    b_lat, b_lon, tag = _best_port_coords_for_origin((olat, olon) if olat is not None else None, f_lat, f_lon)
                    plat, plon, psrc = b_lat, b_lon, tag
            if olat is not None and plat is not None:
                leg1_km = road_km_between((olat, olon), (plat, plon))
                # Add brief computation hint (approx haversine*1.30)
                try:
                    approx_geo = float(leg1_km) / 1.30
                    debug_msgs.append(f"Leg1 km dinámico ({osrc}→{psrc}) ≈ {approx_geo:.0f}*1.30")
                except Exception:
                    debug_msgs.append(f"Leg1 km dinámico ({osrc}→{psrc})")
            else:
                if olat is None and plat is None:
                    debug_msgs.append("No se pudo calcular Leg1: faltan coordenadas de origen y POL")
                elif olat is None:
                    debug_msgs.append("No se pudo calcular Leg1: falta coordenada de origen")
                else:
                    debug_msgs.append("No se pudo calcular Leg1: falta coordenada de POL")
            # Leg3: POD -> destination plant
            dplat, dplon, dpsrc = resolve_point(geo, dc, port=pod) if pod else (None, None, "")
            if dplat is None and pod:
                f_lat, f_lon = _port_coords_from_ports_locations(pod)
                if f_lat is not None:
                    b_lat, b_lon, tag = _best_port_coords_for_origin((tlat, tlon) if 'tlat' in locals() and tlat is not None else None, f_lat, f_lon)
                    dplat, dplon, dpsrc = b_lat, b_lon, tag
            # Parse and validate/enrich destination ZIP similarly
            dc_city_clean, dc_zip_parsed, _ = _parse_city_zip(dest_city, dest_zip)
            dzip_final, dzip_reason = validate_and_enrich_zip(dc, dc_city_clean, dc_zip_parsed)
            if dzip_reason:
                debug_msgs.append(f"ZIP destino corregido por {dzip_reason}: {dc_zip_parsed or '-'}→{dzip_final}")
            tlat, tlon, tsrc = resolve_point(geo, dc, zip_code=dzip_final, city=dc_city_clean, plant=dest_plant_canon)
            if (tlat is None or tlon is None) and dzip_final:
                la, lo = _zip_coords_from_db(dc, dzip_final)
                if la is not None:
                    tlat, tlon, tsrc = la, lo, "zip_coords_db"
            if tlat is None and dc_city_clean:
                la, lo, used = _city_coords_online(dc, dc_city_clean)
                if used and la is not None:
                    tlat, tlon, tsrc = la, lo, "nominatim"
            if dplat is not None and tlat is not None:
                leg3_km = road_km_between((dplat, dplon), (tlat, tlon))
                debug_msgs.append(f"Leg3 km dinámico ({dpsrc}→{tsrc})")
            else:
                debug_msgs.append("No se pudo calcular Leg3: faltan coordenadas (POD/destino)")
        # Special: For DAP, compute origin → destination plant distance for visibility, even if buyer pays 0
        if incoterm_row == 'DAP':
            try:
                olat, olon, osrc = resolve_point(geo, oc, zip_code=origin_zip, city=origin_city, plant=supplier_canon)
                tlat, tlon, tsrc = resolve_point(geo, dc, zip_code=dest_zip, city=dest_city, plant=dest_plant_canon)
                if olat is not None and tlat is not None:
                    dap_km = road_km_between((olat, olon), (tlat, tlon))
                    debug_msgs.append(f"DAP distancia origen→destino ({osrc}→{tsrc}): {dap_km:.1f} km")
            except Exception:
                pass
        if 1 in included_legs and eurpkm_leg1 is None:
            debug_msgs.append("Falta €/km leg1 (origen)")
        if 3 in included_legs and eurpkm_leg3 is None:
            debug_msgs.append("Falta €/km leg3 (destino)")
        # Keep Leg 3 cost aligned with the visible "LEG3/POD Distance (km)" output.
        leg3_cost_km = pod_distance_km if pod_distance_km is not None else leg3_km
        leg1_cost = (eurpkm_leg1 * leg1_km) if (1 in included_legs and eurpkm_leg1 and leg1_km) else (0.0 if 1 in included_legs else None)
        leg2_cost = ocean_rate if 2 in included_legs else None
        leg3_cost = (eurpkm_leg3 * leg3_cost_km) if (3 in included_legs and eurpkm_leg3 and leg3_cost_km) else (0.0 if 3 in included_legs else None)
        if 2 in included_legs and ocean_rate is None:
            debug_msgs.append("Falta tarifa leg2 (ocean)")
        # Mostrar TT si hay par POL/POD aunque el comprador no pague leg 2
        if str(type_of_flow).strip().upper() == "OVERSEAS" and tt_days is None:
            debug_msgs.append("Falta TT leg2")
        # Incoterm suggestions
        try:
            if oc and dc and oc == dc:
                debug_msgs.append("Sugerencia: DAP (mismo país, tráfico inland)")
            if 2 in included_legs and (ocean_rate is None or not pol or not pod):
                debug_msgs.append("Sugerencia: Revisar Incoterm; si el comprador gestiona el flete principal use FOB; si el vendedor lo asume hasta puerto destino use CIF")
            if 1 in included_legs and eurpkm_leg1 is None:
                debug_msgs.append("Sugerencia: Considera un Incoterm que no incluya leg 1 (ej. FOB)")
            if 3 in included_legs and eurpkm_leg3 is None:
                debug_msgs.append("Sugerencia: Considera un Incoterm que no incluya leg 3 (ej. CIF)")
        except Exception:
            pass
        total = 0.0
        for val in (leg1_cost, leg2_cost, leg3_cost):
            if isinstance(val, (int, float)) and val is not None:
                total += float(val)
        rows.append({
            "pn": pn,
            "designation": designation,
            "supplier_plant": supplier,
            "incoterm": incoterm_row,
            "type_of_flow": type_of_flow,
            "origin_cc": oc,
            "dest_cc": dc,
            "dest_plant": dest_plant,
            "POL": pol,
            "POD": pod,
            "pol_distance_km": pol_distance_km,
            "pod_distance_km": pod_distance_km,
            "leg1_eur_per_km": eurpkm_leg1 if 1 in included_legs else None,
            "leg1_km": leg1_km,
            "leg1_cost_eur": leg1_cost,
            "leg2_ocean_rate_eur": ocean_rate if 2 in included_legs else None,
            # Publicar TT cuando el flujo es Overseas (aunque el comprador no pague leg2)
            "leg2_tt_days": tt_days if str(type_of_flow).strip().upper() == "OVERSEAS" else None,
            "transit_time_days": transit_time_days,
            "leg3_eur_per_km": eurpkm_leg3 if 3 in included_legs else None,
            "leg3_km": leg3_cost_km,
            "leg3_cost_eur": leg3_cost,
            "total_cost_eur": total,
            "Red flag/Debug": "; ".join(debug_msgs) if debug_msgs else "",
            "packaging_code_resolved": packaging_code,
            "pkg_volume_m3": pkg_data["pkg_volume_m3"],
            "pkg_snp": pkg_data["pkg_snp"],
            "pkg_weight_part": pkg_data["pkg_weight_part"],
            "pkg_weight_empty": pkg_data["pkg_weight_empty"],
            "pkg_weight_full": pkg_data["pkg_weight_full"],
            "pack_per_cont_40ft": pack_per_container,
            "notes": "Distancias dinámicas: puertos desde Ports Locations; origen/destino desde CITY_ZIPS y GEO_LOCATIONS/CITY_COORDS. Si falta coordenada, la distancia queda vacía."
        })

    quote_df = pd.DataFrame(rows)

    # Build final Quote sheet preserving original column names/order, appending computable fields
    inv_map = {v: k for k, v in STD_COLS.items()}
    raw_like_df = input_df.rename(columns={c: inv_map.get(c, c) for c in input_df.columns})
    # Ensure all original columns exist (we'll order by reference columns if present)
    for c in STD_COLS.keys():
        if c not in raw_like_df.columns:
            raw_like_df[c] = ""

    def _safe_num(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    # Compute Inland as leg1 + leg3
    leg1 = quote_df.get("leg1_cost_eur").fillna(0.0) if "leg1_cost_eur" in quote_df.columns else 0.0
    leg3 = quote_df.get("leg3_cost_eur").fillna(0.0) if "leg3_cost_eur" in quote_df.columns else 0.0
    inland_cost = leg1 + leg3

    # Incoterm series directly from input (column M in the template)
    incoterm_series = input_df.get("incoterm") if "incoterm" in input_df.columns else pd.Series([None] * len(input_df))
    pkg_vol_m3 = pd.to_numeric(quote_df.get("pkg_volume_m3"), errors="coerce")
    pkg_snp = pd.to_numeric(quote_df.get("pkg_snp"), errors="coerce")
    pack_per_cont = pd.to_numeric(quote_df.get("pack_per_cont_40ft"), errors="coerce")
    total_cost_eur = pd.to_numeric(quote_df.get("total_cost_eur"), errors="coerce")
    pkg_vol_m3_out = pkg_vol_m3.round(2)
    pkg_snp_out = pkg_snp.round(2)
    vol_per_cont_m3 = (pack_per_cont * pkg_vol_m3)
    part_vol_m3 = (pkg_vol_m3_out / pkg_snp_out).where(pkg_snp_out > 0)
    plant_to_plant_eur_m3 = (total_cost_eur / vol_per_cont_m3).where(vol_per_cont_m3 > 0)
    plant_to_plant_eur_part = (part_vol_m3 * plant_to_plant_eur_m3)
    pn_unit_cost_eur = pd.to_numeric(raw_like_df.get("PN Unit cost (€)"), errors="coerce")
    floating_stock_eur_part = pn_unit_cost_eur * 0.08 / 365 * pd.to_numeric(
        quote_df.get("transit_time_days"), errors="coerce"
    )
    pn_unit_cost_eur_out = pn_unit_cost_eur.round(2)
    plant_to_plant_eur_part_out = plant_to_plant_eur_part.round(2)
    floating_stock_eur_part_out = floating_stock_eur_part.round(2)
    pa_log_sf_total_eur_part = (pn_unit_cost_eur_out + plant_to_plant_eur_part_out + floating_stock_eur_part_out).round(2)
    annual_needs = pd.to_numeric(raw_like_df.get("Anual Needs (PN / Year)"), errors="coerce")
    daily_need = pd.to_numeric(raw_like_df.get("Daily Need (PN / Day)"), errors="coerce")
    annual_weight_keur = (annual_needs * pa_log_sf_total_eur_part / 1000).round(2)
    fcf_pipe_keur = (daily_need * pn_unit_cost_eur_out * pd.to_numeric(quote_df.get("transit_time_days"), errors="coerce") / 1000).round(2)

    computed_map = {
        "POL": quote_df.get("POL"),
        "POD": quote_df.get("POD"),
        # Unify POL distance and Leg1 distance in a single output column
        # Prefer the explicitly computed leg1_km when present, else fall back to pol_distance_km
        "Leg1/POL Distance (km)": pd.to_numeric(
            quote_df.get("leg1_km").fillna(quote_df.get("pol_distance_km")), errors="coerce"
        ).round(2),
        # Keep LEG3/POD Distance as the single Leg3 distance output
        "LEG3/POD Distance (km)": pd.to_numeric(quote_df.get("pod_distance_km"), errors="coerce").round(2),
        "Transit Time": pd.to_numeric(quote_df.get("transit_time_days"), errors="coerce").round(2),
        # Explicit leg cost columns
        "Leg1 Inland Cost (€)": pd.to_numeric(quote_df.get("leg1_cost_eur"), errors="coerce").fillna(0.0).round(2),
        "Leg2 Overseas Cost (€)": pd.to_numeric(quote_df.get("leg2_ocean_rate_eur"), errors="coerce").fillna(0.0).round(2),
        "Leg 3 Inland Cost (€)": pd.to_numeric(quote_df.get("leg3_cost_eur"), errors="coerce").fillna(0.0).round(2),
        "Total Transportation Cost (€)": total_cost_eur.fillna(0.0).round(2),
        "Packaging Code": quote_df.get("packaging_code_resolved"),
        "Packaging Volume (m³)": pkg_vol_m3_out,
        "SNP_Pack": pkg_snp_out,
        "Part volume(m3/part)": part_vol_m3.round(4),
        "pack/cont 40ft": pack_per_cont.round(2),
        "vol/cont 40ft (m3)": vol_per_cont_m3.round(2),
        "weight/cont 40ft (kg)": (
            pd.to_numeric(quote_df.get("pack_per_cont_40ft"), errors="coerce") *
            pd.to_numeric(quote_df.get("pkg_weight_full"), errors="coerce")
        ).round(2),
        "Plant to plant (€/m3)": plant_to_plant_eur_m3.round(2),
        "Plant to plant (€/part)": plant_to_plant_eur_part_out,
        "Floating Stock €/Part": floating_stock_eur_part_out,
        "PA + LOG + SF TOTAL €/Part": pa_log_sf_total_eur_part.round(2),
        "Annual weight K€": annual_weight_keur,
        "FCF Pipe K€": fcf_pipe_keur,
        "Weight/part (kg)": pd.to_numeric(quote_df.get("pkg_weight_part"), errors="coerce").round(2),
        "Weight empty pack (kg)": pd.to_numeric(quote_df.get("pkg_weight_empty"), errors="coerce").round(2),
        "Weight full pack (kg)": pd.to_numeric(quote_df.get("pkg_weight_full"), errors="coerce").round(2),
        "Type of Flow": quote_df.get("type_of_flow"),
        "Red flag/Debug": quote_df.get("Red flag/Debug", pd.Series([""] * len(quote_df))),
        # Provide 'Incoterm' from input as a fallback if not found in raw_like_df
        "Incoterm": incoterm_series
    }

    # Try to follow exact target order from reference output
    ref_cols = _load_reference_quote_columns()
    if ref_cols:
        cols = ref_cols
    else:
        # Fallback: original input columns first, then computable set in a stable order
        cols = [*STD_COLS.keys(), *[k for k in [
            "POL","POD","Leg1/POL Distance (km)","LEG3/POD Distance (km)","Transit Time",
            "Leg1 Inland Cost (€)","Leg2 Overseas Cost (€)",
            "Leg 3 Inland Cost (€)",
            "Total Transportation Cost (€)","Red flag/Debug"
        ] if k not in STD_COLS.keys()]]

    # Replace any legacy column names from reference with new naming
    # - Remove 'Leg3 Distance (km)'
    # - Rename 'POD Distance (km)' to 'LEG3/POD Distance (km)'
    # - Collapse 'POL Distance (km)' and 'Leg1 Distance (km)' into 'Leg1/POL Distance (km)'
    if "Leg3 Distance (km)" in cols:
        cols = [c for c in cols if c != "Leg3 Distance (km)"]
    if "POD Distance (km)" in cols and "LEG3/POD Distance (km)" not in cols:
        cols = [("LEG3/POD Distance (km)" if c == "POD Distance (km)" else c) for c in cols]
    # Handle both legacy POL/Leg1 distance columns -> unified
    if "Leg1/POL Distance (km)" not in cols and ("POL Distance (km)" in cols or "Leg1 Distance (km)" in cols):
        new_cols = []
        added = False
        for c in cols:
            if c in ("POL Distance (km)", "Leg1 Distance (km)"):
                if not added:
                    new_cols.append("Leg1/POL Distance (km)")
                    added = True
                # skip duplicates
            else:
                new_cols.append(c)
        cols = new_cols
    else:
        # Remove any lingering legacy columns if both exist
        cols = [c for c in cols if c not in ("POL Distance (km)", "Leg1 Distance (km)")]

    # Ensure 'Red flag/Debug' first and 'Type of Flow' second
    FLOW_COL = "Type of Flow"
    DEBUG_COL = "Red flag/Debug"
    # Guarantee presence even if not in reference
    if DEBUG_COL not in cols:
        cols = [DEBUG_COL] + cols
    if FLOW_COL not in cols:
        cols = [DEBUG_COL, FLOW_COL] + [c for c in cols if c != DEBUG_COL]
    # Reorder to put DEBUG first, FLOW second
    cols = [DEBUG_COL, FLOW_COL] + [c for c in cols if c not in (DEBUG_COL, FLOW_COL)]

    # New requirement: hide these location columns in final Quote output
    HIDDEN_QUOTE_COLS = {
        "Origin Country code",
        "Origin City",
        "Origin ZIP Code",
        "Destination country code",
        "Destination City",
        "Destinartion ZIP Code",
    }
    cols = [c for c in cols if c not in HIDDEN_QUOTE_COLS]

    # Ensure 'Incoterm' column remains present in the output (keep the initial template value)
    if "Incoterm" not in cols:
        # Place it right after Debug and Flow for visibility (third overall)
        # cols currently starts with [DEBUG_COL, FLOW_COL, ...]
        cols = [DEBUG_COL, FLOW_COL, "Incoterm"] + [c for c in cols if c not in (DEBUG_COL, FLOW_COL, "Incoterm")]

    # Remove legacy aggregated columns if present
    legacy_cost_cols = {
        "Inland Cost (€)",
        "Overseas Cost (€)",
        "TT (days)",
        "Packaging Volume (mm³)",
        "Part volume (m3)",
        "SNP / Pack (PN / Packaging)",
    }
    cols = [c for c in cols if c not in legacy_cost_cols]

    # Ensure packaging columns are always present (append at end if missing)
    PKG_COLS = [
        "Packaging Code",
        "Packaging Volume (m³)",
        "SNP_Pack",
        "Weight/part (kg)",
        "Weight empty pack (kg)",
        "Weight full pack (kg)",
        "Part volume(m3/part)",
        "pack/cont 40ft",
        "vol/cont 40ft (m3)",
        "weight/cont 40ft (kg)",
        "Plant to plant (€/m3)",
        "Plant to plant (€/part)",
        "Transit Time",
        "Floating Stock €/Part",
        "PA + LOG + SF TOTAL €/Part",
        "Annual weight K€",
        "FCF Pipe K€",
    ]
    for pc in PKG_COLS:
        if pc not in cols:
            cols.append(pc)

    # Ensure POL and POD are present (always visible in output)
    for k in ("POL", "POD"):
        if k not in cols:
            cols.append(k)

    # Ensure Transit Time and the three leg cost columns exist and are placed together.
    leg_cols_in_order = [
        "Leg1 Inland Cost (€)",
        "Leg2 Overseas Cost (€)",
        "Leg 3 Inland Cost (€)",
        "Total Transportation Cost (€)"
    ]
    for lc in ["Transit Time"] + leg_cols_in_order:
        if lc not in cols:
            cols.append(lc)
    # Reinsert leg columns right after Transit Time
    if "Transit Time" in cols:
        # Keep Debug and Flow at the front
        front = [DEBUG_COL, FLOW_COL, "Incoterm"] if "Incoterm" in cols else [DEBUG_COL, FLOW_COL]
        rest = [c for c in cols if c not in set(front)]
        # Remove leg columns from rest
        rest_wo_legs = [c for c in rest if c not in set(leg_cols_in_order)]
        # Ensure POL/POD sit immediately before Transit Time
        rest_wo_legs_no_ports = [c for c in rest_wo_legs if c not in ("POL", "POD", "Leg1/POL Distance (km)", "LEG3/POD Distance (km)")]
        # Find Transit Time index in rest_wo_legs
        if "Transit Time" in rest_wo_legs_no_ports:
            tt_idx = rest_wo_legs_no_ports.index("Transit Time")
            # Insert POL/POD just before TT
            before_tt = rest_wo_legs_no_ports[:tt_idx]
            after_tt = rest_wo_legs_no_ports[tt_idx+1:]
            ports_block = [c for c in ("POL", "POD", "Leg1/POL Distance (km)", "LEG3/POD Distance (km)") if c in rest_wo_legs]
            rest_final = before_tt + ports_block + ["Transit Time"] + leg_cols_in_order + after_tt
        else:
            # If Transit Time not in rest set, append in order
            ports_block = [c for c in ("POL", "POD", "Leg1/POL Distance (km)", "LEG3/POD Distance (km)") if c in rest_wo_legs]
            rest_final = rest_wo_legs_no_ports + ports_block + ["Transit Time"] + leg_cols_in_order
        cols = front + rest_final

    # Force requested business block order.
    REQUESTED_BLOCK_ORDER = [
        "Packaging Volume (m³)",
        "SNP_Pack",
        "Weight/part (kg)",
        "Weight empty pack (kg)",
        "Weight full pack (kg)",
        "Part volume(m3/part)",
        "pack/cont 40ft",
        "vol/cont 40ft (m3)",
        "weight/cont 40ft (kg)",
        "Plant to plant (€/m3)",
        "Plant to plant (€/part)",
        "Transit Time",
        "Floating Stock €/Part",
        "PA + LOG + SF TOTAL €/Part",
        "Annual weight K€",
        "FCF Pipe K€",
    ]
    for c in REQUESTED_BLOCK_ORDER:
        if c not in cols:
            cols.append(c)
    cols = [c for c in cols if c not in REQUESTED_BLOCK_ORDER] + REQUESTED_BLOCK_ORDER

    # Assemble final DataFrame respecting target columns
    data = {}
    n = len(input_df)
    for c in cols:
        if c in raw_like_df.columns:
            data[c] = raw_like_df[c].reset_index(drop=True)
        elif c in computed_map:
            s = computed_map[c]
            # Ensure correct length/index
            if hasattr(s, "reset_index"):
                s = s.reset_index(drop=True)
            data[c] = s
        else:
            data[c] = pd.Series([""] * n)
    final_quote_df = pd.DataFrame(data, columns=cols)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        input_df.to_excel(writer, sheet_name="Input", index=False)
        final_quote_df.to_excel(writer, sheet_name="Quote", index=False)

        # Highlight all Quote headers in yellow.
        quote_ws = writer.book["Quote"]
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for cell in quote_ws[1]:
            cell.fill = yellow_fill

        # Adjust Quote column widths from B to AL based on content length.
        for col_idx in range(2, 39):  # B..AL
            max_len = 0
            for row in quote_ws.iter_rows(min_row=1, max_row=quote_ws.max_row, min_col=col_idx, max_col=col_idx):
                val = row[0].value
                if val is None:
                    continue
                txt = str(val)
                if len(txt) > max_len:
                    max_len = len(txt)
            quote_ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 45))

        # Force 2-decimal display in key numeric result columns.
        two_dec_cols = {
            "Leg1/POL Distance (km)",
            "LEG3/POD Distance (km)",
            "Transit Time",
            "Leg1 Inland Cost (€)",
            "Leg2 Overseas Cost (€)",
            "Leg 3 Inland Cost (€)",
            "Total Transportation Cost (€)",
            "Packaging Volume (m³)",
            "SNP_Pack",
            "Part volume(m3/part)",
            "pack/cont 40ft",
            "vol/cont 40ft (m3)",
            "weight/cont 40ft (kg)",
            "Plant to plant (€/m3)",
            "Plant to plant (€/part)",
            "Floating Stock €/Part",
            "PA + LOG + SF TOTAL €/Part",
            "Annual weight K€",
            "FCF Pipe K€",
            "Weight/part (kg)",
            "Weight empty pack (kg)",
            "Weight full pack (kg)",
        }
        header_to_idx = {str(c.value): i + 1 for i, c in enumerate(quote_ws[1])}
        for h in two_dec_cols:
            idx = header_to_idx.get(h)
            if not idx:
                continue
            for row in quote_ws.iter_rows(min_row=2, max_row=quote_ws.max_row, min_col=idx, max_col=idx):
                cell = row[0]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

        # Specific precision override: part volume needs 4 decimals.
        idx_part_vol = header_to_idx.get("Part volume(m3/part)")
        if idx_part_vol:
            for row in quote_ws.iter_rows(min_row=2, max_row=quote_ws.max_row, min_col=idx_part_vol, max_col=idx_part_vol):
                cell = row[0]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0000"

        if "incoterm" in input_df.columns:
            inc_series = input_df["incoterm"].dropna()
            used_incoterms = sorted(set([str(x).upper().strip() for x in inc_series.tolist()]))
        else:
            used_incoterms = []
        incoterm_summary = ", ".join(used_incoterms) if used_incoterms else DEFAULT_INCOTERM
        summary = pd.DataFrame({
            "Generated": [datetime.now().isoformat(sep=" ", timespec="seconds")],
            "Rows": [len(input_df)],
            "Incoterms": [incoterm_summary],
            "Note": ["Incoterm aplicado por fila (fallback al global si vacío/incorrecto)"],
        })
        summary.to_excel(writer, sheet_name="Summary", index=False)


def main():
    try:
        df = load_input_template(INPUT_FILE, sheet="Input")
    except PermissionError as e:
        print(f"ERROR: No se pudo leer el template porque está abierto o bloqueado: {INPUT_FILE}")
        print(str(e))
        return
    except FileNotFoundError:
        print(f"ERROR: No se encontró el template esperado: {INPUT_FILE}")
        return
    out_path = next_output_path(QTOOL_DIR)
    try:
        build_output(df, out_path)
    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        return
    except PermissionError as e:
        print(f"ERROR: {e}")
        return
    print(out_path)


if __name__ == "__main__":
    main()
