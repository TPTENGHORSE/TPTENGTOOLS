import os
import re
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageDraw, ImageFont


def render_box(label, value):
    return f"""
    <div class='vtt-box'>
        <div class='vtt-box__label'>{label}</div>
        <div class='vtt-box__value' title='{value}'>{value}</div>
    </div>
    """

def _coerce_to_int(val):
    """Attempt to coerce various cell formats to an integer.
    Handles None/NaN, numeric strings with punctuation, and floats.
    Returns 0 on failure.
    """
    try:
        if pd.isna(val):
            return 0
    except Exception:
        pass
    if isinstance(val, (int, float)):
        try:
            return int(round(float(val)))
        except Exception:
            return 0
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return 0
        m = re.search(r"[-+]?\d+(?:[\.,]\d+)?", s)
        if m:
            try:
                num = float(m.group(0).replace(',', '.'))
                return int(round(num))
            except Exception:
                return 0
    try:
        return int(val)
    except Exception:
        return 0


def _due_date_day_plus_value(row, df_vtt):
    try:
        if row is None:
            return 7

        row_id = str(row.get('ID', '')).strip().upper() if 'ID' in df_vtt.columns else ''
        row_pol = str(row.get('POL', '')).strip().upper() if 'POL' in df_vtt.columns else ''
        row_pod = str(row.get('POD', '')).strip().upper() if 'POD' in df_vtt.columns else ''

        if row_id == 'O001' and row_pol == 'CNSHA' and row_pod == 'PTLEI':
            return 5
    except Exception:
        pass
    return 7


def _format_expiration_date(row, df_vtt):
    try:
        if row is None or 'Expiration Date' not in df_vtt.columns:
            return ''
        raw_exp = row.get('Expiration Date', '')
        if pd.isna(raw_exp):
            return ''
        if isinstance(raw_exp, (pd.Timestamp, datetime)):
            return raw_exp.strftime('%d/%m/%Y')
        try:
            return pd.to_datetime(raw_exp).strftime('%d/%m/%Y')
        except Exception:
            return str(raw_exp)
    except Exception:
        try:
            return '' if row is None else str(row.get('Expiration Date', ''))
        except Exception:
            return ''

# Load data from new Excel (VTT DATA.xlsx)
vtt_data_path = os.path.join(os.path.dirname(__file__), "VTT DATA.xlsx")
df_vtt = pd.read_excel(vtt_data_path)

# --- STREAMLIT INTERFACE ---
st.set_page_config(layout="wide")

st.markdown(
    """
    <style>
    :root {
        --vtt-bg: #f3f6fb;
        --vtt-surface: #ffffff;
        --vtt-surface-strong: #f8fbff;
        --vtt-border: #d7e2f0;
        --vtt-border-soft: #e7edf5;
        --vtt-text: #17263c;
        --vtt-text-soft: #5f7088;
        --vtt-primary: #183a63;
        --vtt-primary-strong: #102845;
        --vtt-primary-soft: #eaf2fb;
        --vtt-success: #89e78c;
        --vtt-weekend: #ffdede;
        --vtt-shadow: 0 14px 34px rgba(16, 40, 69, 0.08);
    }
    .stApp {
        background:
            radial-gradient(circle at top left, rgba(24, 58, 99, 0.08), transparent 24%),
            linear-gradient(180deg, #f7f9fc 0%, var(--vtt-bg) 100%);
    }
    .main .block-container {
        padding-top: 0.8rem !important;
        padding-left: 1.1rem !important;
        padding-right: 1.1rem !important;
        padding-bottom: 2rem !important;
        max-width: 86% !important;
    }
    header[data-testid="stHeader"] {
        height: 0px !important;
        min-height: 0px !important;
        padding: 0 !important;
    }
    hr {
        border: none !important;
        height: 1px !important;
        background: linear-gradient(90deg, transparent 0%, var(--vtt-border) 15%, var(--vtt-border) 85%, transparent 100%) !important;
    }
    .vtt-page-title {
        text-align: center;
        color: var(--vtt-primary-strong);
        font-size: 2.2rem;
        font-weight: 800;
        letter-spacing: -0.03em;
        margin: 0.1rem 0 0.2rem 0;
    }
    .vtt-page-subtitle {
        text-align: center;
        color: var(--vtt-text-soft);
        font-size: 0.95rem;
        margin: 0 0 1.35rem 0;
    }
    .vtt-box {
        background: linear-gradient(180deg, var(--vtt-surface-strong) 0%, var(--vtt-surface) 100%);
        border: 1px solid var(--vtt-border);
        border-radius: 14px;
        padding: 0.7rem 0.8rem;
        box-shadow: var(--vtt-shadow);
        min-height: 72px;
    }
    .vtt-box__label {
        color: var(--vtt-text-soft);
        font-size: 0.72rem;
        font-weight: 700;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        margin-bottom: 0.45rem;
    }
    .vtt-box__value {
        color: var(--vtt-text);
        font-size: 0.95rem;
        font-weight: 700;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    .vtt-panel {
        background: var(--vtt-surface);
        border: 1px solid var(--vtt-border);
        border-radius: 18px;
        padding: 1rem 1rem 0.9rem 1rem;
        box-shadow: var(--vtt-shadow);
        margin-top: 0.45rem;
    }
    .vtt-panel__title,
    .vtt-section-title {
        color: var(--vtt-primary-strong);
        font-size: 1.05rem;
        font-weight: 800;
        letter-spacing: -0.02em;
        margin-bottom: 0.75rem;
    }
    .vtt-kpi-row {
        display: flex;
        justify-content: flex-start;
        align-items: center;
    }
    .vtt-kpi-card {
        display: inline-flex;
        align-items: center;
        justify-content: space-between;
        gap: 1rem;
        background: linear-gradient(135deg, #ffffff 0%, #f3f8ff 100%);
        border: 1px solid var(--vtt-border);
        border-radius: 18px;
        padding: 0.95rem 1.1rem;
        box-shadow: var(--vtt-shadow);
        margin: 1rem 0 1.1rem 0;
        width: auto;
        min-width: 320px;
        max-width: 460px;
    }
    .vtt-kpi-card__label {
        color: var(--vtt-primary-strong);
        font-size: 0.95rem;
        font-weight: 800;
        letter-spacing: 0.02em;
    }
    .vtt-kpi-card__value {
        min-width: 72px;
        text-align: center;
        color: var(--vtt-primary-strong);
        font-size: 1.7rem;
        font-weight: 900;
        background: #ffffff;
        border: 1px solid var(--vtt-border);
        border-radius: 14px;
        padding: 0.3rem 0.8rem;
    }
    .vtt-action-bar {
        width: 100%;
        display: flex;
        justify-content: center;
        align-items: center;
        gap: 16px;
        margin: 24px 0 8px 0;
    }
    .vtt-action-btn {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        min-width: 170px;
        background: linear-gradient(135deg, var(--vtt-primary) 0%, #24548d 100%);
        color: #fff;
        border: none;
        border-radius: 14px;
        padding: 12px 18px;
        font-size: 16px;
        font-weight: 700;
        cursor: pointer;
        box-shadow: 0 12px 24px rgba(24, 58, 99, 0.22);
        transition: transform 0.15s ease, box-shadow 0.15s ease;
    }
    .vtt-action-btn:hover {
        transform: translateY(-1px);
        box-shadow: 0 16px 28px rgba(24, 58, 99, 0.28);
    }
    div[data-baseweb="select"] > div {
        background: linear-gradient(180deg, var(--vtt-surface-strong) 0%, var(--vtt-surface) 100%) !important;
        border: 1px solid var(--vtt-border) !important;
        border-radius: 14px !important;
        min-height: 50px !important;
        box-shadow: var(--vtt-shadow) !important;
    }
    div[data-baseweb="select"] span,
    div[data-baseweb="select"] input {
        color: var(--vtt-text) !important;
        font-weight: 600 !important;
    }
    .stButton > button {
        min-height: 50px !important;
        border-radius: 14px !important;
        border: none !important;
        background: linear-gradient(135deg, var(--vtt-primary) 0%, #24548d 100%) !important;
        color: #ffffff !important;
        font-weight: 800 !important;
        letter-spacing: 0.01em;
        box-shadow: 0 12px 24px rgba(24, 58, 99, 0.22) !important;
    }
    .stButton > button:hover {
        background: linear-gradient(135deg, var(--vtt-primary-strong) 0%, var(--vtt-primary) 100%) !important;
    }
    /* vertical text utility for date headers */
    .vtt-vertical-text {
        display: inline-block;
        writing-mode: vertical-rl;
        text-orientation: upright;
        white-space: nowrap;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class='vtt-page-title'>VTT Tool</div>
    <div class='vtt-page-subtitle'>Lead time and transit planning dashboard</div>
    """,
    unsafe_allow_html=True,
)

top_section = st.container()

with top_section:
    all_label = "Todos"

    def _apply_cross_filters(dataframe, pol_value=all_label, pod_value=all_label, id_value=all_label):
        result = dataframe
        if 'POL' in result.columns and pol_value and pol_value != all_label:
            result = result[result['POL'].astype(str) == str(pol_value)]
        if 'POD' in result.columns and pod_value and pod_value != all_label:
            result = result[result['POD'].astype(str) == str(pod_value)]
        if 'ID' in result.columns and id_value and id_value != all_label:
            result = result[result['ID'].astype(str) == str(id_value)]
        return result

    def _set_last_changed_filter(filter_name):
        st.session_state['_last_changed_filter'] = filter_name

    # Defaults in session state
    if 'pol_select' not in st.session_state:
        st.session_state['pol_select'] = all_label
    if 'pod_select' not in st.session_state:
        st.session_state['pod_select'] = all_label
    if 'id_select' not in st.session_state:
        st.session_state['id_select'] = all_label
    if '_last_changed_filter' not in st.session_state:
        st.session_state['_last_changed_filter'] = None

    # POL stays global. POD cascades from POL. ID stays global so it never blocks later POL/POD changes.
    all_pol_options = [all_label] + (df_vtt['POL'].dropna().astype(str).unique().tolist() if 'POL' in df_vtt.columns else [])
    all_pod_options = [all_label] + (df_vtt['POD'].dropna().astype(str).unique().tolist() if 'POD' in df_vtt.columns else [])
    all_id_options = [all_label] + (df_vtt['ID'].dropna().astype(str).unique().tolist() if 'ID' in df_vtt.columns else [])
    pol_options = all_pol_options
    id_options = all_id_options

    selected_pol_value = st.session_state.get('pol_select', all_label)
    selected_pod_value = st.session_state.get('pod_select', all_label)
    selected_id_value = st.session_state.get('id_select', all_label)
    last_changed_filter = st.session_state.get('_last_changed_filter')

    if selected_pol_value != all_label and 'POL' in df_vtt.columns and 'POD' in df_vtt.columns:
        pod_scope_df = df_vtt[df_vtt['POL'].astype(str) == str(selected_pol_value)]
        pod_options = [all_label] + pod_scope_df['POD'].dropna().astype(str).unique().tolist()
    else:
        pod_options = all_pod_options

    if last_changed_filter == 'id_select' and selected_id_value != all_label and 'ID' in df_vtt.columns:
        id_scope_df = df_vtt[df_vtt['ID'].astype(str) == str(selected_id_value)]
        pol_from_id = id_scope_df['POL'].dropna().astype(str).unique().tolist() if 'POL' in id_scope_df.columns else []
        pod_from_id = id_scope_df['POD'].dropna().astype(str).unique().tolist() if 'POD' in id_scope_df.columns else []

        if len(pol_from_id) == 1:
            st.session_state['pol_select'] = pol_from_id[0]
        elif pol_from_id and st.session_state.get('pol_select', all_label) not in pol_from_id:
            st.session_state['pol_select'] = pol_from_id[0]

        if len(pod_from_id) == 1:
            st.session_state['pod_select'] = pod_from_id[0]
        elif pod_from_id and st.session_state.get('pod_select', all_label) not in pod_from_id:
            st.session_state['pod_select'] = pod_from_id[0]

    elif (
        last_changed_filter in ('pol_select', 'pod_select')
        and st.session_state.get('pol_select', all_label) != all_label
        and st.session_state.get('pod_select', all_label) != all_label
        and 'POL' in df_vtt.columns
        and 'POD' in df_vtt.columns
        and 'ID' in df_vtt.columns
    ):
        pol_pod_scope_df = df_vtt[
            (df_vtt['POL'].astype(str) == str(st.session_state.get('pol_select', all_label)))
            & (df_vtt['POD'].astype(str) == str(st.session_state.get('pod_select', all_label)))
        ]
        id_from_pol_pod = pol_pod_scope_df['ID'].dropna().astype(str).unique().tolist()
        current_id_value = st.session_state.get('id_select', all_label)
        if len(id_from_pol_pod) == 1:
            st.session_state['id_select'] = id_from_pol_pod[0]
        elif id_from_pol_pod and current_id_value not in id_from_pol_pod:
            st.session_state['id_select'] = id_from_pol_pod[0]
        elif not id_from_pol_pod:
            st.session_state['id_select'] = all_label

    # Keep current values valid
    if st.session_state.get('pol_select', all_label) not in pol_options:
        st.session_state['pol_select'] = all_label
    if st.session_state.get('pod_select', all_label) not in pod_options:
        st.session_state['pod_select'] = all_label
    if st.session_state.get('id_select', all_label) not in id_options:
        st.session_state['id_select'] = all_label

    # Render the full top row in a single aligned layout.
    top_cols = st.columns([0.9, 0.9, 1.05, 1.3, 1.3, 1.15, 1.15, 1.0, 1.1, 1.15], gap="medium")
    compact_label_style = "font-size:11px; font-weight:800; color:#5f7088; letter-spacing:0.08em; text-transform:uppercase; line-height:1; margin:0 0 8px;"

    with top_cols[0]:
        st.markdown(f"<div style='{compact_label_style}'>POL</div>", unsafe_allow_html=True)
        st.selectbox("POL", pol_options, key="pol_select", label_visibility="collapsed", on_change=_set_last_changed_filter, args=('pol_select',))
    with top_cols[1]:
        st.markdown(f"<div style='{compact_label_style}'>POD</div>", unsafe_allow_html=True)
        st.selectbox("POD", pod_options, key="pod_select", label_visibility="collapsed", on_change=_set_last_changed_filter, args=('pod_select',))
    with top_cols[2]:
        st.markdown(f"<div style='{compact_label_style}'>ID</div>", unsafe_allow_html=True)
        st.selectbox("ID", id_options, key="id_select", label_visibility="collapsed", on_change=_set_last_changed_filter, args=('id_select',))

    # Final filtered dataset from active bidirectional selections
    filtered_df = _apply_cross_filters(
        df_vtt,
        pol_value=st.session_state.get('pol_select', all_label),
        pod_value=st.session_state.get('pod_select', all_label),
        id_value=st.session_state.get('id_select', all_label),
    )

    if not filtered_df.empty:
        row = filtered_df.iloc[0]
    else:
        row = None

    with top_cols[3]:
        if row is not None and 'Carrier' in df_vtt.columns:
            st.markdown(render_box('Carrier', row['Carrier']), unsafe_allow_html=True)
        else:
            st.info("No existe la columna Carrier (Carrier) o no hay coincidencia.")
    with top_cols[4]:
        if row is not None and len(df_vtt.columns) > 10:
            try:
                col_k = df_vtt.columns[10]
                st.markdown(render_box('Shipper', row.get(col_k, "")), unsafe_allow_html=True)
            except Exception:
                st.info("No se pudo leer la columna K (Shipper) o no hay coincidencia.")
        else:
            st.info("No se pudo leer la columna K (Shipper) o no hay coincidencia.")
    with top_cols[5]:
        if row is not None and len(df_vtt.columns) > 8:
            try:
                col_i = df_vtt.columns[8]
                st.markdown(render_box('ILN/FF', row.get(col_i, "")), unsafe_allow_html=True)
            except Exception:
                st.info("No se pudo leer la columna I (ILN/FF) o no hay coincidencia.")
        else:
            st.info("No se pudo leer la columna I (ILN/FF) o no hay coincidencia.")
    with top_cols[6]:
        if row is not None and 'Name Destin Site' in df_vtt.columns:
            st.markdown(render_box('PLANT', row['Name Destin Site']), unsafe_allow_html=True)
        else:
            st.info("No existe la columna Name Destin Site o no hay coincidencia.")
    with top_cols[7]:
        st.markdown(render_box('E/D', _format_expiration_date(row, df_vtt)), unsafe_allow_html=True)
    with top_cols[8]:
        commodity_value = ""
        if row is not None:
            if 'Commodity' in df_vtt.columns:
                commodity_value = row.get('Commodity', "")
            elif 'Comodity' in df_vtt.columns:
                commodity_value = row.get('Comodity', "")
        st.markdown(render_box('Commodity', commodity_value), unsafe_allow_html=True)
    with top_cols[9]:
        st.markdown("<div style='height:21px;'></div>", unsafe_allow_html=True)
        generate_files_clicked = st.button("Generate files", key="generate_files", use_container_width=True)

safety_stock_val = None
if row is not None and 'Safety stock' in df_vtt.columns:
    safety_stock_val = row['Safety stock']

st.markdown("<hr style='margin:16px 0;'>", unsafe_allow_html=True)

st.markdown("<div style='height: 8px'></div>", unsafe_allow_html=True)

time_cols_fixed = 4
today = datetime.today()
start_date = today - timedelta(days=today.weekday())
num_days = int(st.session_state.get("days_slider_timeline", 110))
timeline_days = [start_date + timedelta(days=i) for i in range(num_days)]
time_cols = time_cols_fixed + num_days

# Encabezados fijos y dinámicos
headers = ["Steps", "Day", "Day+", "Final Day"]
table_html = """
<table class='timeline-table' style='width:100%; border-collapse:collapse; margin-top:8px;'>
    <thead>"""
table_html_visible = table_html
# Fila de semana combinada
table_html += "<tr>"
table_html_visible += "<tr>"
for idx_h, h in enumerate(headers):
    if idx_h == 0:
        # Steps column: wider and no wrapping
        table_html += "<th style='border:none; background:none; min-width:80px; white-space:nowrap;'></th>"
        table_html_visible += "<th style='border:none; background:none; min-width:80px; white-space:nowrap;'></th>"
    else:
        table_html += "<th style='border:none; background:none'></th>"
        table_html_visible += "<th style='border:none; background:none'></th>"
# Agrupar días por semana
semana_actual = None
colspan = 0
for idx, day in enumerate(timeline_days):
    semana = day.isocalendar()[1]
    if semana_actual is None:
        semana_actual = semana
        colspan = 1
    elif semana == semana_actual:
        colspan += 1
    else:
        # Imprimir celda combinada para la semana anterior
        week_header = f"<th colspan='{colspan}' style='padding:0 1px; border:1px solid #eee; min-width:28px; text-align:center; background:#fffbe6; font-size:13.5px; font-weight:bold;'>W{semana_actual}</th>"
        table_html += week_header
        table_html_visible += week_header
        semana_actual = semana
        colspan = 1
# Imprimir la última semana
if semana_actual is not None:
    week_header = f"<th colspan='{colspan}' style='padding:0 1px; border:1px solid #eee; min-width:28px; text-align:center; background:#fffbe6; font-size:13.5px; font-weight:bold;'>W{semana_actual}</th>"
    table_html += week_header
    table_html_visible += week_header
table_html += "</tr>"
table_html_visible += "</tr>"
# Fila de encabezados de fechas
table_html += "<tr>"
table_html_visible += "<tr>"
for idx_h, h in enumerate(headers):
    if idx_h == 0:
        fixed_header = f"<th style='padding:5px 7px; border:1px solid #eee; min-width:200px; text-align:center; background:#f5f5f5; white-space:nowrap'>{h}</th>"
    else:
        fixed_header = f"<th style='padding:5px 7px; border:1px solid #eee; min-width:50px; width:50px; text-align:center; background:#f5f5f5'>{h}</th>"
    table_html += fixed_header
    table_html_visible += fixed_header
for idx, day in enumerate(timeline_days):
    # Colorear sábados y domingos
    if day.weekday() in [5, 6]:
        th_style = "padding:0 1px; border:1px solid #eee; min-width:15px; width:18px; height:50px; text-align:center; background:#ffd6d6; font-size:12px; vertical-align:bottom;"
    else:
        th_style = "padding:0 1px; border:1px solid #eee; min-width:20px; width:20px; height:50px; text-align:center; background:#e3eafc; font-size:12px; vertical-align:bottom;"
    # Mostrar solo la letra inicial del día en mayúscula
    vertical_label = day.strftime('%a')[0].upper()  # M, T, W, etc.
    # Centrar verticalmente la letra inicial
    day_header = f"<th style='{th_style}'><span class='vtt-vertical-text' style='display:flex;align-items:center;justify-content:center;height:100%;'>{vertical_label}</span></th>"
    table_html += day_header
    table_html_visible += day_header
table_html += "</tr></thead><tbody>"
table_html_visible += "</tr></thead><tbody>"

# Etiquetas de filas
time_labels = [
    "1. Day Customer Order",
    "2. Day ILN/FF Order",
    "3. First Receipt Days",
    "4. Pack. prep. & load",
    "5. Transport to POL",
    "6. First Day to POL",
    "7. Cut off",
    "8. ETD",
    "9. Transit Duration (ETD>ETA)",
    "10. Days of flexibility",
    "11. Customs clearence",
    "12. Transport to plant",
    "13. Rounding",
    "14. Due Date"
]

time_rows = len(time_labels)
for i in range(time_rows):
    source_step = i if i < 10 else i + 1
    # Reduce row height ~35% (15px -> ~10px)
    table_html += "<tr style='height:15px;'>"
    table_html_visible += "<tr style='height:15px;'>"
    for j in range(time_cols):
        cell_content = ""
        cell_content_visible = ""
        # Alinear la primera columna (etiquetas) a la izquierda
        if j == 0:
            # Steps column: make it wider and prevent wrapping
            cell_style = "padding:4px 6px; border:1px solid #eee; text-align:left; font-weight:bold; background:#f5f5f5; min-width:200px; white-space:nowrap;"
        else:
            cell_style = "padding:4px 6px; border:1px solid #eee; text-align:center;"
        cell_style_visible = cell_style
        # Compactar altura y padding en todas las celdas de steps (≈ -35%)
        cell_style += "height:15px; line-height:15px; padding:1px 4px;"
        cell_style_visible += "height:15px; line-height:15px; padding:1px 4px;"
        # Colorear sábados y domingos en las celdas de fechas
        if j >= 4:
            fecha_actual = timeline_days[j-4] if (j-4) < len(timeline_days) else None
            if fecha_actual is not None and fecha_actual.weekday() in [5, 6]:
                cell_style += "background-color:#ffd6d6;"
                cell_style_visible += "background-color:#ffd6d6;"
        if i == 0:  # 1. Day Customer Order
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '1 Day Customer Order' in df_vtt.columns:
                    cell_content = row['1 Day Customer Order']
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                cell_content = "0"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '1 Day Customer Order' in df_vtt.columns:
                    cell_content = row['1 Day Customer Order']
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['1 Day Customer Order']) if row is not None and '1 Day Customer Order' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                paint_len = 1  # Day+ = 0 -> solo último día
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif i == 1:  # 2. Day ILN Order
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '2 Day ILN Order' in df_vtt.columns:
                    val = row['2 Day ILN Order']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                cell_content = "0"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '2 Day ILN Order' in df_vtt.columns:
                    val = row['2 Day ILN Order']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['2 Day ILN Order']) if row is not None and '2 Day ILN Order' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                paint_len = 1  # Day+ = 0 -> solo último día
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif i == 2:  # 3. First Receipt Days
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '3 First Receipt Days' in df_vtt.columns:
                    val = row['3 First Receipt Days']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "No hay datos para la combinación POL/POD seleccionada"
                cell_content_visible = cell_content
            elif j == 2:
                if row is not None and '3 .1 Time of Recept in ILN' in df_vtt.columns:
                    val = row['3 .1 Time of Recept in ILN']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '3.2 First Receipt Days' in df_vtt.columns:
                    val = row['3.2 First Receipt Days']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['3.2 First Receipt Days']) if row is not None and '3.2 First Receipt Days' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                day_plus_val = _coerce_to_int(row['3 .1 Time of Recept in ILN']) if row is not None and '3 .1 Time of Recept in ILN' in df_vtt.columns else 0
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif i == 4:  # 5. Transport ILN to POL
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '5.1 Transport ILN to POL' in df_vtt.columns:
                    val = row['5.1 Transport ILN to POL']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                if row is not None and '5.2 Transport ILN to POL' in df_vtt.columns:
                    val = row['5.2 Transport ILN to POL']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '5.3 Transport ILN to POL' in df_vtt.columns:
                    val = row['5.3 Transport ILN to POL']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['5.3 Transport ILN to POL']) if row is not None and '5.3 Transport ILN to POL' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                day_plus_val = _coerce_to_int(row['5.2 Transport ILN to POL']) if row is not None and '5.2 Transport ILN to POL' in df_vtt.columns else 0
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif i == 5:  # 6. First Day to POL
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '6 First Day to POL' in df_vtt.columns:
                    val = row['6 First Day to POL']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                if row is not None and 'First Day to POL' in df_vtt.columns:
                    val = row['First Day to POL']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "0"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '6 First Day to POL' in df_vtt.columns:
                    val = row['6 First Day to POL']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['6 First Day to POL']) if row is not None and '6 First Day to POL' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                day_plus_val = _coerce_to_int(row['First Day to POL']) if row is not None and 'First Day to POL' in df_vtt.columns else 0
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif i == 6:  # 7. Cut off
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '7 Cutt off' in df_vtt.columns:
                    val = row['7 Cutt off']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                cell_content = "0"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '7 Cutt off' in df_vtt.columns:
                    val = row['7 Cutt off']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['7 Cutt off']) if row is not None and '7 Cutt off' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                paint_len = 1  # Day+ = 0
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif i == 7:  # 8. ETD
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '8 ETD' in df_vtt.columns:
                    val = row['8 ETD']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                cell_content = "0"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '8 ETD' in df_vtt.columns:
                    val = row['8 ETD']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['8 ETD']) if row is not None and '8 ETD' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                paint_len = 1  # Day+ = 0
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif i == 8:  # 9. TT (ETD> ETA)
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '8 ETD' in df_vtt.columns:
                    val = row['8 ETD']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                if row is not None and 'Transit time' in df_vtt.columns:
                    val = row['Transit time']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 3:
                final_col = None
                if row is not None:
                    if '9 ETD> ETA' in df_vtt.columns:
                        final_col = '9 ETD> ETA'
                    elif '9 ETD>ETA' in df_vtt.columns:
                        final_col = '9 ETD>ETA'
                if final_col is not None:
                    val = row[final_col]
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                dias_final_day = 0
                try:
                    if row is not None and '9 ETD> ETA' in df_vtt.columns:
                        dias_final_day = int(row['9 ETD> ETA'])
                    elif row is not None and '9 ETD>ETA' in df_vtt.columns:
                        dias_final_day = int(row['9 ETD>ETA'])
                except Exception:
                    dias_final_day = 0
                day_plus_val = _coerce_to_int(row['Transit time']) if row is not None and 'Transit time' in df_vtt.columns else 0
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = "<span style='color:#ffffff; font-size:12px; line-height:1;'>🚢</span>"
                    # Azul más claro para Transit Duration (ETD>ETA)
                    cell_style += "background-color:#4a90e2;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif i == 9:  # 10. Days of flexibility
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:  # Day = Final day of step 9 + 1
                # Base: '9 ETD> ETA' o '9 ETD>ETA'
                base_val = None
                if row is not None:
                    candidate = None
                    if '9 ETD> ETA' in df_vtt.columns:
                        candidate = row.get('9 ETD> ETA')
                    if (candidate is None or pd.isna(candidate)) and '9 ETD>ETA' in df_vtt.columns:
                        candidate = row.get('9 ETD>ETA')
                    if pd.isna(candidate) if isinstance(candidate, (int, float, pd.Series, pd.Timestamp)) else (candidate is None):
                        cell_content = "-"
                    else:
                        # Convertir a número y sumar 1
                        try:
                            num_val = pd.to_numeric(candidate, errors='coerce')
                            if pd.isna(num_val):
                                # regex fallback
                                matches = re.findall(r"[-+]?\d*\.?\d+", str(candidate))
                                num_val = float(matches[0]) if matches else float('nan')
                            if pd.isna(num_val):
                                cell_content = "-"
                            else:
                                cell_content = str(int(float(num_val)) + 1)
                        except Exception:
                            cell_content = "-"
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:  # Day+ = Time for security + Time for security2 buffer
                if row is not None and ('Time for security' in df_vtt.columns or 'Time for security2 buffer' in df_vtt.columns):
                    flex_1 = _coerce_to_int(row['Time for security']) if 'Time for security' in df_vtt.columns else 0
                    flex_2 = _coerce_to_int(row['Time for security2 buffer']) if 'Time for security2 buffer' in df_vtt.columns else 0
                    cell_content = str(flex_1 + flex_2)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 3:  # Final Day
                if row is not None and '11 Days flexibility 2' in df_vtt.columns:
                    val = row['11 Days flexibility 2']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                    try:
                        base = row['10 Days flexibility 1'] if row is not None and '10 Days flexibility 1' in df_vtt.columns else None
                        bnum = pd.to_numeric(base, errors='coerce') if base is not None else float('nan')
                        if pd.isna(bnum):
                            m = re.findall(r"[-+]?\.?\d+", str(base)) if base is not None else []
                            bnum = float(m[0]) if m else float('nan')
                        plus = _coerce_to_int(row['Time for security2 buffer']) if row is not None and 'Time for security2 buffer' in df_vtt.columns else 0
                        if not pd.isna(bnum):
                            cell_content = str(int(float(bnum)) + 1 + int(plus))
                    except Exception:
                        cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = 0
                    if row is not None and '11 Days flexibility 2' in df_vtt.columns:
                        dias_final_day = int(row['11 Days flexibility 2'])
                    elif row is not None and '10 Days flexibility 1' in df_vtt.columns:
                        base_final = pd.to_numeric(row['10 Days flexibility 1'], errors='coerce')
                        if pd.isna(base_final):
                            matches = re.findall(r"[-+]?\d*\.?\d+", str(row['10 Days flexibility 1']))
                            base_final = float(matches[0]) if matches else float('nan')
                        plus_2 = _coerce_to_int(row['Time for security2 buffer']) if 'Time for security2 buffer' in df_vtt.columns else 0
                        if not pd.isna(base_final):
                            dias_final_day = int(float(base_final)) + 1 + int(plus_2)
                except Exception:
                    dias_final_day = 0
                flex_1 = _coerce_to_int(row['Time for security']) if row is not None and 'Time for security' in df_vtt.columns else 0
                flex_2 = _coerce_to_int(row['Time for security2 buffer']) if row is not None and 'Time for security2 buffer' in df_vtt.columns else 0
                day_plus_val = flex_1 + flex_2
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style

                if start_idx <= (j-3) <= dias_final_day and flex_1 > 0:
                    time_for_security_end = start_idx + flex_1 - 1
                    if start_idx <= (j-3) <= time_for_security_end:
                        cell_style_visible = cell_style_visible.replace("background-color:#90ee90;", "background-color:#87ceeb;")
        elif source_step == 11:  # 11. Customs clearence
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '11 Days flexibility 2' in df_vtt.columns:
                    base_val = row['11 Days flexibility 2']
                    num_val = pd.to_numeric(base_val, errors='coerce')
                    if pd.isna(num_val):
                        try:
                            matches = re.findall(r"[-+]?\d*\.?\d+", str(base_val))
                            num_val = float(matches[0]) if matches else float('nan')
                        except Exception:
                            num_val = float('nan')
                    if pd.isna(num_val):
                        cell_content = "-"
                    else:
                        try:
                            cell_content = str(int(float(num_val)) + 1)
                        except Exception:
                            cell_content = "-"
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                if row is not None and 'Cust.' in df_vtt.columns:
                    val = row['Cust.']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 3:
                final_col = None
                if row is not None:
                    if '12 Customs Clearance' in df_vtt.columns:
                        final_col = '12 Customs Clearance'
                    elif '12 Customs clearence' in df_vtt.columns:
                        final_col = '12 Customs clearence'
                if final_col is not None:
                    val = row[final_col]
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                dias_final_day = 0
                try:
                    if row is not None and '12 Customs Clearance' in df_vtt.columns:
                        dias_final_day = int(row['12 Customs Clearance'])
                    elif row is not None and '12 Customs clearence' in df_vtt.columns:
                        dias_final_day = int(row['12 Customs clearence'])
                except Exception:
                    dias_final_day = 0
                day_plus_val = _coerce_to_int(row['Cust.']) if row is not None and 'Cust.' in df_vtt.columns else 0
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif source_step == 12:  # 12. Transport to plant
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                base_val = None
                if row is not None:
                    if '12 Customs Clearance' in df_vtt.columns and pd.notna(row.get('12 Customs Clearance')):
                        base_val = row.get('12 Customs Clearance')
                    elif '12 Customs clearence' in df_vtt.columns and pd.notna(row.get('12 Customs clearence')):
                        base_val = row.get('12 Customs clearence')
                if base_val is not None:
                    num_val = pd.to_numeric(base_val, errors='coerce')
                    if pd.isna(num_val):
                        try:
                            matches = re.findall(r"[-+]?\d*\.?\d+", str(base_val))
                            num_val = float(matches[0]) if matches else float('nan')
                        except Exception:
                            num_val = float('nan')
                    if pd.isna(num_val):
                        cell_content = "-"
                    else:
                        try:
                            cell_content = str(int(float(num_val)) + 1)
                        except Exception:
                            cell_content = "-"
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                if row is not None and 'Trpt POD/PFI vers Usine' in df_vtt.columns:
                    val = row['Trpt POD/PFI vers Usine']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '13 Transport to Plant' in df_vtt.columns:
                    val = row['13 Transport to Plant']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['13 Transport to Plant']) if row is not None and '13 Transport to Plant' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                day_plus_val = _coerce_to_int(row['Trpt POD/PFI vers Usine']) if row is not None and 'Trpt POD/PFI vers Usine' in df_vtt.columns else 0
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif source_step == 13:  # 13. Rounding
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '13 Transport to Plant' in df_vtt.columns:
                    base_val = row['13 Transport to Plant']
                    num_val = pd.to_numeric(base_val, errors='coerce')
                    if pd.isna(num_val):
                        try:
                            matches = re.findall(r"[-+]?\d*\.?\d+", str(base_val))
                            num_val = float(matches[0]) if matches else float('nan')
                        except Exception:
                            num_val = float('nan')
                    if pd.isna(num_val):
                        cell_content = "-"
                    else:
                        try:
                            cell_content = str(int(float(num_val)) + 1)
                        except Exception:
                            cell_content = "-"
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                val = None
                if row is not None:
                    if 'Round.' in df_vtt.columns:
                        val = row['Round.']
                    elif 'Round' in df_vtt.columns:
                        val = row['Round']
                if val is not None:
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '14 Rounding' in df_vtt.columns:
                    val = row['14 Rounding']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['14 Rounding']) if row is not None and '14 Rounding' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                day_plus_val = None
                if row is not None:
                    if 'Round.' in df_vtt.columns:
                        day_plus_val = _coerce_to_int(row['Round.'])
                    elif 'Round' in df_vtt.columns:
                        day_plus_val = _coerce_to_int(row['Round'])
                day_plus_val = day_plus_val if day_plus_val is not None else 0
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif source_step == 14:  # 14. Due Date
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '14 Rounding' in df_vtt.columns:
                    base_val = row['14 Rounding']
                    num_val = pd.to_numeric(base_val, errors='coerce')
                    if pd.isna(num_val):
                        try:
                            matches = re.findall(r"[-+]?\d*\.?\d+", str(base_val))
                            num_val = float(matches[0]) if matches else float('nan')
                        except Exception:
                            num_val = float('nan')
                    if pd.isna(num_val):
                        cell_content = "-"
                    else:
                        try:
                            cell_content = str(int(float(num_val)) + 1)
                        except Exception:
                            cell_content = "-"
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                cell_content = str(_due_date_day_plus_value(row, df_vtt))
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '15 Due Date' in df_vtt.columns:
                    val = row['15 Due Date']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['15 Due Date']) if row is not None and '15 Due Date' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                day_plus_val = _due_date_day_plus_value(row, df_vtt)
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif i == 15:  # 16. Manufacturing
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '15 Due Date' in df_vtt.columns:
                    base_val = row['15 Due Date']
                    num_val = pd.to_numeric(base_val, errors='coerce')
                    if pd.isna(num_val):
                        try:
                            matches = re.findall(r"[-+]?\d*\.?\d+", str(base_val))
                            num_val = float(matches[0]) if matches else float('nan')
                        except Exception:
                            num_val = float('nan')
                    if pd.isna(num_val):
                        cell_content = "-"
                    else:
                        try:
                            cell_content = str(int(float(num_val)) + 1)
                        except Exception:
                            cell_content = "-"
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                cell_content = "7"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '16 Manufacturing' in df_vtt.columns:
                    val = row['16 Manufacturing']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['16 Manufacturing']) if row is not None and '16 Manufacturing' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                day_plus_val = 7
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        elif i == 3:  # 4. Pack. prep. & load
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            elif j == 1:
                if row is not None and '4.1 Packaging préparation & loading' in df_vtt.columns:
                    val = row['4.1 Packaging préparation & loading']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 2:
                if row is not None and '4.2 Packaging préparation & loading' in df_vtt.columns:
                    val = row['4.2 Packaging préparation & loading']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j == 3:
                if row is not None and '4.3 Packaging préparation & loading' in df_vtt.columns:
                    val = row['4.3 Packaging préparation & loading']
                    if pd.isna(val):
                        cell_content = "-"
                    elif val == 0:
                        cell_content = "0"
                    else:
                        cell_content = str(val)
                else:
                    cell_content = "-"
                cell_content_visible = cell_content
            elif j >= 4:
                try:
                    dias_final_day = int(row['4.3 Packaging préparation & loading']) if row is not None and '4.3 Packaging préparation & loading' in df_vtt.columns else 0
                except Exception:
                    dias_final_day = 0
                day_plus_val = _coerce_to_int(row['4.2 Packaging préparation & loading']) if row is not None and '4.2 Packaging préparation & loading' in df_vtt.columns else 0
                paint_len = day_plus_val if (day_plus_val and day_plus_val > 0) else 1
                start_idx = max(1, dias_final_day - paint_len + 1)
                if start_idx <= (j-3) <= dias_final_day:
                    cell_content = ""
                    cell_style += "background-color:#90ee90;"
                cell_content_visible = cell_content
                cell_style_visible = cell_style
        else:
            if j == 0:
                cell_content = time_labels[i]
                cell_content_visible = cell_content
            else:
                cell_content = ""
                cell_content_visible = cell_content
        table_html += f"<td style='{cell_style}'>{cell_content}</td>"
        table_html_visible += f"<td style='{cell_style_visible}'>{cell_content_visible}</td>"
    table_html += "</tr>"
    table_html_visible += "</tr>"

table_html += "</tbody></table>"
table_html_visible += "</tbody></table>"
# Render visible table as before, but with a distinct id to avoid capture conflicts
wrapped_html_visible = (
    "<div class='vtt-panel'>"
    "<div class='vtt-panel__title'>Timeline Overview</div>"
    f"<div id='timeline_capture_table' style='display:inline-block'>{table_html_visible}</div>"
    "</div>"
)
st.markdown(wrapped_html_visible, unsafe_allow_html=True)



# --- KPIs al final ---
st.markdown("<hr style='margin:32px 0;'>", unsafe_allow_html=True)

"""Cálculo base de KPIs (vista numérica oculta, se usa para Gantt y export)."""
# CUSTOMER LEADTIME (CLT) toma el valor de '14 Rounding' (Final Day)

# POL>POD (Transit time + Time for security)
total_tt = None
if row is not None:
    t1 = pd.to_numeric(row.get("Transit time", None), errors="coerce") if "Transit time" in df_vtt.columns else None
    t2 = pd.to_numeric(row.get("Time for security", None), errors="coerce") if "Time for security" in df_vtt.columns else None
    parts = [v for v in (t1, t2) if v is not None and pd.notna(v)]
    if parts:
        total_tt = float(sum(parts))

# POD DETENTION (Customs clearence final day minus Days flexibility 1)
pod_det = None
try:
    if row is not None:
        customs_val = None
        flex1_val = None
        if '12 Customs Clearance' in df_vtt.columns:
            customs_val = _coerce_to_int(row.get('12 Customs Clearance'))
        elif '12 Customs clearence' in df_vtt.columns:
            customs_val = _coerce_to_int(row.get('12 Customs clearence'))
        if '10 Days flexibility 1' in df_vtt.columns:
            flex1_val = _coerce_to_int(row.get('10 Days flexibility 1'))
        if customs_val and flex1_val:
            pod_det = customs_val - flex1_val
except Exception:
    pod_det = None

# POD>PLANT (Rounding final day minus Customs clearence)
pod_plant = None
try:
    if row is not None:
        customs_val = None
        rounding_val = None
        if '12 Customs Clearance' in df_vtt.columns:
            customs_val = _coerce_to_int(row.get('12 Customs Clearance'))
        elif '12 Customs clearence' in df_vtt.columns:
            customs_val = _coerce_to_int(row.get('12 Customs clearence'))
        if '14 Rounding' in df_vtt.columns:
            rounding_val = _coerce_to_int(row.get('14 Rounding'))
        if rounding_val and customs_val:
            pod_plant = rounding_val - customs_val
except Exception:
    pod_plant = None

# --- KPI Gantt view (duraciones en formato barra de días) ---
def _final_day_for_step(i, row, df_vtt):
    try:
        if i == 0:
            return int(row['1 Day Customer Order']) if row is not None and '1 Day Customer Order' in df_vtt.columns else 0
        if i == 1:
            return int(row['2 Day ILN Order']) if row is not None and '2 Day ILN Order' in df_vtt.columns else 0
        if i == 2:
            return int(row['3.2 First Receipt Days']) if row is not None and '3.2 First Receipt Days' in df_vtt.columns else 0
        if i == 3:
            return int(row['4.3 Packaging préparation & loading']) if row is not None and '4.3 Packaging préparation & loading' in df_vtt.columns else 0
        if i == 4:
            return int(row['5.3 Transport ILN to POL']) if row is not None and '5.3 Transport ILN to POL' in df_vtt.columns else 0
        if i == 5:
            return int(row['6 First Day to POL']) if row is not None and '6 First Day to POL' in df_vtt.columns else 0
        if i == 6:
            return int(row['7 Cutt off']) if row is not None and '7 Cutt off' in df_vtt.columns else 0
        if i == 7:
            return int(row['8 ETD']) if row is not None and '8 ETD' in df_vtt.columns else 0
        if i == 8:
            if row is not None and '9 ETD> ETA' in df_vtt.columns:
                return int(row['9 ETD> ETA'])
            if row is not None and '9 ETD>ETA' in df_vtt.columns:
                return int(row['9 ETD>ETA'])
            return 0
        if i == 9:
            if row is not None and '10 Days flexibility 1' in df_vtt.columns and pd.notna(row['10 Days flexibility 1']):
                return int(row['10 Days flexibility 1'])
            base = None
            if row is not None and '9 ETD> ETA' in df_vtt.columns:
                base = row['9 ETD> ETA']
            elif row is not None and '9 ETD>ETA' in df_vtt.columns:
                base = row['9 ETD>ETA']
            bnum = pd.to_numeric(base, errors='coerce') if base is not None else float('nan')
            if pd.isna(bnum):
                m = re.findall(r"[-+]?\.?\d+", str(base)) if base is not None else []
                bnum = float(m[0]) if m else float('nan')
            plus = _coerce_to_int(row['Time for security']) if row is not None and 'Time for security' in df_vtt.columns else 0
            return int(float(bnum)) + 1 + int(plus) if not pd.isna(bnum) else 0
        if i == 10:
            return int(row['11 Days flexibility 2']) if row is not None and '11 Days flexibility 2' in df_vtt.columns else 0
        if i == 11:
            if row is not None and '12 Customs Clearance' in df_vtt.columns:
                return int(row['12 Customs Clearance'])
            if row is not None and '12 Customs clearence' in df_vtt.columns:
                return int(row['12 Customs clearence'])
            return 0
        if i == 12:
            return int(row['13 Transport to Plant']) if row is not None and '13 Transport to Plant' in df_vtt.columns else 0
        if i == 13:
            return int(row['14 Rounding']) if row is not None and '14 Rounding' in df_vtt.columns else 0
        if i == 14:
            return int(row['15 Due Date']) if row is not None and '15 Due Date' in df_vtt.columns else 0
        if i == 15:
            return int(row['16 Manufacturing']) if row is not None and '16 Manufacturing' in df_vtt.columns else 0
        return 0
    except Exception:
        return 0


def _step_start_index(step_index, row, df_vtt):
    try:
        final_day = _final_day_for_step(step_index, row, df_vtt)
        if not final_day:
            return 0

        day_plus = _day_plus_value_for_step(step_index, row, df_vtt)
        paint_len = day_plus if isinstance(day_plus, int) and day_plus > 0 else 1
        if step_index in (0, 1, 5, 6, 7):
            paint_len = 1
        return max(1, final_day - paint_len + 1)
    except Exception:
        return 0


def _day_plus_value_for_step(i, row, df_vtt):
    if i in (0, 1, 6, 7):
        return 0
    if i == 2:
        return _coerce_to_int(row['3 .1 Time of Recept in ILN']) if row is not None and '3 .1 Time of Recept in ILN' in df_vtt.columns else 0
    if i == 3:
        return _coerce_to_int(row['4.2 Packaging préparation & loading']) if row is not None and '4.2 Packaging préparation & loading' in df_vtt.columns else 0
    if i == 4:
        return _coerce_to_int(row['5.2 Transport ILN to POL']) if row is not None and '5.2 Transport ILN to POL' in df_vtt.columns else 0
    if i == 5:
        return _coerce_to_int(row['First Day to POL']) if row is not None and 'First Day to POL' in df_vtt.columns else 0
    if i == 8:
        return _coerce_to_int(row['Transit time']) if row is not None and 'Transit time' in df_vtt.columns else 0
    if i == 9:
        return _coerce_to_int(row['Time for security']) if row is not None and 'Time for security' in df_vtt.columns else 0
    if i == 10:
        return _coerce_to_int(row['Time for security2 buffer']) if row is not None and 'Time for security2 buffer' in df_vtt.columns else 0
    if i == 11:
        return _coerce_to_int(row['Cust.']) if row is not None and 'Cust.' in df_vtt.columns else 0
    if i == 12:
        return _coerce_to_int(row['Trpt POD/PFI vers Usine']) if row is not None and 'Trpt POD/PFI vers Usine' in df_vtt.columns else 0
    if i == 13:
        if row is not None and 'Round.' in df_vtt.columns:
            return _coerce_to_int(row['Round.'])
        if row is not None and 'Round' in df_vtt.columns:
            return _coerce_to_int(row['Round'])
        return 0
    if i == 14:
        return _due_date_day_plus_value(row, df_vtt)
    if i == 15:
        return 7
    return 0


def _build_kpi_rows(row, df_vtt):
    total_tt = None
    if row is not None:
        t1 = pd.to_numeric(row.get('Transit time', None), errors='coerce') if 'Transit time' in df_vtt.columns else None
        t2 = pd.to_numeric(row.get('Time for security', None), errors='coerce') if 'Time for security' in df_vtt.columns else None
        parts = [value for value in (t1, t2) if value is not None and pd.notna(value)]
        if parts:
            total_tt = float(sum(parts))

    pod_det = None
    try:
        if row is not None:
            customs_val = None
            flex1_val = None
            if '12 Customs Clearance' in df_vtt.columns:
                customs_val = _coerce_to_int(row.get('12 Customs Clearance'))
            elif '12 Customs clearence' in df_vtt.columns:
                customs_val = _coerce_to_int(row.get('12 Customs clearence'))
            if '10 Days flexibility 1' in df_vtt.columns:
                flex1_val = _coerce_to_int(row.get('10 Days flexibility 1'))
            if customs_val and flex1_val:
                pod_det = customs_val - flex1_val
    except Exception:
        pod_det = None

    pod_plant = None
    try:
        if row is not None:
            customs_val = None
            rounding_val = None
            if '12 Customs Clearance' in df_vtt.columns:
                customs_val = _coerce_to_int(row.get('12 Customs Clearance'))
            elif '12 Customs clearence' in df_vtt.columns:
                customs_val = _coerce_to_int(row.get('12 Customs clearence'))
            if '14 Rounding' in df_vtt.columns:
                rounding_val = _coerce_to_int(row.get('14 Rounding'))
            if rounding_val and customs_val:
                pod_plant = rounding_val - customs_val
    except Exception:
        pod_plant = None

    try:
        final_day_8 = _final_day_for_step(7, row, df_vtt)
        day_3 = _coerce_to_int(row['3 First Receipt Days']) if (row is not None and '3 First Receipt Days' in df_vtt.columns) else 0
        kpi_sup_pol = final_day_8 - day_3 + 1
    except Exception:
        day_3 = 0
        kpi_sup_pol = 0

    kpi_pol_pod = _coerce_to_int(total_tt) if total_tt is not None and not pd.isna(total_tt) else 0
    kpi_pod_det = _coerce_to_int(pod_det) if pod_det is not None else 0
    kpi_pod_plant = _coerce_to_int(pod_plant) if pod_plant is not None else 0

    start_sup = day_3 if day_3 > 0 else 0
    start_pol_pod = max(1, start_sup + kpi_sup_pol - 1) if (start_sup and kpi_sup_pol > 0) else 0
    start_pod_det = start_pol_pod + kpi_pol_pod if (start_pol_pod and kpi_pol_pod > 0) else 0
    start_pod_plant = start_pod_det + kpi_pod_det if (start_pod_det and kpi_pod_det > 0) else 0

    try:
        final_day_14 = _final_day_for_step(13, row, df_vtt)
        final_day_1 = _final_day_for_step(0, row, df_vtt)
        customer_leadtime = final_day_14 - final_day_1 + 1
    except Exception:
        customer_leadtime = 0

    try:
        final_day_14 = _final_day_for_step(13, row, df_vtt)
        day_5 = _coerce_to_int(row['5.1 Transport ILN to POL']) if (row is not None and '5.1 Transport ILN to POL' in df_vtt.columns) else 0
        transportation_duration = final_day_14 - day_5 + 1
    except Exception:
        transportation_duration = 0

    return [
        ("CUSTOMER LEADTIME (CLT)", customer_leadtime, _step_start_index(0, row, df_vtt)),
        ("OVS SAP STAGES", None, None),
        ("Transportation Duration", transportation_duration, _step_start_index(4, row, df_vtt)),
        ("SUPPLIER>POL", kpi_sup_pol, start_sup),
        ("POL>POD", kpi_pol_pod, start_pol_pod),
        ("POD DETENTION", kpi_pod_det, start_pod_det),
        ("POD>PLANT", kpi_pod_plant, start_pod_plant),
    ]

# Guardar HTML del VTT SUMMARY para reutilizarlo en la captura de imagen
kpi_gantt_html = ""
try:
    kpi_rows = _build_kpi_rows(row, df_vtt)

    # Escala de días: usar la misma línea de tiempo que la zona superior
    max_days_kpi = len(timeline_days)

    if max_days_kpi > 0:
        summary_ui_label_width = 200
        summary_ui_value_width = 50
        summary_ui_spacer_col_width = 50
        kpi_gantt_html = "<div class='vtt-panel' style='margin-top:16px;'><div class='vtt-panel__title'>VTT SUMMARY</div>"
        # Usar mismo tamaño base de fuente que la tabla superior
        kpi_gantt_html += "<table style='border-collapse:collapse; width:auto; font-size:12px;'>"

        # Cabecero de semanas alineado con la zona de tiempos (cálculo local)
        kpi_gantt_html += "<thead><tr>"
        # 2 columnas fijas para etiqueta y valor, mas 2 separadores invisibles que replican Day+ y Final Day
        kpi_gantt_html += (
            f"<th style='border:none; min-width:{summary_ui_label_width}px; width:{summary_ui_label_width}px; max-width:{summary_ui_label_width}px; padding:0;'></th>"
            f"<th style='border:none; min-width:{summary_ui_value_width}px; width:{summary_ui_value_width}px; max-width:{summary_ui_value_width}px; padding:0;'></th>"
            f"<th style='border:none; min-width:{summary_ui_spacer_col_width}px; width:{summary_ui_spacer_col_width}px; padding:0;'></th>"
            f"<th style='border:none; min-width:{summary_ui_spacer_col_width}px; width:{summary_ui_spacer_col_width}px; padding:0;'></th>"
        )
        current_week = None
        span_count = 0
        for idx, d_week in enumerate(timeline_days):
            w = d_week.isocalendar()[1]
            if current_week is None:
                current_week = w
                span_count = 1
            elif w == current_week:
                span_count += 1
            else:
                # Copiar estilo de cabecera de semanas de la tabla principal
                kpi_gantt_html += f"<th colspan='{span_count}' style='padding:0 1px; border:1px solid #eee; min-width:28px; text-align:center; background:#fffbe6; font-size:13.5px; font-weight:bold;'>W{current_week}</th>"
                current_week = w
                span_count = 1
        if current_week is not None and span_count > 0:
            kpi_gantt_html += f"<th colspan='{span_count}' style='padding:0 1px; border:1px solid #eee; min-width:28px; text-align:center; background:#fffbe6; font-size:13.5px; font-weight:bold;'>W{current_week}</th>"
        kpi_gantt_html += "</tr>"

        # Fila de días (M,T,W,...) también alineada
        kpi_gantt_html += "<tr>"
        # 2 columnas vacías equivalentes a etiqueta y valor, mas 2 separadores invisibles para alinear el inicio del calendario
        kpi_gantt_html += (
            f"<th style='border:none; min-width:{summary_ui_label_width}px; width:{summary_ui_label_width}px; max-width:{summary_ui_label_width}px; padding:0;'></th>"
            f"<th style='border:none; min-width:{summary_ui_value_width}px; width:{summary_ui_value_width}px; max-width:{summary_ui_value_width}px; padding:0;'></th>"
            f"<th style='border:none; min-width:{summary_ui_spacer_col_width}px; width:{summary_ui_spacer_col_width}px; padding:0;'></th>"
            f"<th style='border:none; min-width:{summary_ui_spacer_col_width}px; width:{summary_ui_spacer_col_width}px; padding:0;'></th>"
        )
        for d_day in timeline_days:
            if d_day.weekday() in (5, 6):
                th_style = "padding:0 1px; border:1px solid #eee; min-width:15px; width:18px; height:50px; text-align:center; background:#ffd6d6; font-size:12px; vertical-align:bottom;"
            else:
                th_style = "padding:0 1px; border:1px solid #eee; min-width:20px; width:20px; height:50px; text-align:center; background:#e3eafc; font-size:12px; vertical-align:bottom;"
            label_day = d_day.strftime('%a')[0].upper()
            # Usar la misma etiqueta vertical que la tabla principal
            kpi_gantt_html += f"<th style='{th_style}'><span class='vtt-vertical-text' style='display:flex;align-items:center;justify-content:center;height:100%;'>{label_day}</span></th>"
        kpi_gantt_html += "</tr></thead><tbody>"

        total_kpi_columns = 4 + len(timeline_days)

        for label_txt, val, start_day in kpi_rows:
            if val is None and start_day is None:
                kpi_gantt_html += (
                    "<tr>"
                    f"<td colspan='{total_kpi_columns}' style='padding:6px 4px 6px 2cm; border:1px solid #1f4e79; text-align:left; font-weight:bold; color:#ffffff; background:#1f4e79; min-width:200px; white-space:nowrap; font-size:18px;'>"
                    f"{label_txt}</td>"
                    "</tr>"
                )
                continue

            kpi_gantt_html += "<tr>"
            # Etiqueta KPI (columna Steps)
            kpi_gantt_html += (
                f"<td style='padding:1px 4px; border:1px solid #eee; text-align:left; font-weight:bold; background:#f5f5f5; min-width:{summary_ui_label_width}px; width:{summary_ui_label_width}px; max-width:{summary_ui_label_width}px; white-space:nowrap; height:15px; line-height:15px; font-size:14px;'>"
                f"{label_txt}</td>"
            )
            # Valor numérico (columna Day)
            display_val = str(val)  # Mostrar siempre el valor, incluso si es 0 o negativo, para depuración
            kpi_gantt_html += (
                f"<td style='padding:1px 4px; border:1px solid #eee; text-align:center; min-width:{summary_ui_value_width}px; width:{summary_ui_value_width}px; max-width:{summary_ui_value_width}px; height:15px; line-height:15px; font-size:14px;'>"
                f"{display_val}</td>"
            )
            # Separadores invisibles para mantener alineado el inicio de semanas con Day+ y Final Day de la tabla superior
            kpi_gantt_html += (
                f"<td style='padding:0; border:none; min-width:{summary_ui_spacer_col_width}px; width:{summary_ui_spacer_col_width}px; height:15px; background:transparent;'></td>"
                f"<td style='padding:0; border:none; min-width:{summary_ui_spacer_col_width}px; width:{summary_ui_spacer_col_width}px; height:15px; background:transparent;'></td>"
            )
            # Barras de días (Gantt secuencial) alineadas con timeline_days
            for idx, _day in enumerate(timeline_days, start=1):
                if val and val > 0 and start_day:
                    end_day = start_day + val - 1
                    is_active = start_day <= idx <= end_day
                else:
                    is_active = False

                if is_active:
                    # Usar azul claro y barco para POL>POD y verde para el resto
                    bg = "#4a90e2" if label_txt == "POL>POD" else "#90ee90"
                    content = "<span style='color:#ffffff; font-size:12px; line-height:1;'>🚢</span>" if label_txt == "POL>POD" else ""
                else:
                    bg = "#ffffff"
                    content = ""

                # Altura y ancho similares a las celdas de días de la tabla superior
                kpi_gantt_html += (
                    f"<td style='border:1px solid #f0f0f0; width:20px; height:15px; padding:0 1px; background:{bg}; text-align:center; vertical-align:middle;'>{content}</td>"
                )
            kpi_gantt_html += "</tr>"
        kpi_gantt_html += "</tbody></table></div>"
        st.markdown(kpi_gantt_html, unsafe_allow_html=True)
except Exception:
    # Si algo falla, no romper la app; simplemente no mostrar el gantt de KPIs
    pass

# Mostrar Customer Safety STOCK debajo del Gantt de KPIs
if safety_stock_val is not None:
    st.markdown(
        f"""
        <div class='vtt-kpi-row'>
            <div class='vtt-kpi-card'>
                <div class='vtt-kpi-card__label'>Customer Safety STOCK</div>
                <div class='vtt-kpi-card__value'>{safety_stock_val}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# Controles de Timeline al final (sin mover la tabla de gantt)
st.markdown("<div class='vtt-section-title'>Timeline Controls</div>", unsafe_allow_html=True)
st.slider(
    "Days to Show",
    min_value=7,
    max_value=150,
    value=st.session_state.get("days_slider_timeline", 110),
    step=1,
    key="days_slider_timeline",
)

# Build an off-screen composite capture area that includes table + KPIs + selection context
capture_pol = st.session_state.get('pol_select','')
capture_pod = st.session_state.get('pod_select','')
capture_days = st.session_state.get('days_slider_timeline', 100)
composite_html = ""
composite_html += "<div id='timeline_capture' style='position:absolute; left:-100000px; top:0; background:#fff; padding:8px; font-family:Arial, sans-serif; display:inline-block; width:max-content; max-width:none; overflow:visible;'>"
composite_html += "<div style='font-size:22px; font-weight:700; margin-bottom:8px;'>VTT View</div>"
composite_html += f"<div style='margin-bottom:8px;'><b>POL:</b> {capture_pol} &nbsp;&nbsp; <b>POD:</b> {capture_pod} &nbsp;&nbsp; <b>Days to Show:</b> {capture_days}</div>"

# Add ID, Carrier, Shipper, ILN/FF, PLANT, E/D y Commodity en la cabecera de la captura
_id_val = _carrier_val = _shipper_val = _iln_val = _plant_val = _commodity_val = _ed_val = ""
if row is not None:
    try:
        _id_val = str(row.get('ID', '')) if 'ID' in df_vtt.columns else ''
    except Exception:
        _id_val = ''
    try:
        _carrier_val = str(row.get('Carrier', '')) if 'Carrier' in df_vtt.columns else ''
    except Exception:
        _carrier_val = ''
    # Shipper from column K (index 10) if present
    try:
        _shipper_col = df_vtt.columns[10] if len(df_vtt.columns) > 10 else None
        _shipper_val = str(row.get(_shipper_col, '')) if _shipper_col and _shipper_col in df_vtt.columns else ''
    except Exception:
        _shipper_val = ''
    # ILN from column I (index 8) if present
    try:
        _iln_col = df_vtt.columns[8] if len(df_vtt.columns) > 8 else None
        _iln_val = str(row.get(_iln_col, '')) if _iln_col and _iln_col in df_vtt.columns else ''
    except Exception:
        _iln_val = ''
    try:
        _plant_val = str(row.get('Name Destin Site', '')) if 'Name Destin Site' in df_vtt.columns else ''
    except Exception:
        _plant_val = ''
    # Commodity (admite nombre de columna 'Commodity' o 'Comodity')
    try:
        if 'Commodity' in df_vtt.columns:
            _commodity_val = str(row.get('Commodity', ''))
        elif 'Comodity' in df_vtt.columns:
            _commodity_val = str(row.get('Comodity', ''))
        else:
            _commodity_val = ''
    except Exception:
        _commodity_val = ''

_ed_val = _format_expiration_date(row, df_vtt)

composite_html += "<div style='display:grid; grid-template-columns: repeat(7, minmax(150px, 1fr)); gap:12px; align-items:start; margin:6px 0 10px 0;'>"
composite_html += render_box('ID', _id_val)
composite_html += render_box('Carrier', _carrier_val)
composite_html += render_box('Shipper', _shipper_val)
composite_html += render_box('ILN/FF', _iln_val)
composite_html += render_box('PLANT', _plant_val)
composite_html += render_box('E/D', _ed_val)
composite_html += render_box('Commodity', _commodity_val)
composite_html += "</div>"

# Wrap the table to allow full-width capture (no fixed width)
composite_html += f"<div style='display:inline-block; width:max-content; overflow:visible;'>{table_html}</div>"

composite_html += "<hr style='margin:16px 0;'>"

# Incluir el mismo Gantt de KPIs (VTT SUMMARY) que se ve en la UI (ya incluye su propio título)
try:
    if kpi_gantt_html:
        composite_html += kpi_gantt_html
except Exception:
    pass

# Mostrar Customer Safety STOCK debajo del VTT SUMMARY en la captura, igual que en la UI
if safety_stock_val is not None:
    composite_html += "<div class='vtt-kpi-row' style='margin-top:12px;'>"
    composite_html += "<div class='vtt-kpi-card'>"
    composite_html += "<div class='vtt-kpi-card__label'>Customer Safety STOCK</div>"
    composite_html += f"<div class='vtt-kpi-card__value'>{safety_stock_val}</div>"
    composite_html += "</div>"
    composite_html += "</div>"

composite_html += "</div>"  # end capture root
st.markdown(composite_html, unsafe_allow_html=True)

# --- Descargar Excel con la visualización completa ---
def _hex_to_fill(hex_color):
    if not hex_color:
        return None
    h = hex_color.lstrip('#')
    if len(h) == 6:
        h = 'FF' + h.upper()
    return PatternFill(fill_type='solid', start_color=h, end_color=h)

def _compute_week_spans(days):
    spans = []


    def _load_snapshot_font(size, bold=False):
        candidates = []
        if os.name == 'nt':
            font_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
            candidates.extend([
                os.path.join(font_dir, 'arialbd.ttf' if bold else 'arial.ttf'),
                os.path.join(font_dir, 'calibrib.ttf' if bold else 'calibri.ttf'),
            ])
        candidates.extend(['arialbd.ttf' if bold else 'arial.ttf', 'DejaVuSans-Bold.ttf' if bold else 'DejaVuSans.ttf'])
        for candidate in candidates:
            try:
                return ImageFont.truetype(candidate, size)
            except Exception:
                continue
        return ImageFont.load_default()


    def _text_size(draw, text, font):
        left, top, right, bottom = draw.textbbox((0, 0), str(text), font=font)
        return right - left, bottom - top


    def _draw_centered_text(draw, box, text, font, fill):
        x1, y1, x2, y2 = box
        text_w, text_h = _text_size(draw, text, font)
        draw.text((x1 + max(0, (x2 - x1 - text_w) / 2), y1 + max(0, (y2 - y1 - text_h) / 2)), str(text), font=font, fill=fill)


    def _draw_cell(draw, box, text='', *, fill='#ffffff', outline='#dddddd', font=None, text_fill='#111111', align='left'):
        draw.rectangle(box, fill=fill, outline=outline, width=1)
        if text == '':
            return
        x1, y1, x2, y2 = box
        text = str(text)
        font = font or _load_snapshot_font(12)
        text_w, text_h = _text_size(draw, text, font)
        if align == 'center':
            tx = x1 + max(0, (x2 - x1 - text_w) / 2)
        else:
            tx = x1 + 6
        ty = y1 + max(0, (y2 - y1 - text_h) / 2)
        draw.text((tx, ty), text, font=font, fill=text_fill)


    def _snapshot_info_pairs(row, df_vtt):
        shipper_col = df_vtt.columns[10] if len(df_vtt.columns) > 10 else None
        iln_col = df_vtt.columns[8] if len(df_vtt.columns) > 8 else None
        commodity_col = 'Commodity' if 'Commodity' in df_vtt.columns else ('Comodity' if 'Comodity' in df_vtt.columns else None)

        exp_value = ''
        try:
            if row is not None and 'Expiration Date' in df_vtt.columns:
                raw_exp = row.get('Expiration Date', '')
                if pd.notnull(raw_exp):
                    if isinstance(raw_exp, (pd.Timestamp, datetime)):
                        exp_value = raw_exp.strftime('%d/%m/%Y')
                    else:
                        exp_value = pd.to_datetime(raw_exp).strftime('%d/%m/%Y')
        except Exception:
            exp_value = str(row.get('Expiration Date', '')) if row is not None else ''

        info_spec = [
            ('ID', 'ID'),
            ('Carrier', 'Carrier'),
            ('Shipper', shipper_col),
            ('ILN/FF', iln_col),
            ('PLANT', 'Name Destin Site'),
        ]

        pairs = []
        for label, column in info_spec:
            value = ''
            try:
                if row is not None and column and column in df_vtt.columns:
                    value = row.get(column, '')
            except Exception:
                value = ''
            pairs.append((label, '' if pd.isna(value) else str(value)))
        commodity_value = ''
        try:
            if row is not None and commodity_col and commodity_col in df_vtt.columns:
                commodity_value = row.get(commodity_col, '')
        except Exception:
            commodity_value = ''
        pairs.append(('E/D', exp_value))
        pairs.append(('Commodity', '' if pd.isna(commodity_value) else str(commodity_value)))
        return pairs
    current_week = None
    count = 0
    for d in days:
        w = d.isocalendar()[1]
        if current_week is None:
            current_week = w
            count = 1
        elif w == current_week:
            count += 1
        else:
            spans.append((current_week, count))
            current_week = w
            count = 1
    if current_week is not None:
        spans.append((current_week, count))
    return spans


def _load_snapshot_font(size, bold=False):
    candidates = []
    if os.name == 'nt':
        font_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
        candidates.extend([
            os.path.join(font_dir, 'arialbd.ttf' if bold else 'arial.ttf'),
            os.path.join(font_dir, 'calibrib.ttf' if bold else 'calibri.ttf'),
        ])
    candidates.extend(['arialbd.ttf' if bold else 'arial.ttf', 'DejaVuSans-Bold.ttf' if bold else 'DejaVuSans.ttf'])
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _text_size(draw, text, font):
    left, top, right, bottom = draw.textbbox((0, 0), str(text), font=font)
    return right - left, bottom - top


def _draw_cell(draw, box, text='', *, fill='#ffffff', outline='#dddddd', font=None, text_fill='#111111', align='left'):
    draw.rectangle(box, fill=fill, outline=outline, width=1)
    if text == '':
        return
    x1, y1, x2, y2 = box
    text = str(text)
    font = font or _load_snapshot_font(12)
    text_w, text_h = _text_size(draw, text, font)
    if align == 'center':
        tx = x1 + max(0, (x2 - x1 - text_w) / 2)
    else:
        tx = x1 + 6
    ty = y1 + max(0, (y2 - y1 - text_h) / 2)
    draw.text((tx, ty), text, font=font, fill=text_fill)


def _snapshot_info_pairs(row, df_vtt):
    shipper_col = df_vtt.columns[10] if len(df_vtt.columns) > 10 else None
    iln_col = df_vtt.columns[8] if len(df_vtt.columns) > 8 else None
    commodity_col = 'Commodity' if 'Commodity' in df_vtt.columns else ('Comodity' if 'Comodity' in df_vtt.columns else None)

    exp_value = ''
    try:
        if row is not None and 'Expiration Date' in df_vtt.columns:
            raw_exp = row.get('Expiration Date', '')
            if pd.notnull(raw_exp):
                if isinstance(raw_exp, (pd.Timestamp, datetime)):
                    exp_value = raw_exp.strftime('%d/%m/%Y')
                else:
                    exp_value = pd.to_datetime(raw_exp).strftime('%d/%m/%Y')
    except Exception:
        exp_value = str(row.get('Expiration Date', '')) if row is not None else ''

    info_spec = [
        ('ID', 'ID'),
        ('Carrier', 'Carrier'),
        ('Shipper', shipper_col),
        ('ILN/FF', iln_col),
        ('PLANT', 'Name Destin Site'),
    ]

    pairs = []
    for label, column in info_spec:
        value = ''
        try:
            if row is not None and column and column in df_vtt.columns:
                value = row.get(column, '')
        except Exception:
            value = ''
        pairs.append((label, '' if pd.isna(value) else str(value)))
    commodity_value = ''
    try:
        if row is not None and commodity_col and commodity_col in df_vtt.columns:
            commodity_value = row.get(commodity_col, '')
    except Exception:
        commodity_value = ''
    pairs.append(('E/D', exp_value))
    pairs.append(('Commodity', '' if pd.isna(commodity_value) else str(commodity_value)))
    return pairs


def _build_snapshot_image(row, df_vtt, selected_pol, selected_pod, time_labels, headers, timeline_days, scale=2, font_multiplier=1):
    def s(value):
        return max(1, int(round(value * scale)))

    def fs(value):
        return max(1, int(round(value * scale * font_multiplier)))

    margin = s(24)
    title_h = fs(68)
    subtitle_h = fs(48)
    info_h = fs(48)
    section_gap = s(18)
    week_h = fs(36)
    date_h = fs(36)
    row_h = fs(34)
    label_w = s(360)
    metric_w = s(88)
    final_w = s(108)
    day_w = s(28)
    fixed_widths = [label_w, metric_w, metric_w, final_w]
    table_left = margin
    grid_left = table_left + sum(fixed_widths)
    total_width = margin * 2 + sum(fixed_widths) + day_w * len(timeline_days)
    kpi_rows = _build_kpi_rows(row, df_vtt)
    total_height = (
        margin * 2 + title_h + subtitle_h + info_h * 3 + section_gap +
        week_h + date_h + row_h * len(time_labels) + section_gap +
        title_h + week_h + date_h + row_h * len(kpi_rows) +
        section_gap + max(s(120), fs(80))
    )

    image = PILImage.new('RGB', (max(total_width, s(1200)), total_height), '#ffffff')
    draw = ImageDraw.Draw(image)
    font_title = _load_snapshot_font(fs(40), bold=True)
    font_heading = _load_snapshot_font(fs(28), bold=True)
    font_bold = _load_snapshot_font(fs(24), bold=True)
    font_text = _load_snapshot_font(fs(24))
    font_small = _load_snapshot_font(s(20))

    draw.text((margin, margin), 'VTT View', font=font_title, fill='#111111')
    draw.text((margin, margin + title_h), f'POL: {selected_pol}   POD: {selected_pod}   Days to Show: {len(timeline_days)}', font=font_heading, fill='#222222')

    info_y = margin + title_h + subtitle_h
    info_pairs = _snapshot_info_pairs(row, df_vtt)
    for index, (label, value) in enumerate(info_pairs):
        col = index % 3
        row_index = index // 3
        x = margin + col * ((image.width - margin * 2) // 3)
        y = info_y + row_index * info_h
        label_text = f'{label}:'
        draw.text((x, y), label_text, font=font_bold, fill='#111111')
        label_width, _ = _text_size(draw, label_text, font_bold)
        draw.text((x + label_width + s(24), y), value, font=font_text, fill='#333333')

    y = info_y + info_h * 3 + s(8)
    week_spans = _compute_week_spans(timeline_days)
    summary_fixed_widths = [label_w, metric_w]
    summary_grid_left = table_left + sum(summary_fixed_widths)

    x = summary_grid_left
    for week, span in week_spans:
        _draw_cell(draw, (x, y, x + span * day_w, y + week_h), f'W{week}', fill='#fffbe6', font=font_bold, align='center')
        x += span * day_w
    y += week_h

    x = table_left
    for width, header in zip(fixed_widths, headers):
        _draw_cell(draw, (x, y, x + width, y + date_h), header, fill='#f5f5f5', font=font_bold, align='center' if header != 'Steps' else 'left')
        x += width
    for day in timeline_days:
        fill = '#ffd6d6' if day.weekday() in (5, 6) else '#e3eafc'
        _draw_cell(draw, (x, y, x + day_w, y + date_h), day.strftime('%d'), fill=fill, font=font_small, align='center')
        x += day_w
    y += date_h

    for index, label in enumerate(time_labels):
        x = table_left
        _draw_cell(draw, (x, y, x + label_w, y + row_h), label, fill='#f5f5f5', font=font_bold)
        x += label_w
        _draw_cell(draw, (x, y, x + metric_w, y + row_h), _day_value_for_step(index, row, df_vtt), font=font_text, align='center')
        x += metric_w
        day_plus = _day_plus_for_step(index, row, df_vtt)
        _draw_cell(draw, (x, y, x + metric_w, y + row_h), str(day_plus) if day_plus != 0 else '0', font=font_text, align='center')
        x += metric_w
        final_day = _final_day_for_step(index, row, df_vtt)
        _draw_cell(draw, (x, y, x + final_w, y + row_h), str(final_day) if final_day else '-', font=font_text, align='center')
        x += final_w

        paint_len = day_plus if isinstance(day_plus, int) and day_plus > 0 else 1
        if index in (0, 1, 5, 6, 7):
            paint_len = 1
        start_idx = max(1, final_day - paint_len + 1) if final_day else 0
        for day_index, day in enumerate(timeline_days, start=1):
            fill = '#ffd6d6' if day.weekday() in (5, 6) else '#ffffff'
            if final_day and start_idx <= day_index <= final_day:
                fill = '#4a90e2' if index == 8 else '#90ee90'
            _draw_cell(draw, (x, y, x + day_w, y + row_h), fill=fill)
            x += day_w
        y += row_h

    y += section_gap
    draw.text((margin, y), 'VTT SUMMARY', font=font_title, fill='#111111')
    y += title_h

    x = grid_left
    for week, span in week_spans:
        _draw_cell(draw, (x, y, x + span * day_w, y + week_h), f'W{week}', fill='#fffbe6', font=font_bold, align='center')
        x += span * day_w
    y += week_h

    x = table_left
    for width in summary_fixed_widths:
        _draw_cell(draw, (x, y, x + width, y + date_h), fill='#ffffff')
        x += width
    for day in timeline_days:
        fill = '#ffd6d6' if day.weekday() in (5, 6) else '#e3eafc'
        _draw_cell(draw, (x, y, x + day_w, y + date_h), day.strftime('%a')[0].upper(), fill=fill, font=font_small, align='center')
        x += day_w
    y += date_h

    for label, value, start_day in kpi_rows:
        if label == 'OVS SAP STAGES' and value is None and start_day is None:
            _draw_cell(
                draw,
                (table_left, y, table_left + sum(summary_fixed_widths) + day_w * len(timeline_days), y + row_h),
                label,
                fill='#1f4e79',
                outline='#1f4e79',
                font=font_bold,
                text_fill='#ffffff',
                align='left',
            )
            y += row_h
            continue
        x = table_left
        _draw_cell(draw, (x, y, x + label_w, y + row_h), label, fill='#f5f5f5', font=font_bold)
        x += label_w
        _draw_cell(draw, (x, y, x + metric_w, y + row_h), value if value else '-', font=font_text, align='center')
        x += metric_w
        end_day = start_day + value - 1 if value and start_day else 0
        for day_index, _day in enumerate(timeline_days, start=1):
            fill = '#ffffff'
            if value and start_day and start_day <= day_index <= end_day:
                fill = '#4a90e2' if label == 'POL>POD' else '#90ee90'
            _draw_cell(draw, (x, y, x + day_w, y + row_h), fill=fill)
            x += day_w
        y += row_h

    if row is not None and 'Safety stock' in df_vtt.columns:
        y += section_gap
        safety_stock_label = 'Customer Safety STOCK'
        safety_stock_value = str(row['Safety stock'])
        card_pad_x = s(22)
        card_gap = s(24)
        badge_pad_x = s(18)
        badge_pad_y = s(10)
        label_width, label_height = _text_size(draw, safety_stock_label, font_bold)
        value_width, value_height = _text_size(draw, safety_stock_value, font_heading)
        badge_width = max(s(72), value_width + badge_pad_x * 2)
        badge_height = max(s(52), value_height + badge_pad_y * 2)
        card_height = max(s(74), label_height + s(28))
        card_width = label_width + badge_width + card_gap + card_pad_x * 2
        card_box = (margin, y, margin + card_width, y + card_height)
        draw.rounded_rectangle(card_box, radius=s(18), fill='#f3f8ff', outline='#d7e2f0', width=max(1, s(1)))

        label_x = margin + card_pad_x
        label_y = y + (card_height - label_height) // 2
        draw.text((label_x, label_y), safety_stock_label, font=font_bold, fill='#102845')

        badge_x1 = card_box[2] - card_pad_x - badge_width
        badge_y1 = y + (card_height - badge_height) // 2
        badge_x2 = badge_x1 + badge_width
        badge_y2 = badge_y1 + badge_height
        draw.rounded_rectangle((badge_x1, badge_y1, badge_x2, badge_y2), radius=s(14), fill='#ffffff', outline='#d7e2f0', width=max(1, s(1)))
        value_x = badge_x1 + (badge_width - value_width) // 2
        value_y = badge_y1 + (badge_height - value_height) // 2 - s(1)
        draw.text((value_x, value_y), safety_stock_value, font=font_heading, fill='#102845')

    return image


def _build_snapshot_png_bytes(row, df_vtt, selected_pol, selected_pod, time_labels, headers, timeline_days, scale=2, font_multiplier=1):
    snapshot_image = _build_snapshot_image(
        row=row,
        df_vtt=df_vtt,
        selected_pol=selected_pol,
        selected_pod=selected_pod,
        time_labels=time_labels,
        headers=headers,
        timeline_days=timeline_days,
        scale=scale,
        font_multiplier=font_multiplier,
    )
    image_buffer = BytesIO()
    snapshot_image.save(image_buffer, format='PNG', optimize=True)
    return image_buffer.getvalue()

def _get_value_safe(val):
    if pd.isna(val):
        return "-"
    try:
        return str(val)
    except Exception:
        return "-"

def _final_day_for_step(i, row, df_vtt):
    try:
        if i == 0:
            return int(row['1 Day Customer Order']) if row is not None and '1 Day Customer Order' in df_vtt.columns else 0
        if i == 1:
            return int(row['2 Day ILN Order']) if row is not None and '2 Day ILN Order' in df_vtt.columns else 0
        if i == 2:
            return int(row['3.2 First Receipt Days']) if row is not None and '3.2 First Receipt Days' in df_vtt.columns else 0
        if i == 3:
            return int(row['4.3 Packaging préparation & loading']) if row is not None and '4.3 Packaging préparation & loading' in df_vtt.columns else 0
        if i == 4:
            return int(row['5.3 Transport ILN to POL']) if row is not None and '5.3 Transport ILN to POL' in df_vtt.columns else 0
        if i == 5:
            return int(row['6 First Day to POL']) if row is not None and '6 First Day to POL' in df_vtt.columns else 0
        if i == 6:
            return int(row['7 Cutt off']) if row is not None and '7 Cutt off' in df_vtt.columns else 0
        if i == 7:
            return int(row['8 ETD']) if row is not None and '8 ETD' in df_vtt.columns else 0
        if i == 8:
            if row is not None and '9 ETD> ETA' in df_vtt.columns:
                return int(row['9 ETD> ETA'])
            if row is not None and '9 ETD>ETA' in df_vtt.columns:
                return int(row['9 ETD>ETA'])
            return 0
        if i == 9:
            if row is not None and '10 Days flexibility 1' in df_vtt.columns and pd.notna(row['10 Days flexibility 1']):
                return int(row['10 Days flexibility 1'])
            # derive: base(9) + 1 + buffer
            base = None
            if row is not None and '9 ETD> ETA' in df_vtt.columns:
                base = row['9 ETD> ETA']
            elif row is not None and '9 ETD>ETA' in df_vtt.columns:
                base = row['9 ETD>ETA']
            bnum = pd.to_numeric(base, errors='coerce') if base is not None else float('nan')
            if pd.isna(bnum):
                # FIX: regex string was split across lines, causing unterminated string literal error
                m = re.findall(r"[-+]?\.?\d+", str(base)) if base is not None else []
                bnum = float(m[0]) if m else float('nan')
            plus = _coerce_to_int(row['Time for security']) if row is not None and 'Time for security' in df_vtt.columns else 0
            return int(float(bnum)) + 1 + int(plus) if not pd.isna(bnum) else 0
        if i == 10:
            return int(row['11 Days flexibility 2']) if row is not None and '11 Days flexibility 2' in df_vtt.columns else 0
        if i == 11:
            if row is not None and '12 Customs Clearance' in df_vtt.columns:
                return int(row['12 Customs Clearance'])
            if row is not None and '12 Customs clearence' in df_vtt.columns:
                return int(row['12 Customs clearence'])
            return 0
        if i == 12:
            return int(row['13 Transport to Plant']) if row is not None and '13 Transport to Plant' in df_vtt.columns else 0
        if i == 13:
            return int(row['14 Rounding']) if row is not None and '14 Rounding' in df_vtt.columns else 0
        if i == 14:
            return int(row['15 Due Date']) if row is not None and '15 Due Date' in df_vtt.columns else 0
        if i == 15:
            return int(row['16 Manufacturing']) if row is not None and '16 Manufacturing' in df_vtt.columns else 0
        return 0
    except Exception:
        return 0

def _day_plus_for_step(i, row, df_vtt):
    return _day_plus_value_for_step(i, row, df_vtt)

def _day_value_for_step(i, row, df_vtt):
    # Returns the Day column value per step as string
    try:
        if i == 0:
            return _get_value_safe(row['1 Day Customer Order']) if row is not None and '1 Day Customer Order' in df_vtt.columns else '-'
        if i == 1:
            return _get_value_safe(row['2 Day ILN Order']) if row is not None and '2 Day ILN Order' in df_vtt.columns else '-'
        if i == 2:
            return _get_value_safe(row['3 First Receipt Days']) if row is not None and '3 First Receipt Days' in df_vtt.columns else 'No hay datos para la combinación POL/POD seleccionada'
        if i == 3:
            return _get_value_safe(row['4.1 Packaging préparation & loading']) if row is not None and '4.1 Packaging préparation & loading' in df_vtt.columns else '-'
        if i == 4:
            return _get_value_safe(row['5.1 Transport ILN to POL']) if row is not None and '5.1 Transport ILN to POL' in df_vtt.columns else '-'
        if i == 5:
            return _get_value_safe(row['6 First Day to POL']) if row is not None and '6 First Day to POL' in df_vtt.columns else '-'
        if i == 6:
            return _get_value_safe(row['7 Cutt off']) if row is not None and '7 Cutt off' in df_vtt.columns else '-'
        if i == 7:
            return _get_value_safe(row['8 ETD']) if row is not None and '8 ETD' in df_vtt.columns else '-'
        if i == 8:
            # Day for step 9 is ETD
            return _get_value_safe(row['8 ETD']) if row is not None and '8 ETD' in df_vtt.columns else '-'
        if i == 9:
            # base final of 9 + 1
            base = _final_day_for_step(8, row, df_vtt)
            return str(base + 1) if base else '-'
        if i == 10:
            base = _final_day_for_step(9, row, df_vtt)
            return str(base + 1) if base else '-'
        if i == 11:
            base = _final_day_for_step(10, row, df_vtt)
            return str(base + 1) if base else '-'
        if i == 12:
            base = _final_day_for_step(11, row, df_vtt)
            return str(base + 1) if base else '-'
        if i == 13:
            base = _final_day_for_step(12, row, df_vtt)
            return str(base + 1) if base else '-'
        if i == 14:
            base = _final_day_for_step(13, row, df_vtt)
            return str(base + 1) if base else '-'
        if i == 15:
            base = _final_day_for_step(14, row, df_vtt)
            return str(base + 1) if base else '-'
        return '-'
    except Exception:
        return '-'


def build_excel_workbook(row, df_vtt, selected_pol, selected_pod, time_labels, headers, timeline_days, include_snapshot_sheet=True):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Timeline'
    ws.sheet_view.showGridLines = False

    # styles
    bold = Font(bold=True)
    hfill = _hex_to_fill('#f5f5f5')
    weekfill = _hex_to_fill('#fffbe6')
    weekendfill = _hex_to_fill('#ffd6d6')
    weekdayfill = _hex_to_fill('#e3eafc')
    paintfill = _hex_to_fill('#90ee90')
    # Azul más claro para Transit Duration (ETD>ETA) en Excel para que coincida con la vista HTML
    darkbluefill = _hex_to_fill('#4a90e2')
    border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'), top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))

    r = 1
    ws.cell(row=r, column=1, value='POL:').font = bold; ws.cell(row=r, column=2, value=selected_pol)
    ws.cell(row=r, column=3, value='POD:').font = bold; ws.cell(row=r, column=4, value=selected_pod)
    r += 1
    # Info row if available
    if row is not None:
        info_pairs = []
        # Detectar columna de Commodity/Comodity si existe
        commodity_col = None
        if 'Commodity' in df_vtt.columns:
            commodity_col = 'Commodity'
        elif 'Comodity' in df_vtt.columns:
            commodity_col = 'Comodity'

        for label, colname in [
            ('ID','ID'),
            ('Carrier','Carrier'),
            ('Shipper', df_vtt.columns[10] if len(df_vtt.columns) > 10 else None),
            ('ILN/FF', df_vtt.columns[8] if len(df_vtt.columns) > 8 else None),
            ('PLANT','Name Destin Site'),
            ('Commodity', commodity_col),
            ('E/D','Expiration Date')
        ]:
            val = row.get(colname, '') if (colname and colname in df_vtt.columns) else ''
            info_pairs.append((label, val))
        c = 1
        for label, val in info_pairs:
            ws.cell(row=r, column=c, value=f'{label}:').font = bold
            ws.cell(row=r, column=c+1, value=str(val))
            c += 2
        r += 1

    r += 1
    # Week header merged cells
    start_col = 5  # dates start at column 5
    ws.cell(row=r, column=1, value='')
    # leave placeholders for Steps/Day/Day+/Final Day
    spans = _compute_week_spans(timeline_days)
    c = start_col
    for week, span in spans:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
        cell = ws.cell(row=r, column=c, value=f'W{week}')
        cell.fill = weekfill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        # borders
        for cc in range(c, c+span):
            ws.cell(row=r, column=cc).border = border
        c += span
    r += 1

    # Header row (Steps, Day, Day+, Final Day, then dates)
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.fill = hfill
        cell.font = bold
        cell.border = border
        cell.alignment = Alignment(horizontal='center') if ci > 1 else Alignment(horizontal='left')
    for idx, d in enumerate(timeline_days):
        ci = start_col + idx
        cell = ws.cell(row=r, column=ci, value=d.strftime('%d-%b'))
        cell.fill = weekendfill if d.weekday() in (5,6) else weekdayfill
        cell.border = border
        # Rotate text vertically (top-to-bottom)
        cell.alignment = Alignment(horizontal='center', vertical='bottom', textRotation=90)
    r += 1

    # Row content
    for i, label in enumerate(time_labels):
        # Reduce Excel row height for step rows (~35% smaller than default ~15pt)
        try:
            ws.row_dimensions[r+i].height = 10.5
        except Exception:
            pass
        ws.cell(row=r+i, column=1, value=label).fill = hfill
        ws.cell(row=r+i, column=1).font = bold
        ws.cell(row=r+i, column=1).border = border
        ws.cell(row=r+i, column=1).alignment = Alignment(horizontal='left')

        # Day
        day_val = _day_value_for_step(i, row, df_vtt)
        ws.cell(row=r+i, column=2, value=day_val).border = border
        ws.cell(row=r+i, column=2).alignment = Alignment(horizontal='center')

        # Day+
        day_plus = _day_plus_for_step(i, row, df_vtt)
        ws.cell(row=r+i, column=3, value=str(day_plus) if day_plus != 0 else ("0" if i in (3,4,8,9,11,12,13,14,15) else "0" if day_plus==0 else "-")).border = border
        ws.cell(row=r+i, column=3).alignment = Alignment(horizontal='center')

        # Final Day
        fday = _final_day_for_step(i, row, df_vtt)
        ws.cell(row=r+i, column=4, value=str(fday) if fday != 0 else "-").border = border
        ws.cell(row=r+i, column=4).alignment = Alignment(horizontal='center')

        # Paint date cells
        paint_len = day_plus if (isinstance(day_plus, int) and day_plus > 0) else 1
        if i in (0,1,5,6,7):
            paint_len = 1  # Day+ = 0 -> only final day for these steps
        start_idx = max(1, fday - paint_len + 1) if fday else 0
        for idx, d in enumerate(timeline_days, start=0):
            ci = start_col + idx
            cell = ws.cell(row=r+i, column=ci, value="")
            cell.border = border
            # weekend shading
            if d.weekday() in (5,6):
                cell.fill = weekendfill
            # paint range overrides shade
            if fday and start_idx <= (idx+1) <= fday:
                if i == 8:
                    cell.fill = darkbluefill
                else:
                    cell.fill = paintfill

    # Column widths
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    for k in range(start_col, start_col + len(timeline_days)):
        ws.column_dimensions[get_column_letter(k)].width = 4

    # VTT SUMMARY block under the table (mirror of UI summary with mini-Gantt)
    rr = r + len(time_labels) + 2
    ws.cell(row=rr, column=1, value='VTT SUMMARY').font = Font(bold=True, size=14)
    rr += 1

    kpi_rows = _build_kpi_rows(row, df_vtt)

    summary_start_col = 3

    for label_txt, val, start_day in kpi_rows:
        if label_txt == 'OVS SAP STAGES' and val is None and start_day is None:
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=summary_start_col + len(timeline_days) - 1)
            ovs_cell = ws.cell(row=rr, column=1, value=label_txt)
            ovs_cell.font = Font(bold=True, color='FFFFFF')
            ovs_cell.fill = PatternFill(fill_type='solid', fgColor='1F4E79')
            ovs_cell.border = border
            ovs_cell.alignment = Alignment(horizontal='left')
            for ci in range(1, summary_start_col + len(timeline_days)):
                ws.cell(row=rr, column=ci).border = border
                ws.cell(row=rr, column=ci).fill = PatternFill(fill_type='solid', fgColor='1F4E79')
            rr += 1
            continue

        ws.cell(row=rr, column=1, value=label_txt).font = bold
        ws.cell(row=rr, column=1).border = border
        ws.cell(row=rr, column=1).alignment = Alignment(horizontal='left')

        display_val = str(val) if val and val > 0 else "-"
        ws.cell(row=rr, column=2, value=display_val).border = border
        ws.cell(row=rr, column=2).alignment = Alignment(horizontal='center')

        # Pintar mini-Gantt en las columnas de días usando mismo eje temporal
        for idx, d in enumerate(timeline_days, start=0):
            ci = summary_start_col + idx
            cell = ws.cell(row=rr, column=ci, value="")
            cell.border = border
            if val and val > 0 and start_day:
                end_day = start_day + val - 1
                day_index = idx + 1
                if start_day <= day_index <= end_day:
                    # Azul claro para POL>POD, verde para el resto
                    cell.fill = darkbluefill if label_txt == "POL>POD" else paintfill
        rr += 1

    # Customer Safety STOCK debajo del resumen
    ws.cell(row=rr, column=1, value='Customer Safety STOCK').font = bold
    if row is not None and 'Safety stock' in df_vtt.columns:
        ws.cell(row=rr, column=2, value=str(row['Safety stock']))

    if include_snapshot_sheet:
        # Snapshot sheet with the same timeline rendered as an embedded image.
        snapshot_ws = wb.create_sheet('UI Snapshot')
        snapshot_ws.sheet_view.showGridLines = False
        try:
            snapshot_png_bytes = _build_snapshot_png_bytes(
                row=row,
                df_vtt=df_vtt,
                selected_pol=selected_pol,
                selected_pod=selected_pod,
                time_labels=time_labels,
                headers=headers,
                timeline_days=timeline_days,
            )
            image_buffer = BytesIO(snapshot_png_bytes)
            xl_image = XLImage(image_buffer)
            xl_image._source_buffer = image_buffer
            xl_image.anchor = 'A1'
            snapshot_ws.add_image(xl_image)
        except Exception:
            snapshot_ws['A1'] = 'UI Snapshot could not be generated.'

    # Return bytes
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()



# --- Single 'Generate files' button, then show download buttons in English ---
st.markdown("<hr style='margin:32px 0;'>", unsafe_allow_html=True)

if generate_files_clicked:
    snapshot_png_bytes = _build_snapshot_png_bytes(
        row=row,
        df_vtt=df_vtt,
        selected_pol=st.session_state.get('pol_select',''),
        selected_pod=st.session_state.get('pod_select',''),
        time_labels=time_labels,
        headers=headers,
        timeline_days=timeline_days,
        font_multiplier=1,
    )
    image_b64 = base64.b64encode(snapshot_png_bytes).decode('utf-8') if snapshot_png_bytes else ''

    excel_bytes = build_excel_workbook(
        row=row,
        df_vtt=df_vtt,
        selected_pol=st.session_state.get('pol_select',''),
        selected_pod=st.session_state.get('pod_select',''),
        time_labels=time_labels,
        headers=headers,
        timeline_days=timeline_days,
        include_snapshot_sheet=True,
    )
    excel_b64 = base64.b64encode(excel_bytes).decode('utf-8') if excel_bytes else ''

    # Obtener valores para el nombre del archivo
    pol_val = st.session_state.get('pol_select', '').replace(' ', '_')
    pod_val = st.session_state.get('pod_select', '').replace(' ', '_')
    # Obtener shipper de la fila seleccionada (columna 10)
    if row is not None and len(df_vtt.columns) > 10:
        shipper_col = df_vtt.columns[10]
        shipper_val = str(row.get(shipper_col, '')).replace(' ', '_')
    else:
        shipper_val = ''
    # Si alguno está vacío, poner UNKNOWN
    pol_val = pol_val if pol_val else 'UNKNOWN'
    pod_val = pod_val if pod_val else 'UNKNOWN'
    shipper_val = shipper_val if shipper_val else 'UNKNOWN'
    base_file_name = f"VTT_{pol_val}_{pod_val}_{shipper_val}"
    excel_file_name = f"{base_file_name}.xlsx"
    image_file_name = f"{base_file_name}.png"

    st.markdown(f"""
    <div class='vtt-action-bar'>
        <button id='excelBtn' class='vtt-action-btn'>Excel file</button>
        <button id='imgBtn' class='vtt-action-btn'>Image</button>
    </div>
    """, unsafe_allow_html=True)
    components.html(
        """ 
        <script>
        (function(){
            function parentDoc(){
                try { return window.parent && window.parent.document ? window.parent.document : document; } catch(e){ return document; }
            }
            function getBtn(){ return parentDoc().getElementById('imgBtn'); }
            function getExcelBtn(){ return parentDoc().getElementById('excelBtn'); }
            function base64ToBlob(base64, mimeType){
                var binary = atob(base64);
                var bytes = new Uint8Array(binary.length);
                for (var i = 0; i < binary.length; i++) {
                    bytes[i] = binary.charCodeAt(i);
                }
                return new Blob([bytes], { type: mimeType });
            }
            function downloadBlob(blob, fileName){
                var d = parentDoc();
                var url = URL.createObjectURL(blob);
                var a = d.createElement('a');
                a.href = url;
                a.download = fileName;
                d.body.appendChild(a);
                a.click();
                setTimeout(function(){ d.body.removeChild(a); URL.revokeObjectURL(url); }, 100);
            }
            function bind(){
                var imageButton = getBtn();
                var excelButton = getExcelBtn();
                if (!imageButton || !excelButton) { setTimeout(bind, 250); return; }
                imageButton.addEventListener('click', function(){
                    if (!'__IMAGE_B64__') { alert('No se pudo generar la imagen'); return; }
                    downloadBlob(base64ToBlob('__IMAGE_B64__', 'image/png'), '__IMAGE_FILE_NAME__');
                });
                excelButton.addEventListener('click', function(){
                    if (!'__EXCEL_B64__') { alert('No se pudo generar el Excel'); return; }
                    downloadBlob(
                        base64ToBlob('__EXCEL_B64__', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                        '__EXCEL_FILE_NAME__'
                    );
                });
            }
            bind();
        })();
        </script>
        """.replace('__IMAGE_FILE_NAME__', image_file_name).replace('__EXCEL_FILE_NAME__', excel_file_name).replace('__EXCEL_B64__', excel_b64).replace('__IMAGE_B64__', image_b64),
        height=10,
    )