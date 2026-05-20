import os
import re
from datetime import datetime, timedelta
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


TIME_LABELS = [
    "1. Day Customer Order",
    "2. Day ILN/FF Order",
    "3. First Receipt Days",
    "4. Pack. prep. & load",
    "5. Transport to POL",
    "6. First Day to POL",
    "7. Cut off",
    "8. ETD",
    "9. Transit Duration (ETD>ETA)",
    "10. Days of flexibility",
    "11. Customs clearence",
    "12. Transport to plant",
    "13. Rounding",
    "14. Due Date",
]

HEADERS = ["Steps", "Day", "Day+", "Final Day"]


def _coerce_to_int(val):
    try:
        if pd.isna(val):
            return 0
    except Exception:
        pass
    if isinstance(val, (int, float)):
        try:
            return int(round(float(val)))
        except Exception:
            return 0
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return 0
        match = re.search(r"[-+]?\d+(?:[\.,]\d+)?", s)
        if match:
            try:
                num = float(match.group(0).replace(',', '.'))
                return int(round(num))
            except Exception:
                return 0
    try:
        return int(val)
    except Exception:
        return 0


def _format_expiration_date(row, df_vtt):
    try:
        if row is None or 'Expiration Date' not in df_vtt.columns:
            return ''
        raw_exp = row.get('Expiration Date', '')
        if pd.isna(raw_exp):
            return ''
        if isinstance(raw_exp, (pd.Timestamp, datetime)):
            return raw_exp.strftime('%d/%m/%Y')
        try:
            return pd.to_datetime(raw_exp).strftime('%d/%m/%Y')
        except Exception:
            return str(raw_exp)
    except Exception:
        try:
            return '' if row is None else str(row.get('Expiration Date', ''))
        except Exception:
            return ''


def _due_date_day_plus_value(row, df_vtt):
    try:
        if row is None:
            return 7

        if '14 Rounding' in df_vtt.columns and '15 Due Date' in df_vtt.columns:
            rounding_val = _coerce_to_int(row.get('14 Rounding'))
            due_date_val = _coerce_to_int(row.get('15 Due Date'))
            if due_date_val >= rounding_val > 0:
                return due_date_val - rounding_val

        row_id = str(row.get('ID', '')).strip().upper() if 'ID' in df_vtt.columns else ''
        row_pol = str(row.get('POL', '')).strip().upper() if 'POL' in df_vtt.columns else ''
        row_pod = str(row.get('POD', '')).strip().upper() if 'POD' in df_vtt.columns else ''
        if row_id == 'O001' and row_pol == 'CNSHA' and row_pod == 'PTLEI':
            return 5
    except Exception:
        pass
    return 7


def _get_value_safe(val):
    if pd.isna(val):
        return "-"
    try:
        return str(val)
    except Exception:
        return "-"


def _final_day_for_step(step_index, row, df_vtt):
    try:
        if step_index == 0:
            return int(row['1 Day Customer Order']) if row is not None and '1 Day Customer Order' in df_vtt.columns else 0
        if step_index == 1:
            return int(row['2 Day ILN Order']) if row is not None and '2 Day ILN Order' in df_vtt.columns else 0
        if step_index == 2:
            return int(row['3.2 First Receipt Days']) if row is not None and '3.2 First Receipt Days' in df_vtt.columns else 0
        if step_index == 3:
            return int(row['4.3 Packaging préparation & loading']) if row is not None and '4.3 Packaging préparation & loading' in df_vtt.columns else 0
        if step_index == 4:
            return int(row['5.3 Transport ILN to POL']) if row is not None and '5.3 Transport ILN to POL' in df_vtt.columns else 0
        if step_index == 5:
            return int(row['6 First Day to POL']) if row is not None and '6 First Day to POL' in df_vtt.columns else 0
        if step_index == 6:
            return int(row['7 Cutt off']) if row is not None and '7 Cutt off' in df_vtt.columns else 0
        if step_index == 7:
            return int(row['8 ETD']) if row is not None and '8 ETD' in df_vtt.columns else 0
        if step_index == 8:
            if row is not None and '9 ETD> ETA' in df_vtt.columns:
                return int(row['9 ETD> ETA'])
            if row is not None and '9 ETD>ETA' in df_vtt.columns:
                return int(row['9 ETD>ETA'])
            return 0
        if step_index == 9:
            if row is not None and '10 Days flexibility 1' in df_vtt.columns and pd.notna(row['10 Days flexibility 1']):
                return int(row['10 Days flexibility 1'])
            base = None
            if row is not None and '9 ETD> ETA' in df_vtt.columns:
                base = row['9 ETD> ETA']
            elif row is not None and '9 ETD>ETA' in df_vtt.columns:
                base = row['9 ETD>ETA']
            base_number = pd.to_numeric(base, errors='coerce') if base is not None else float('nan')
            if pd.isna(base_number):
                matches = re.findall(r"[-+]?\.?\d+", str(base)) if base is not None else []
                base_number = float(matches[0]) if matches else float('nan')
            plus = _coerce_to_int(row['Time for security']) if row is not None and 'Time for security' in df_vtt.columns else 0
            return int(float(base_number)) + 1 + int(plus) if not pd.isna(base_number) else 0
        if step_index == 10:
            return int(row['11 Days flexibility 2']) if row is not None and '11 Days flexibility 2' in df_vtt.columns else 0
        if step_index == 11:
            if row is not None and '12 Customs Clearance' in df_vtt.columns:
                return int(row['12 Customs Clearance'])
            if row is not None and '12 Customs clearence' in df_vtt.columns:
                return int(row['12 Customs clearence'])
            return 0
        if step_index == 12:
            return int(row['13 Transport to Plant']) if row is not None and '13 Transport to Plant' in df_vtt.columns else 0
        if step_index == 13:
            return int(row['14 Rounding']) if row is not None and '14 Rounding' in df_vtt.columns else 0
        if step_index == 14:
            return int(row['15 Due Date']) if row is not None and '15 Due Date' in df_vtt.columns else 0
        if step_index == 15:
            return int(row['16 Manufacturing']) if row is not None and '16 Manufacturing' in df_vtt.columns else 0
        return 0
    except Exception:
        return 0


def _day_plus_value_for_step(step_index, row, df_vtt):
    if step_index in (0, 1, 6, 7):
        return 0
    if step_index == 2:
        return _coerce_to_int(row['3 .1 Time of Recept in ILN']) if row is not None and '3 .1 Time of Recept in ILN' in df_vtt.columns else 0
    if step_index == 3:
        return _coerce_to_int(row['4.2 Packaging préparation & loading']) if row is not None and '4.2 Packaging préparation & loading' in df_vtt.columns else 0
    if step_index == 4:
        return _coerce_to_int(row['5.2 Transport ILN to POL']) if row is not None and '5.2 Transport ILN to POL' in df_vtt.columns else 0
    if step_index == 5:
        return _coerce_to_int(row['First Day to POL']) if row is not None and 'First Day to POL' in df_vtt.columns else 0
    if step_index == 8:
        return _coerce_to_int(row['Transit time']) if row is not None and 'Transit time' in df_vtt.columns else 0
    if step_index == 9:
        return _coerce_to_int(row['Time for security']) if row is not None and 'Time for security' in df_vtt.columns else 0
    if step_index == 10:
        return _coerce_to_int(row['Time for security2 buffer']) if row is not None and 'Time for security2 buffer' in df_vtt.columns else 0
    if step_index == 11:
        return _coerce_to_int(row['Cust.']) if row is not None and 'Cust.' in df_vtt.columns else 0
    if step_index == 12:
        return _coerce_to_int(row['Trpt POD/PFI vers Usine']) if row is not None and 'Trpt POD/PFI vers Usine' in df_vtt.columns else 0
    if step_index == 13:
        if row is not None and 'Round.' in df_vtt.columns:
            return _coerce_to_int(row['Round.'])
        if row is not None and 'Round' in df_vtt.columns:
            return _coerce_to_int(row['Round'])
        return 0
    if step_index == 14:
        return _due_date_day_plus_value(row, df_vtt)
    if step_index == 15:
        return 7
    return 0


def _step_start_index(step_index, row, df_vtt):
    try:
        final_day = _final_day_for_step(step_index, row, df_vtt)
        if not final_day:
            return 0
        day_plus = _day_plus_value_for_step(step_index, row, df_vtt)
        paint_len = day_plus if isinstance(day_plus, int) and day_plus > 0 else 1
        if step_index in (0, 1, 5, 6, 7):
            paint_len = 1
        return max(1, final_day - paint_len + 1)
    except Exception:
        return 0


def _build_kpi_rows(row, df_vtt):
    total_tt = None
    if row is not None:
        transit_time = pd.to_numeric(row.get('Transit time', None), errors='coerce') if 'Transit time' in df_vtt.columns else None
        security_time = pd.to_numeric(row.get('Time for security', None), errors='coerce') if 'Time for security' in df_vtt.columns else None
        parts = [value for value in (transit_time, security_time) if value is not None and pd.notna(value)]
        if parts:
            total_tt = float(sum(parts))

    pod_det = None
    try:
        if row is not None:
            customs_val = None
            flex1_val = None
            if '12 Customs Clearance' in df_vtt.columns:
                customs_val = _coerce_to_int(row.get('12 Customs Clearance'))
            elif '12 Customs clearence' in df_vtt.columns:
                customs_val = _coerce_to_int(row.get('12 Customs clearence'))
            if '10 Days flexibility 1' in df_vtt.columns:
                flex1_val = _coerce_to_int(row.get('10 Days flexibility 1'))
            if customs_val and flex1_val:
                pod_det = customs_val - flex1_val
    except Exception:
        pod_det = None

    pod_plant = None
    try:
        if row is not None:
            customs_val = None
            rounding_val = None
            if '12 Customs Clearance' in df_vtt.columns:
                customs_val = _coerce_to_int(row.get('12 Customs Clearance'))
            elif '12 Customs clearence' in df_vtt.columns:
                customs_val = _coerce_to_int(row.get('12 Customs clearence'))
            if '14 Rounding' in df_vtt.columns:
                rounding_val = _coerce_to_int(row.get('14 Rounding'))
            if rounding_val and customs_val:
                pod_plant = rounding_val - customs_val
    except Exception:
        pod_plant = None

    try:
        transit_day = _coerce_to_int(row['8 ETD']) if row is not None and '8 ETD' in df_vtt.columns else 0
        pack_day = _coerce_to_int(row['4.1 Packaging préparation & loading']) if row is not None and '4.1 Packaging préparation & loading' in df_vtt.columns else 0
        kpi_sup_pol = transit_day - pack_day
    except Exception:
        pack_day = 0
        kpi_sup_pol = 0

    kpi_pol_pod = _coerce_to_int(total_tt) if total_tt is not None and not pd.isna(total_tt) else 0
    kpi_pod_det = _coerce_to_int(pod_det) if pod_det is not None else 0
    kpi_pod_plant = _coerce_to_int(pod_plant) if pod_plant is not None else 0

    start_sup = pack_day if pack_day > 0 else 0
    start_pol_pod = start_sup + kpi_sup_pol if start_sup and kpi_sup_pol > 0 else 0
    start_pod_det = start_pol_pod + kpi_pol_pod if start_pol_pod and kpi_pol_pod > 0 else 0
    start_pod_plant = start_pod_det + kpi_pod_det if start_pod_det and kpi_pod_det > 0 else 0

    try:
        final_day_14 = _final_day_for_step(13, row, df_vtt)
        final_day_1 = _final_day_for_step(0, row, df_vtt)
        customer_leadtime = final_day_14 - final_day_1 + 1
    except Exception:
        customer_leadtime = 0

    try:
        final_day_14 = _final_day_for_step(13, row, df_vtt)
        pack_day = _coerce_to_int(row['4.1 Packaging préparation & loading']) if row is not None and '4.1 Packaging préparation & loading' in df_vtt.columns else 0
        transportation_duration = final_day_14 - pack_day + 1
    except Exception:
        pack_day = 0
        transportation_duration = 0

    start_transportation = pack_day if pack_day > 0 else 0

    return [
        ("CUSTOMER LEADTIME (CLT)", customer_leadtime, _step_start_index(0, row, df_vtt)),
        ("OVS SAP STAGES", None, None),
        ("Transportation Duration", transportation_duration, start_transportation),
        ("SUPPLIER>POL", kpi_sup_pol, start_sup),
        ("POL>POD", kpi_pol_pod, start_pol_pod),
        ("POD DETENTION", kpi_pod_det, start_pod_det),
        ("POD>PLANT", kpi_pod_plant, start_pod_plant),
    ]


def _visible_timeline_step_data(display_index, row, df_vtt):
    if display_index < 9 or row is None:
        return None

    if display_index == 9:
        base_final = 0
        if '9 ETD> ETA' in df_vtt.columns:
            base_final = _coerce_to_int(row.get('9 ETD> ETA'))
        elif '9 ETD>ETA' in df_vtt.columns:
            base_final = _coerce_to_int(row.get('9 ETD>ETA'))

        flex_1 = _coerce_to_int(row.get('Time for security')) if 'Time for security' in df_vtt.columns else 0
        flex_2 = _coerce_to_int(row.get('Time for security2 buffer')) if 'Time for security2 buffer' in df_vtt.columns else 0
        day_plus = flex_1 + flex_2
        final_day = _coerce_to_int(row.get('11 Days flexibility 2')) if '11 Days flexibility 2' in df_vtt.columns else 0
        if not final_day:
            base_flex = _coerce_to_int(row.get('10 Days flexibility 1')) if '10 Days flexibility 1' in df_vtt.columns else 0
            if base_flex:
                final_day = base_flex + 1 + flex_2

        segments = []
        if final_day and day_plus > 0:
            start_idx = max(1, final_day - day_plus + 1)
            if flex_1 > 0:
                segments.append({'start': start_idx, 'end': min(final_day, start_idx + flex_1 - 1), 'fill': '#87ceeb'})
            green_start = start_idx + max(flex_1, 0)
            if green_start <= final_day:
                segments.append({'start': green_start, 'end': final_day, 'fill': '#90ee90'})

        return {
            'day': str(base_final + 1) if base_final else '-',
            'day_plus': day_plus,
            'final_day': final_day,
            'segments': segments,
        }

    if display_index == 10:
        day_plus = _coerce_to_int(row.get('Cust.')) if 'Cust.' in df_vtt.columns else 0
        final_day = 0
        if '12 Customs Clearance' in df_vtt.columns:
            final_day = _coerce_to_int(row.get('12 Customs Clearance'))
        elif '12 Customs clearence' in df_vtt.columns:
            final_day = _coerce_to_int(row.get('12 Customs clearence'))
        previous_final = _coerce_to_int(row.get('11 Days flexibility 2')) if '11 Days flexibility 2' in df_vtt.columns else 0
        day_value = final_day if day_plus <= 0 and final_day else (previous_final + 1 if previous_final else '-')
        segments = []
        if final_day and day_plus > 0:
            segments.append({'start': max(1, final_day - day_plus + 1), 'end': final_day, 'fill': '#90ee90'})
        return {'day': str(day_value) if day_value != '-' else '-', 'day_plus': day_plus, 'final_day': final_day, 'segments': segments}

    if display_index == 11:
        day_plus = _coerce_to_int(row.get('Trpt POD/PFI vers Usine')) if 'Trpt POD/PFI vers Usine' in df_vtt.columns else 0
        final_day = _coerce_to_int(row.get('13 Transport to Plant')) if '13 Transport to Plant' in df_vtt.columns else 0
        previous_final = 0
        if '12 Customs Clearance' in df_vtt.columns:
            previous_final = _coerce_to_int(row.get('12 Customs Clearance'))
        elif '12 Customs clearence' in df_vtt.columns:
            previous_final = _coerce_to_int(row.get('12 Customs clearence'))
        day_value = previous_final + 1 if previous_final else '-'
        segments = []
        if final_day and day_plus > 0:
            segments.append({'start': max(1, final_day - day_plus + 1), 'end': final_day, 'fill': '#90ee90'})
        return {'day': str(day_value) if day_value != '-' else '-', 'day_plus': day_plus, 'final_day': final_day, 'segments': segments}

    if display_index == 12:
        if 'Round.' in df_vtt.columns:
            day_plus = _coerce_to_int(row.get('Round.'))
        elif 'Round' in df_vtt.columns:
            day_plus = _coerce_to_int(row.get('Round'))
        else:
            day_plus = 0
        final_day = _coerce_to_int(row.get('14 Rounding')) if '14 Rounding' in df_vtt.columns else 0
        previous_final = _coerce_to_int(row.get('13 Transport to Plant')) if '13 Transport to Plant' in df_vtt.columns else 0
        day_value = final_day if day_plus <= 0 and final_day else (previous_final + 1 if previous_final else '-')
        segments = []
        if final_day and day_plus > 0:
            segments.append({'start': max(1, final_day - day_plus + 1), 'end': final_day, 'fill': '#90ee90'})
        return {'day': str(day_value) if day_value != '-' else '-', 'day_plus': day_plus, 'final_day': final_day, 'segments': segments}

    if display_index == 13:
        day_plus = _due_date_day_plus_value(row, df_vtt)
        final_day = _coerce_to_int(row.get('15 Due Date')) if '15 Due Date' in df_vtt.columns else 0
        previous_final = _coerce_to_int(row.get('14 Rounding')) if '14 Rounding' in df_vtt.columns else 0
        day_value = previous_final + 1 if previous_final else '-'
        segments = []
        if final_day and day_plus > 0:
            segments.append({'start': max(1, final_day - day_plus + 1), 'end': final_day, 'fill': '#90ee90'})
        return {'day': str(day_value) if day_value != '-' else '-', 'day_plus': day_plus, 'final_day': final_day, 'segments': segments}

    return None


def _day_value_for_step(step_index, row, df_vtt):
    try:
        if step_index == 0:
            return _get_value_safe(row['1 Day Customer Order']) if row is not None and '1 Day Customer Order' in df_vtt.columns else '-'
        if step_index == 1:
            return _get_value_safe(row['2 Day ILN Order']) if row is not None and '2 Day ILN Order' in df_vtt.columns else '-'
        if step_index == 2:
            return _get_value_safe(row['3 First Receipt Days']) if row is not None and '3 First Receipt Days' in df_vtt.columns else '-'
        if step_index == 3:
            return _get_value_safe(row['4.1 Packaging préparation & loading']) if row is not None and '4.1 Packaging préparation & loading' in df_vtt.columns else '-'
        if step_index == 4:
            return _get_value_safe(row['5.1 Transport ILN to POL']) if row is not None and '5.1 Transport ILN to POL' in df_vtt.columns else '-'
        if step_index == 5:
            return _get_value_safe(row['6 First Day to POL']) if row is not None and '6 First Day to POL' in df_vtt.columns else '-'
        if step_index == 6:
            return _get_value_safe(row['7 Cutt off']) if row is not None and '7 Cutt off' in df_vtt.columns else '-'
        if step_index == 7:
            return _get_value_safe(row['8 ETD']) if row is not None and '8 ETD' in df_vtt.columns else '-'
        if step_index == 8:
            return _get_value_safe(row['8 ETD']) if row is not None and '8 ETD' in df_vtt.columns else '-'
        if step_index == 9:
            base = _final_day_for_step(8, row, df_vtt)
            return str(base + 1) if base else '-'
        if step_index == 10:
            base = _final_day_for_step(9, row, df_vtt)
            return str(base + 1) if base else '-'
        if step_index == 11:
            base = _final_day_for_step(10, row, df_vtt)
            return str(base + 1) if base else '-'
        if step_index == 12:
            base = _final_day_for_step(11, row, df_vtt)
            return str(base + 1) if base else '-'
        if step_index == 13:
            base = _final_day_for_step(12, row, df_vtt)
            return str(base + 1) if base else '-'
        if step_index == 14:
            base = _final_day_for_step(13, row, df_vtt)
            return str(base + 1) if base else '-'
        if step_index == 15:
            base = _final_day_for_step(14, row, df_vtt)
            return str(base + 1) if base else '-'
        return '-'
    except Exception:
        return '-'


def _ui_timeline_day_value(display_index, row, df_vtt):
    visible_step_data = _visible_timeline_step_data(display_index, row, df_vtt)
    if visible_step_data is not None:
        return visible_step_data['day']
    if display_index < 9:
        return _day_value_for_step(display_index, row, df_vtt)
    return '-'


def _ui_timeline_day_plus(display_index, row, df_vtt):
    visible_step_data = _visible_timeline_step_data(display_index, row, df_vtt)
    if visible_step_data is not None:
        return visible_step_data['day_plus']
    if display_index < 9:
        return _day_plus_value_for_step(display_index, row, df_vtt)
    return 0


def _ui_timeline_final_day(display_index, row, df_vtt):
    visible_step_data = _visible_timeline_step_data(display_index, row, df_vtt)
    if visible_step_data is not None:
        return visible_step_data['final_day']
    if display_index < 9:
        return _final_day_for_step(display_index, row, df_vtt)
    return 0


def _ui_timeline_paint_segments(display_index, row, df_vtt):
    visible_step_data = _visible_timeline_step_data(display_index, row, df_vtt)
    if visible_step_data is not None:
        return visible_step_data['segments']

    final_day = _ui_timeline_final_day(display_index, row, df_vtt)
    if not final_day:
        return []

    if display_index in (10, 12):
        day_plus = _ui_timeline_day_plus(display_index, row, df_vtt)
        if day_plus <= 0:
            return []

    day_plus = _ui_timeline_day_plus(display_index, row, df_vtt)
    paint_len = day_plus if isinstance(day_plus, int) and day_plus > 0 else 1
    if display_index in (0, 1, 5, 6, 7):
        paint_len = 1

    start_idx = max(1, final_day - paint_len + 1)
    segment = {
        'start': start_idx,
        'end': final_day,
        'fill': '#4a90e2' if display_index == 8 else '#90ee90',
    }
    return [segment]


def _hex_to_fill(hex_color):
    if not hex_color:
        return None
    value = hex_color.lstrip('#')
    if len(value) == 6:
        value = 'FF' + value.upper()
    return PatternFill(fill_type='solid', start_color=value, end_color=value)


def _compute_week_spans(days):
    spans = []
    current_week = None
    count = 0
    for day in days:
        week = day.isocalendar()[1]
        if current_week is None:
            current_week = week
            count = 1
        elif week == current_week:
            count += 1
        else:
            spans.append((current_week, count))
            current_week = week
            count = 1
    if current_week is not None:
        spans.append((current_week, count))
    return spans


def _segment_duration_for_step(display_index, row, df_vtt):
    total_duration = 0
    for segment in _ui_timeline_paint_segments(display_index, row, df_vtt):
        try:
            total_duration += max(0, int(segment['end']) - int(segment['start']) + 1)
        except Exception:
            continue
    return total_duration


def _segment_bounds_by_fill(display_index, row, df_vtt, fill):
    matching_segments = [segment for segment in _ui_timeline_paint_segments(display_index, row, df_vtt) if segment.get('fill') == fill]
    if not matching_segments:
        return '', ''

    try:
        start = min(int(segment['start']) for segment in matching_segments)
        end = max(int(segment['end']) for segment in matching_segments)
        return start, end
    except Exception:
        return '', ''


def _write_dashboard_sheet(ws, row, df_vtt, selected_pol, selected_pod, time_labels, headers, timeline_days):
    ws.sheet_view.showGridLines = False

    bold = Font(bold=True)
    section_title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    hfill = _hex_to_fill('#f5f5f5')
    weekfill = _hex_to_fill('#fffbe6')
    weekendfill = _hex_to_fill('#ffd6d6')
    weekdayfill = _hex_to_fill('#e3eafc')
    paintfill = _hex_to_fill('#90ee90')
    darkbluefill = _hex_to_fill('#4a90e2')
    lightbluefill = _hex_to_fill('#87ceeb')

    row_cursor = 1
    ws.cell(row=row_cursor, column=1, value='POL:').font = bold
    ws.cell(row=row_cursor, column=2, value=selected_pol)
    ws.cell(row=row_cursor, column=3, value='POD:').font = bold
    ws.cell(row=row_cursor, column=4, value=selected_pod)
    row_cursor += 1

    commodity_col = 'Commodity' if 'Commodity' in df_vtt.columns else ('Comodity' if 'Comodity' in df_vtt.columns else None)
    shipper_col = df_vtt.columns[10] if len(df_vtt.columns) > 10 else None
    iln_col = df_vtt.columns[8] if len(df_vtt.columns) > 8 else None
    info_pairs = [
        ('ID', row.get('ID', '') if row is not None and 'ID' in df_vtt.columns else ''),
        ('Carrier', row.get('Carrier', '') if row is not None and 'Carrier' in df_vtt.columns else ''),
        ('Shipper', row.get(shipper_col, '') if row is not None and shipper_col and shipper_col in df_vtt.columns else ''),
        ('ILN/FF', row.get(iln_col, '') if row is not None and iln_col and iln_col in df_vtt.columns else ''),
        ('PLANT', row.get('Name Destin Site', '') if row is not None and 'Name Destin Site' in df_vtt.columns else ''),
        ('Commodity', row.get(commodity_col, '') if row is not None and commodity_col and commodity_col in df_vtt.columns else ''),
        ('E/D', _format_expiration_date(row, df_vtt)),
    ]
    col_cursor = 1
    for label, value in info_pairs:
        ws.cell(row=row_cursor, column=col_cursor, value=f'{label}:').font = bold
        ws.cell(row=row_cursor, column=col_cursor + 1, value='' if pd.isna(value) else str(value))
        col_cursor += 2
    row_cursor += 2

    start_col = 5
    for week, span in _compute_week_spans(timeline_days):
        ws.merge_cells(start_row=row_cursor, start_column=start_col, end_row=row_cursor, end_column=start_col + span - 1)
        cell = ws.cell(row=row_cursor, column=start_col, value=f'W{week}')
        cell.fill = weekfill
        cell.font = bold
        cell.alignment = Alignment(horizontal='center')
        for current_col in range(start_col, start_col + span):
            ws.cell(row=row_cursor, column=current_col).border = border
        start_col += span
    row_cursor += 1

    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_cursor, column=column_index, value=header)
        cell.fill = hfill
        cell.font = bold
        cell.border = border
        cell.alignment = Alignment(horizontal='center' if column_index > 1 else 'left')
    for offset, day in enumerate(timeline_days):
        column_index = 5 + offset
        cell = ws.cell(row=row_cursor, column=column_index, value=day.strftime('%d-%b'))
        cell.fill = weekendfill if day.weekday() in (5, 6) else weekdayfill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='bottom', textRotation=90)
    row_cursor += 1

    for step_index, label in enumerate(time_labels):
        current_row = row_cursor + step_index
        ws.row_dimensions[current_row].height = 10.5

        step_cell = ws.cell(row=current_row, column=1, value=label)
        step_cell.fill = hfill
        step_cell.font = bold
        step_cell.border = border
        step_cell.alignment = Alignment(horizontal='left')

        day_cell = ws.cell(row=current_row, column=2, value=_ui_timeline_day_value(step_index, row, df_vtt))
        day_cell.border = border
        day_cell.alignment = Alignment(horizontal='center')

        day_plus = _ui_timeline_day_plus(step_index, row, df_vtt)
        day_plus_cell = ws.cell(row=current_row, column=3, value=str(day_plus) if day_plus != 0 else '0')
        day_plus_cell.border = border
        day_plus_cell.alignment = Alignment(horizontal='center')

        final_day = _ui_timeline_final_day(step_index, row, df_vtt)
        final_day_cell = ws.cell(row=current_row, column=4, value=str(final_day) if final_day else '-')
        final_day_cell.border = border
        final_day_cell.alignment = Alignment(horizontal='center')

        paint_segments = _ui_timeline_paint_segments(step_index, row, df_vtt)
        for day_offset, day in enumerate(timeline_days):
            column_index = 5 + day_offset
            cell = ws.cell(row=current_row, column=column_index, value='')
            cell.border = border
            if day.weekday() in (5, 6):
                cell.fill = weekendfill
            for segment in paint_segments:
                if segment['start'] <= (day_offset + 1) <= segment['end']:
                    fill = segment['fill']
                    if fill == '#4a90e2':
                        cell.fill = darkbluefill
                    elif fill == '#87ceeb':
                        cell.fill = lightbluefill
                    else:
                        cell.fill = paintfill
                    break

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    for column_index in range(5, 5 + len(timeline_days)):
        ws.column_dimensions[get_column_letter(column_index)].width = 4

    summary_row = row_cursor + len(time_labels) + 2
    ws.cell(row=summary_row, column=1, value='VTT SUMMARY').font = section_title_font
    summary_row += 1

    summary_start_col = 3
    current_col = summary_start_col
    for week, span in _compute_week_spans(timeline_days):
        ws.merge_cells(start_row=summary_row, start_column=current_col, end_row=summary_row, end_column=current_col + span - 1)
        cell = ws.cell(row=summary_row, column=current_col, value=f'W{week}')
        cell.fill = weekfill
        cell.font = bold
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        for merged_col in range(current_col, current_col + span):
            ws.cell(row=summary_row, column=merged_col).border = border
        current_col += span
    summary_row += 1

    for label, value, start_day in _build_kpi_rows(row, df_vtt):
        if label == 'OVS SAP STAGES' and value is None and start_day is None:
            ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=summary_start_col + len(timeline_days) - 1)
            ovs_cell = ws.cell(row=summary_row, column=1, value=label)
            ovs_cell.font = Font(bold=True, color='102845')
            ovs_cell.fill = PatternFill(fill_type='solid', fgColor='EAF2FB')
            ovs_cell.border = border
            ovs_cell.alignment = Alignment(horizontal='left')
            for merged_col in range(1, summary_start_col + len(timeline_days)):
                ws.cell(row=summary_row, column=merged_col).border = border
                ws.cell(row=summary_row, column=merged_col).fill = PatternFill(fill_type='solid', fgColor='EAF2FB')
            summary_row += 1
            continue

        label_cell = ws.cell(row=summary_row, column=1, value=label)
        label_cell.font = bold
        label_cell.border = border
        label_cell.alignment = Alignment(horizontal='left')

        value_cell = ws.cell(row=summary_row, column=2, value=str(value) if value and value > 0 else '-')
        value_cell.border = border
        value_cell.alignment = Alignment(horizontal='center')

        for day_offset, _day in enumerate(timeline_days):
            column_index = summary_start_col + day_offset
            cell = ws.cell(row=summary_row, column=column_index, value='')
            cell.border = border
            if value and value > 0 and start_day:
                end_day = start_day + value - 1
                if start_day <= (day_offset + 1) <= end_day:
                    cell.fill = darkbluefill if label == 'POL>POD' else paintfill
        summary_row += 1

    safety_label = ws.cell(row=summary_row, column=1, value='Customer Safety STOCK')
    safety_label.font = bold
    if row is not None and 'Safety stock' in df_vtt.columns:
        ws.cell(row=summary_row, column=2, value=str(row['Safety stock']))


def _sanitize_sheet_title(raw_title):
    clean = re.sub(r'[\\/*?:\[\]]', '_', str(raw_title).strip() or 'Dashboard')
    return clean[:31] if clean else 'Dashboard'


def _sheet_title_from_row(row, df_vtt):
    id_value = str(row.get('ID', '')).strip() if row is not None and 'ID' in df_vtt.columns else ''
    pol_value = str(row.get('POL', '')).strip() if row is not None and 'POL' in df_vtt.columns else ''
    pod_value = str(row.get('POD', '')).strip() if row is not None and 'POD' in df_vtt.columns else ''

    parts = [part for part in (id_value, pol_value, pod_value) if part]
    return _sanitize_sheet_title('_'.join(parts) or 'Dashboard')


def _unique_sheet_title(base_title, used_titles):
    if base_title not in used_titles:
        used_titles.add(base_title)
        return base_title

    suffix = 2
    while True:
        candidate = f'{base_title[:28]}_{suffix}'
        if candidate not in used_titles:
            used_titles.add(candidate)
            return candidate
        suffix += 1


def build_all_vtt_workbook(df_vtt, timeline_days):
    workbook = Workbook()
    first_sheet = True
    used_titles = set()

    for _, row in df_vtt.iterrows():
        sheet_base = _sheet_title_from_row(row, df_vtt)
        sheet_title = _unique_sheet_title(sheet_base, used_titles)
        worksheet = workbook.active if first_sheet else workbook.create_sheet()
        worksheet.title = sheet_title
        first_sheet = False
        _write_dashboard_sheet(
            worksheet,
            row=row,
            df_vtt=df_vtt,
            selected_pol=str(row.get('POL', '')) if 'POL' in df_vtt.columns else '',
            selected_pod=str(row.get('POD', '')) if 'POD' in df_vtt.columns else '',
            time_labels=TIME_LABELS,
            headers=HEADERS,
            timeline_days=timeline_days,
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_dynamic_single_sheet_workbook(df_vtt, timeline_days):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'ALL_VTT_DYNAMIC'
    worksheet.sheet_view.showGridLines = False
    raw_worksheet = workbook.create_sheet('Raw Data')
    raw_worksheet.sheet_view.showGridLines = False

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except Exception:
        pass

    bold = Font(bold=True)
    section_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    header_fill = _hex_to_fill('#f5f5f5')
    info_fill = _hex_to_fill('#eaf2fb')
    week_fill = _hex_to_fill('#fffbe6')
    weekend_fill = _hex_to_fill('#ffd6d6')
    weekday_fill = _hex_to_fill('#e3eafc')
    green_fill = _hex_to_fill('#90ee90')
    blue_fill = _hex_to_fill('#4a90e2')
    lightblue_fill = _hex_to_fill('#87ceeb')

    commodity_col = 'Commodity' if 'Commodity' in df_vtt.columns else ('Comodity' if 'Comodity' in df_vtt.columns else None)
    shipper_col = df_vtt.columns[10] if len(df_vtt.columns) > 10 else None
    iln_col = df_vtt.columns[8] if len(df_vtt.columns) > 8 else None

    metadata_rows = []
    step_rows = []
    kpi_rows = []

    for _, row in df_vtt.iterrows():
        id_value = str(row.get('ID', '')).strip() if 'ID' in df_vtt.columns else ''
        pol_value = str(row.get('POL', '')).strip() if 'POL' in df_vtt.columns else ''
        pod_value = str(row.get('POD', '')).strip() if 'POD' in df_vtt.columns else ''
        carrier_value = '' if 'Carrier' not in df_vtt.columns or pd.isna(row.get('Carrier', '')) else str(row.get('Carrier', ''))
        shipper_value = '' if not shipper_col or shipper_col not in df_vtt.columns or pd.isna(row.get(shipper_col, '')) else str(row.get(shipper_col, ''))
        iln_value = '' if not iln_col or iln_col not in df_vtt.columns or pd.isna(row.get(iln_col, '')) else str(row.get(iln_col, ''))
        plant_value = '' if 'Name Destin Site' not in df_vtt.columns or pd.isna(row.get('Name Destin Site', '')) else str(row.get('Name Destin Site', ''))
        commodity_value = '' if not commodity_col or commodity_col not in df_vtt.columns or pd.isna(row.get(commodity_col, '')) else str(row.get(commodity_col, ''))
        expiration_value = _format_expiration_date(row, df_vtt)
        safety_stock_value = '' if 'Safety stock' not in df_vtt.columns or pd.isna(row.get('Safety stock', '')) else row.get('Safety stock', '')

        metadata_rows.append([
            id_value,
            pol_value,
            pod_value,
            carrier_value,
            shipper_value,
            iln_value,
            plant_value,
            commodity_value,
            expiration_value,
            safety_stock_value,
        ])

        for step_index, label in enumerate(TIME_LABELS):
            green_start, green_end = _segment_bounds_by_fill(step_index, row, df_vtt, '#90ee90')
            blue_start, blue_end = _segment_bounds_by_fill(step_index, row, df_vtt, '#4a90e2')
            lightblue_start, lightblue_end = _segment_bounds_by_fill(step_index, row, df_vtt, '#87ceeb')
            step_rows.append([
                id_value,
                pol_value,
                pod_value,
                label,
                _ui_timeline_day_value(step_index, row, df_vtt),
                _ui_timeline_day_plus(step_index, row, df_vtt),
                _ui_timeline_final_day(step_index, row, df_vtt),
                green_start,
                green_end,
                blue_start,
                blue_end,
                lightblue_start,
                lightblue_end,
                _segment_duration_for_step(step_index, row, df_vtt),
                f'{id_value}|{label}',
            ])

        for label, value, start_day in _build_kpi_rows(row, df_vtt):
            if value is None:
                continue
            end_day = start_day + value - 1 if value and start_day else ''
            kpi_rows.append([id_value, label, value, start_day, end_day, f'{id_value}|{label}'])

    def lookup_formula(value_range, key_range, key_expression, fallback):
        return f'=IFERROR(INDEX({value_range}, MATCH({key_expression}, {key_range}, 0)), {fallback})'

    timeline_week_row = 4
    timeline_header_row = 5
    timeline_first_row = 6
    timeline_last_row = timeline_first_row + len(TIME_LABELS) - 1
    timeline_start_col = 5
    timeline_end_col = timeline_start_col + len(timeline_days) - 1

    summary_title_row = timeline_last_row + 2
    summary_week_row = summary_title_row + 1
    summary_first_row = summary_week_row + 1
    summary_labels = [
        'CUSTOMER LEADTIME (CLT)',
        'OVS SAP STAGES',
        'Transportation Duration',
        'SUPPLIER>POL',
        'POL>POD',
        'POD DETENTION',
        'POD>PLANT',
    ]
    summary_last_row = summary_first_row + len(summary_labels) - 1
    safety_stock_row = summary_last_row + 1
    raw_title_row = 1
    raw_header_row = 2
    raw_data_start_row = raw_header_row + 1
    raw_data_end_row = raw_header_row + len(step_rows)

    timeline_helper_start_col = timeline_end_col + 2
    summary_helper_start_col = timeline_helper_start_col + 6
    metadata_start_col = summary_helper_start_col + 2
    kpi_start_col = metadata_start_col + 10

    metadata_col_map = {
        'ID': metadata_start_col,
        'POL': metadata_start_col + 1,
        'POD': metadata_start_col + 2,
        'Carrier': metadata_start_col + 3,
        'Shipper': metadata_start_col + 4,
        'ILN/FF': metadata_start_col + 5,
        'PLANT': metadata_start_col + 6,
        'Commodity': metadata_start_col + 7,
        'E/D': metadata_start_col + 8,
        'Safety stock': metadata_start_col + 9,
    }
    metadata_start_row = 2
    metadata_end_row = metadata_start_row + len(metadata_rows) - 1

    kpi_col_map = {
        'ID': kpi_start_col,
        'KPI': kpi_start_col + 1,
        'Value': kpi_start_col + 2,
        'Start': kpi_start_col + 3,
        'End': kpi_start_col + 4,
        'Key': kpi_start_col + 5,
    }
    kpi_start_row = 2
    kpi_end_row = kpi_start_row + len(kpi_rows) - 1

    selected_id_cell = worksheet['B1']
    worksheet['A1'] = 'Selected ID:'
    worksheet['A1'].font = bold
    worksheet['D1'] = 'POL:'
    worksheet['D1'].font = bold
    worksheet['F1'] = 'POD:'
    worksheet['F1'].font = bold

    metadata_headers = ['ID', 'POL', 'POD', 'Carrier', 'Shipper', 'ILN/FF', 'PLANT', 'Commodity', 'E/D', 'Safety stock']
    for offset, header in enumerate(metadata_headers):
        cell = worksheet.cell(row=1, column=metadata_start_col + offset, value=header)
        cell.font = bold
        cell.fill = header_fill
    for row_index, values in enumerate(metadata_rows, start=metadata_start_row):
        for offset, value in enumerate(values):
            worksheet.cell(row=row_index, column=metadata_start_col + offset, value=value)

    for offset, header in enumerate(['ID', 'KPI', 'Value', 'Start Day', 'End Day', 'Key']):
        cell = worksheet.cell(row=1, column=kpi_start_col + offset, value=header)
        cell.font = bold
        cell.fill = header_fill
    for row_index, values in enumerate(kpi_rows, start=kpi_start_row):
        for offset, value in enumerate(values):
            worksheet.cell(row=row_index, column=kpi_start_col + offset, value=value)

    if metadata_rows:
        id_col_letter = get_column_letter(metadata_col_map['ID'])
        validation = DataValidation(type='list', formula1=f'=${id_col_letter}$2:${id_col_letter}${metadata_end_row}', allow_blank=False)
        worksheet.add_data_validation(validation)
        validation.add(selected_id_cell)
        selected_id_cell.value = metadata_rows[0][0]

    worksheet['E1'] = lookup_formula(
        f'${get_column_letter(metadata_col_map["POL"])}$2:${get_column_letter(metadata_col_map["POL"])}${metadata_end_row}',
        f'${get_column_letter(metadata_col_map["ID"])}$2:${get_column_letter(metadata_col_map["ID"])}${metadata_end_row}',
        '$B$1',
        '""',
    )
    worksheet['G1'] = lookup_formula(
        f'${get_column_letter(metadata_col_map["POD"])}$2:${get_column_letter(metadata_col_map["POD"])}${metadata_end_row}',
        f'${get_column_letter(metadata_col_map["ID"])}$2:${get_column_letter(metadata_col_map["ID"])}${metadata_end_row}',
        '$B$1',
        '""',
    )

    info_layout = [
        ('ID', 1, 2),
        ('Carrier', 3, 4),
        ('Shipper', 5, 6),
        ('ILN/FF', 7, 8),
        ('PLANT', 9, 10),
        ('Commodity', 11, 12),
        ('E/D', 13, 14),
    ]
    for label, label_col, value_col in info_layout:
        label_cell = worksheet.cell(row=2, column=label_col, value=f'{label}:')
        label_cell.font = bold
        formula_cell = worksheet.cell(row=2, column=value_col)
        formula_cell.value = lookup_formula(
            f'${get_column_letter(metadata_col_map[label])}$2:${get_column_letter(metadata_col_map[label])}${metadata_end_row}',
            f'${get_column_letter(metadata_col_map["ID"])}$2:${get_column_letter(metadata_col_map["ID"])}${metadata_end_row}',
            '$B$1',
            '""',
        )

    spans = _compute_week_spans(timeline_days)
    current_col = timeline_start_col
    for week, span in spans:
        worksheet.merge_cells(start_row=timeline_week_row, start_column=current_col, end_row=timeline_week_row, end_column=current_col + span - 1)
        cell = worksheet.cell(row=timeline_week_row, column=current_col, value=f'W{week}')
        cell.fill = week_fill
        cell.font = bold
        cell.alignment = Alignment(horizontal='center')
        for merged_col in range(current_col, current_col + span):
            worksheet.cell(row=timeline_week_row, column=merged_col).border = border
        current_col += span

    for column_index, header in enumerate(HEADERS, start=1):
        cell = worksheet.cell(row=timeline_header_row, column=column_index, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center' if column_index > 1 else 'left')
    for day_offset, day in enumerate(timeline_days):
        column_index = timeline_start_col + day_offset
        cell = worksheet.cell(row=timeline_header_row, column=column_index, value=day.strftime('%d-%b'))
        cell.fill = weekend_fill if day.weekday() in (5, 6) else weekday_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='bottom', textRotation=90)

    raw_key_col_letter = 'O'
    raw_day_col_letter = 'E'
    raw_day_plus_col_letter = 'F'
    raw_final_col_letter = 'G'
    raw_green_start_col_letter = 'H'
    raw_green_end_col_letter = 'I'
    raw_blue_start_col_letter = 'J'
    raw_blue_end_col_letter = 'K'
    raw_light_start_col_letter = 'L'
    raw_light_end_col_letter = 'M'

    timeline_helper_letters = [get_column_letter(timeline_helper_start_col + offset) for offset in range(6)]
    summary_helper_letters = [get_column_letter(summary_helper_start_col + offset) for offset in range(2)]

    raw_sheet_name = "'Raw Data'"
    timeline_key_range = f'{raw_sheet_name}!${raw_key_col_letter}${raw_data_start_row}:${raw_key_col_letter}${raw_data_end_row}'
    for step_offset, label in enumerate(TIME_LABELS):
        current_row = timeline_first_row + step_offset
        worksheet.row_dimensions[current_row].height = 10.5

        step_cell = worksheet.cell(row=current_row, column=1, value=label)
        step_cell.font = bold
        step_cell.fill = header_fill
        step_cell.border = border
        step_cell.alignment = Alignment(horizontal='left')

        worksheet.cell(row=current_row, column=2, value=lookup_formula(f'{raw_sheet_name}!${raw_day_col_letter}${raw_data_start_row}:${raw_day_col_letter}${raw_data_end_row}', timeline_key_range, f'$B$1&"|"&$A{current_row}', '"-"')).border = border
        worksheet.cell(row=current_row, column=3, value=lookup_formula(f'{raw_sheet_name}!${raw_day_plus_col_letter}${raw_data_start_row}:${raw_day_plus_col_letter}${raw_data_end_row}', timeline_key_range, f'$B$1&"|"&$A{current_row}', '0')).border = border
        worksheet.cell(row=current_row, column=4, value=lookup_formula(f'{raw_sheet_name}!${raw_final_col_letter}${raw_data_start_row}:${raw_final_col_letter}${raw_data_end_row}', timeline_key_range, f'$B$1&"|"&$A{current_row}', '"-"')).border = border

        helper_ranges = [
            raw_green_start_col_letter,
            raw_green_end_col_letter,
            raw_blue_start_col_letter,
            raw_blue_end_col_letter,
            raw_light_start_col_letter,
            raw_light_end_col_letter,
        ]
        for helper_letter, raw_letter in zip(timeline_helper_letters, helper_ranges):
            helper_cell = worksheet[f'{helper_letter}{current_row}']
            helper_cell.value = lookup_formula(
                f'{raw_sheet_name}!${raw_letter}${raw_data_start_row}:${raw_letter}${raw_data_end_row}',
                timeline_key_range,
                f'$B$1&"|"&$A{current_row}',
                '""',
            )

        for day_offset, day in enumerate(timeline_days):
            column_index = timeline_start_col + day_offset
            cell = worksheet.cell(row=current_row, column=column_index, value='')
            cell.border = border
            if day.weekday() in (5, 6):
                cell.fill = weekend_fill

    timeline_range = f'{get_column_letter(timeline_start_col)}{timeline_first_row}:{get_column_letter(timeline_end_col)}{timeline_last_row}'
    worksheet.conditional_formatting.add(
        timeline_range,
        FormulaRule(formula=[f'AND(${timeline_helper_letters[4]}{timeline_first_row}<>"",COLUMN()-{timeline_start_col - 1}>=${timeline_helper_letters[4]}{timeline_first_row},COLUMN()-{timeline_start_col - 1}<=${timeline_helper_letters[5]}{timeline_first_row})'], fill=lightblue_fill),
    )
    worksheet.conditional_formatting.add(
        timeline_range,
        FormulaRule(formula=[f'AND(${timeline_helper_letters[2]}{timeline_first_row}<>"",COLUMN()-{timeline_start_col - 1}>=${timeline_helper_letters[2]}{timeline_first_row},COLUMN()-{timeline_start_col - 1}<=${timeline_helper_letters[3]}{timeline_first_row})'], fill=blue_fill),
    )
    worksheet.conditional_formatting.add(
        timeline_range,
        FormulaRule(formula=[f'AND(${timeline_helper_letters[0]}{timeline_first_row}<>"",COLUMN()-{timeline_start_col - 1}>=${timeline_helper_letters[0]}{timeline_first_row},COLUMN()-{timeline_start_col - 1}<=${timeline_helper_letters[1]}{timeline_first_row})'], fill=green_fill),
    )

    worksheet.cell(row=summary_title_row, column=1, value='VTT SUMMARY').font = section_font
    current_col = 3
    for week, span in spans:
        worksheet.merge_cells(start_row=summary_week_row, start_column=current_col, end_row=summary_week_row, end_column=current_col + span - 1)
        cell = worksheet.cell(row=summary_week_row, column=current_col, value=f'W{week}')
        cell.fill = week_fill
        cell.font = bold
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        for merged_col in range(current_col, current_col + span):
            worksheet.cell(row=summary_week_row, column=merged_col).border = border
        current_col += span

    kpi_id_col_letter = get_column_letter(kpi_col_map['ID'])
    kpi_value_col_letter = get_column_letter(kpi_col_map['Value'])
    kpi_start_col_letter = get_column_letter(kpi_col_map['Start'])
    kpi_end_col_letter = get_column_letter(kpi_col_map['End'])
    kpi_key_col_letter = get_column_letter(kpi_col_map['Key'])
    kpi_key_range = f'${kpi_key_col_letter}$2:${kpi_key_col_letter}${kpi_end_row}'

    for label_offset, label in enumerate(summary_labels):
        current_row = summary_first_row + label_offset
        if label == 'OVS SAP STAGES':
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2 + len(timeline_days))
            ovs_cell = worksheet.cell(row=current_row, column=1, value=label)
            ovs_cell.font = Font(bold=True, color='102845')
            ovs_cell.fill = info_fill
            ovs_cell.border = border
            ovs_cell.alignment = Alignment(horizontal='left')
            for merged_col in range(1, 3 + len(timeline_days)):
                worksheet.cell(row=current_row, column=merged_col).border = border
                worksheet.cell(row=current_row, column=merged_col).fill = info_fill
            worksheet[f'{summary_helper_letters[0]}{current_row}'] = ''
            worksheet[f'{summary_helper_letters[1]}{current_row}'] = ''
            continue

        label_cell = worksheet.cell(row=current_row, column=1, value=label)
        label_cell.font = bold
        label_cell.border = border
        label_cell.alignment = Alignment(horizontal='left')

        value_cell = worksheet.cell(
            row=current_row,
            column=2,
            value=lookup_formula(f'${kpi_value_col_letter}$2:${kpi_value_col_letter}${kpi_end_row}', kpi_key_range, f'$B$1&"|"&$A{current_row}', '"-"'),
        )
        value_cell.border = border
        value_cell.alignment = Alignment(horizontal='center')

        worksheet[f'{summary_helper_letters[0]}{current_row}'] = lookup_formula(
            f'${kpi_start_col_letter}$2:${kpi_start_col_letter}${kpi_end_row}',
            kpi_key_range,
            f'$B$1&"|"&$A{current_row}',
            '""',
        )
        worksheet[f'{summary_helper_letters[1]}{current_row}'] = lookup_formula(
            f'${kpi_end_col_letter}$2:${kpi_end_col_letter}${kpi_end_row}',
            kpi_key_range,
            f'$B$1&"|"&$A{current_row}',
            '""',
        )

        for day_offset in range(len(timeline_days)):
            column_index = 3 + day_offset
            worksheet.cell(row=current_row, column=column_index, value='').border = border

    summary_range = f'C{summary_first_row}:{get_column_letter(2 + len(timeline_days))}{summary_last_row}'
    worksheet.conditional_formatting.add(
        summary_range,
        FormulaRule(formula=[f'AND($A{summary_first_row}="POL>POD",${summary_helper_letters[0]}{summary_first_row}<>"",COLUMN()-2>=${summary_helper_letters[0]}{summary_first_row},COLUMN()-2<=${summary_helper_letters[1]}{summary_first_row})'], fill=blue_fill),
    )
    worksheet.conditional_formatting.add(
        summary_range,
        FormulaRule(formula=[f'AND($A{summary_first_row}<>"POL>POD",$A{summary_first_row}<>"OVS SAP STAGES",${summary_helper_letters[0]}{summary_first_row}<>"",COLUMN()-2>=${summary_helper_letters[0]}{summary_first_row},COLUMN()-2<=${summary_helper_letters[1]}{summary_first_row})'], fill=green_fill),
    )

    safety_label = worksheet.cell(row=safety_stock_row, column=1, value='Customer Safety STOCK')
    safety_label.font = bold
    safety_value_cell = worksheet.cell(row=safety_stock_row, column=2)
    safety_value_cell.value = lookup_formula(
        f'${get_column_letter(metadata_col_map["Safety stock"])}$2:${get_column_letter(metadata_col_map["Safety stock"])}${metadata_end_row}',
        f'${get_column_letter(metadata_col_map["ID"])}$2:${get_column_letter(metadata_col_map["ID"])}${metadata_end_row}',
        '$B$1',
        '""',
    )

    raw_worksheet.cell(row=raw_title_row, column=1, value='Raw data (filter by ID, POL or POD)').font = section_font
    raw_headers = ['ID', 'POL', 'POD', 'Step', 'Day', 'Day+', 'Final Day', 'Green Start', 'Green End', 'Blue Start', 'Blue End', 'Light Blue Start', 'Light Blue End', 'Paint Duration', 'Key']
    for column_index, header in enumerate(raw_headers, start=1):
        cell = raw_worksheet.cell(row=raw_header_row, column=column_index, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
    for row_index, values in enumerate(step_rows, start=raw_data_start_row):
        for column_index, value in enumerate(values, start=1):
            cell = raw_worksheet.cell(row=row_index, column=column_index, value=value)
            cell.border = border

    if step_rows:
        table = Table(displayName='DynamicStepData', ref=f'A{raw_header_row}:O{raw_data_end_row}')
        table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        raw_worksheet.add_table(table)

    worksheet.freeze_panes = 'A5'
    raw_worksheet.freeze_panes = 'A3'
    worksheet.column_dimensions['A'].width = 36
    worksheet.column_dimensions['B'].width = 10
    worksheet.column_dimensions['C'].width = 10
    worksheet.column_dimensions['D'].width = 12
    for column_index in range(timeline_start_col, timeline_end_col + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 4
    worksheet.column_dimensions['N'].width = 14
    worksheet.column_dimensions['O'].width = 28
    for raw_col in ['A', 'B', 'C', 'D']:
        raw_worksheet.column_dimensions[raw_col].width = 18
    for raw_col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        raw_worksheet.column_dimensions[raw_col].width = 14
    raw_worksheet.column_dimensions['O'].width = 28

    for helper_col in range(timeline_helper_start_col, timeline_helper_start_col + 8):
        worksheet.column_dimensions[get_column_letter(helper_col)].hidden = True
    for metadata_col in range(metadata_start_col, metadata_start_col + 16):
        worksheet.column_dimensions[get_column_letter(metadata_col)].hidden = True
    for hidden_col in ['H', 'I', 'J', 'K', 'L', 'M']:
        worksheet.column_dimensions[hidden_col].width = 12

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def run():
    st.set_page_config(page_title='ALL_VTT', layout='wide')
    st.title('ALL_VTT')
    st.caption('Descarga masiva de dashboards VTT en Excel. Un ID por hoja y solo formato tabla.')

    data_path = os.path.join(os.path.dirname(__file__), 'VTT DATA.xlsx')
    if not os.path.exists(data_path):
        st.error(f'No se encontró el archivo fuente: {data_path}')
        return

    try:
        df_vtt = pd.read_excel(data_path)
    except Exception as exc:
        st.error(f'No se pudo leer el archivo VTT DATA.xlsx: {exc}')
        st.exception(exc)
        return

    if df_vtt.empty:
        st.warning('El archivo VTT DATA.xlsx no tiene registros para exportar.')
        return

    today = datetime.today()
    start_date = today - timedelta(days=today.weekday())
    num_days = 110
    timeline_days = [start_date + timedelta(days=index) for index in range(num_days)]

    id_count = df_vtt['ID'].astype(str).str.strip().nunique() if 'ID' in df_vtt.columns else len(df_vtt)
    col1, col2, col3 = st.columns(3)
    col1.metric('Rows', len(df_vtt))
    col2.metric('IDs', id_count)
    col3.metric('Sheets to export', len(df_vtt))

    preview_columns = [column for column in ['ID', 'POL', 'POD', 'Carrier', 'Name Destin Site'] if column in df_vtt.columns]
    if preview_columns:
        st.dataframe(df_vtt[preview_columns], use_container_width=True, hide_index=True)

    action_col_1, action_col_2 = st.columns(2)

    with action_col_1:
        if st.button('Generate ALL_VTT Excel', type='primary', use_container_width=True):
            with st.spinner('Generando workbook masivo...'):
                excel_bytes = build_all_vtt_workbook(df_vtt, timeline_days)
            download_date = datetime.now().strftime('%Y-%m-%d')
            st.download_button(
                'Download ALL_VTT.xlsx',
                data=excel_bytes,
                file_name=f'HORSE_ALL_VTT_{download_date}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )

    with action_col_2:
        if st.button('Generate Dynamic Excel (1 sheet)', use_container_width=True):
            with st.spinner('Generando workbook dinamico de una sola hoja...'):
                dynamic_excel_bytes = build_dynamic_single_sheet_workbook(df_vtt, timeline_days)
            download_date = datetime.now().strftime('%Y-%m-%d')
            st.download_button(
                'Download Dynamic ALL_VTT.xlsx',
                data=dynamic_excel_bytes,
                file_name=f'HORSE_ALL_VTT_DYNAMIC_{download_date}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )


if __name__ == '__main__':
    run()